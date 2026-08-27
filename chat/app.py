#!/usr/bin/env python3
"""Aperyn's authenticated Ollama control plane and local-model workspaces."""

import base64
import glob
import hashlib
import hmac
import io
import json
import math
import os
import queue
import re
import secrets
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import uuid
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone, timedelta
from pathlib import Path
from urllib.parse import quote

import pandas as pd
import requests
from bs4 import BeautifulSoup
from docx import Document
from flask import Flask, Response, after_this_request, g, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import load_workbook
from pypdf import PdfReader
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
OLLAMA_API = os.environ.get('OLLAMA_API', 'http://ollama:11434').rstrip('/')
MODELS_DIR = os.environ.get('MODELS_DIR', '/models')
DATABASE_PATH = os.environ.get('DATABASE_PATH', '/data/ollama-manager.db')
PROXY_INTERNAL_URL = os.environ.get('PROXY_INTERNAL_URL', 'http://127.0.0.1:11435').rstrip('/')
PROXY_PUBLIC_PORT = int(os.environ.get('PROXY_PUBLIC_PORT', '11435'))
MANAGER_LISTEN_PORT = int(os.environ.get('MANAGER_LISTEN_PORT', '3000'))
APP_VERSION = '1.27.11'
HELPER_TOKEN = os.environ.get('OLLAMA_CONTROL_HELPER_TOKEN', '')

# The session key is generated once in persistent application data. It is never
# embedded in the image or release archive, so sessions survive image upgrades.
SESSION_SECRET_PATH = Path(os.environ.get('SESSION_SECRET_PATH', str(Path(DATABASE_PATH).parent / 'session.secret')))
SESSION_SECRET_PATH.parent.mkdir(parents=True, exist_ok=True)
if not SESSION_SECRET_PATH.exists():
    SESSION_SECRET_PATH.write_text(secrets.token_urlsafe(64), encoding='utf-8')
    os.chmod(SESSION_SECRET_PATH, 0o600)
app.secret_key = SESSION_SECRET_PATH.read_text(encoding='utf-8').strip()


def _telemetry_clear_identity():
    """Derive a scoped proxy credential without transmitting the session secret."""
    return hmac.new(
        str(app.secret_key).encode('utf-8'),
        b'aperyn-telemetry-clear-v1',
        hashlib.sha256,
    ).hexdigest()


app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    SESSION_COOKIE_SECURE=os.environ.get('SESSION_COOKIE_SECURE', '').lower() in ('1', 'true', 'yes'),
    PERMANENT_SESSION_LIFETIME=timedelta(days=14),
)

@app.after_request
def _control_cache_headers(response):
    # HTML/API changes should show immediately after a container rebuild; versioned
    # static asset URLs can still be cached normally. This is especially important
    # for Mobile Safari, which otherwise tends to keep older dashboard JS/CSS.
    if request.path.startswith('/api/') or response.mimetype == 'text/html':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
    return response

# Converter configuration
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB max file size
app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', '/tmp/uploads')
app.config['OUTPUT_FOLDER'] = os.environ.get('OUTPUT_FOLDER', '/data/training')

TABLE_UPLOAD_SUFFIXES = {'.xlsx', '.xls', '.csv'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

def _table_upload():
    upload = request.files.get('file')
    if upload is None:
        raise ValueError('No file uploaded')
    original = str(upload.filename or '').strip()
    if not original:
        raise ValueError('No file selected')
    safe_name = secure_filename(original)
    suffix = Path(safe_name).suffix.lower()
    if suffix not in TABLE_UPLOAD_SUFFIXES:
        raise ValueError('Invalid file type. Use .xlsx, .xls, or .csv')
    temporary = tempfile.NamedTemporaryFile(prefix='aperyn-table-', suffix=suffix, dir=app.config['UPLOAD_FOLDER'], delete=False)
    temporary.close()
    upload.save(temporary.name)
    return Path(temporary.name), safe_name


def _read_table(path, row_limit=None):
    options = {'nrows': row_limit} if row_limit else {}
    return pd.read_csv(path, **options) if path.suffix.lower() == '.csv' else pd.read_excel(path, **options)


def _discard(path):
    if path:
        try:
            Path(path).unlink(missing_ok=True)
        except OSError:
            app.logger.warning('Could not remove temporary table file %s', path)

@app.route('/converter')
def converter_page():
    return render_template('converter.html')

@app.route('/library')
@app.route('/wizard')  # backwards-compatible URL from older releases
def wizard_page():
    return render_template('wizard.html')

@app.route('/settings')
def settings_page():
    return render_template('settings.html')

@app.post('/api/converter/convert')
def convert_dataset():
    source_path = jsonl_path = None
    download_scheduled = False
    try:
        source_path, original_name = _table_upload()
        table = _read_table(source_path)
        prompt_column = str(request.form.get('instruction_col') or 'instruction')
        answer_column = str(request.form.get('output_col') or 'output')
        missing = [name for name in (prompt_column, answer_column) if name not in table.columns]
        if missing:
            return jsonify({'error': f'Column "{missing[0]}" not found in file'}), 400

        download_name = f'{Path(original_name).stem}.jsonl'
        handle = tempfile.NamedTemporaryFile(mode='w', suffix='.jsonl', delete=False, encoding='utf-8')
        jsonl_path = Path(handle.name)
        converted = 0
        with handle:
            for prompt_value, answer_value in table[[prompt_column, answer_column]].itertuples(index=False, name=None):
                if pd.isna(prompt_value) or pd.isna(answer_value):
                    continue
                handle.write(json.dumps({'instruction': str(prompt_value).strip(), 'output': str(answer_value).strip()}, ensure_ascii=False) + '\n')
                converted += 1

        _discard(source_path)
        source_path = None
        if str(request.form.get('save_to_server') or '').lower() == 'true':
            destination = Path(app.config['OUTPUT_FOLDER']) / download_name
            os.replace(jsonl_path, destination)
            jsonl_path = None
            return jsonify({'success': True, 'message': f'Converted and saved to {download_name}', 'output_file': download_name, 'rows_converted': converted})

        cleanup_target = jsonl_path
        @after_this_request
        def _remove_download(response):
            _discard(cleanup_target)
            return response
        download_scheduled = True
        return send_file(jsonl_path, as_attachment=True, download_name=download_name, mimetype='application/x-ndjson')
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        app.logger.exception('Dataset conversion failed')
        return jsonify({'error': f'Could not convert dataset: {exc}'}), 500
    finally:
        _discard(source_path)
        if jsonl_path and not download_scheduled:
            _discard(jsonl_path)

@app.post('/api/converter/preview')
def preview_dataset():
    source_path = None
    try:
        source_path, _ = _table_upload()
        sample = _read_table(source_path, row_limit=5).where(lambda frame: frame.notna(), None)
        return jsonify({'columns': [str(column) for column in sample.columns], 'preview': sample.to_dict(orient='records')})
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception as exc:
        app.logger.exception('Dataset preview failed')
        return jsonify({'error': f'Could not preview dataset: {exc}'}), 500
    finally:
        _discard(source_path)

_WIZARD_PARAMETER_SETS = {
    'chatbot': (('temperature', .7), ('num_ctx', 4096), ('top_p', .9)),
    'code': (('temperature', .3), ('num_ctx', 8192), ('top_p', .9), ('repeat_penalty', 1.1)),
    'support': (('temperature', .3), ('num_ctx', 4096), ('top_p', .9)),
    'creative': (('temperature', 1.2), ('num_ctx', 8192), ('top_p', .95)),
    'translator': (('temperature', .2), ('num_ctx', 4096), ('top_p', .9)),
    'data': (('temperature', .1), ('num_ctx', 2048), ('top_p', .9)),
}


@app.post('/api/wizard/generate')
def wizard_modelfile_api():
    payload = request.get_json(silent=True) or {}
    source_model = str(payload.get('base_model') or 'llama3.2:1b').strip()
    if not source_model or any(character in source_model for character in '\r\n'):
        return jsonify({'error': 'Choose a valid base model'}), 400
    preset = _WIZARD_PARAMETER_SETS.get(str(payload.get('use_case') or 'chatbot'), _WIZARD_PARAMETER_SETS['chatbot'])
    sections = [f'FROM {source_model}', '\n'.join(f'PARAMETER {key} {value}' for key, value in preset)]

    persona = str(payload.get('personality') or '').strip()
    rules = [str(value).strip() for value in (payload.get('rules') or []) if str(value).strip()]
    system_lines = ([persona] if persona else []) + ((['Rules:'] + [f'- {rule}' for rule in rules]) if rules else [])
    if system_lines:
        sections.append('SYSTEM """\n' + '\n\n'.join(system_lines[:1]) + ('\n' if len(system_lines) > 1 else '') + '\n'.join(system_lines[1:]) + '\n"""')

    dialogue = []
    for example in payload.get('examples') or []:
        if not isinstance(example, dict):
            continue
        question = str(example.get('question') or '').strip().replace('"', '\\"')
        answer = str(example.get('answer') or '').strip().replace('"', '\\"')
        if question and answer:
            dialogue.extend((f'MESSAGE user "{question}"', f'MESSAGE assistant "{answer}"'))
    if dialogue:
        sections.append('\n'.join(dialogue))
    return jsonify({'content': '\n\n'.join(sections)})

@app.route('/')
def index():
    # Chat is the default product landing page; Dashboard remains available at /dashboard.
    return redirect('/chat')

@app.route('/dashboard')
def dashboard_page():
    return render_template('manager.html')

@app.route('/chat')
def legacy_chat():
    return render_template('chat.html')

@app.route('/api/models')
def installed_models_api():
    try:
        response = requests.get(f'{OLLAMA_API}/api/tags', timeout=(3, 20))
        response.raise_for_status()
        payload = response.json()
        return jsonify(payload if isinstance(payload, dict) else {'models': []})
    except (requests.RequestException, ValueError) as exc:
        app.logger.warning('Installed model lookup failed: %s', exc)
        return jsonify({'error': 'Installed models are unavailable'}), 503


def _models_root():
    return Path(MODELS_DIR).resolve()


def _modelfile_path(candidate, require_existing=True, custom_only=False):
    root = ((_models_root() / 'custom') if custom_only else _models_root()).resolve()
    path = Path(candidate).resolve()
    if root != path and root not in path.parents:
        raise ValueError('Modelfile must be inside the custom models directory')
    if path.name != 'Modelfile':
        raise ValueError('Expected a Modelfile path')
    if require_existing and not path.is_file():
        raise FileNotFoundError('Modelfile not found')
    return path


def _custom_modelfile(candidate, require_existing=True):
    return _modelfile_path(candidate, require_existing=require_existing, custom_only=True)

@app.route('/api/modelfiles')
def local_modelfiles_api():
    try:
        root = _models_root()
        entries = [{'path': str(path), 'name': str(path.relative_to(root))} for path in root.rglob('Modelfile') if path.is_file()]
        return jsonify({'modelfiles': sorted(entries, key=lambda item: item['name'].casefold())})
    except OSError as exc:
        app.logger.warning('Could not enumerate Modelfiles: %s', exc)
        return jsonify({'error': 'Could not enumerate local Modelfiles'}), 500

@app.post('/api/pull-model')
def pull_installed_model():
    model_name = str((request.get_json(silent=True) or {}).get('name') or '').strip()
    if not model_name:
        return jsonify({'error': 'Missing model name'}), 400

    def stream_pull():
        try:
            with requests.post(f'{OLLAMA_API}/api/pull', json={'model': model_name}, stream=True, timeout=(5, 3600)) as upstream:
                if not upstream.ok:
                    yield json.dumps({'error': upstream.text or f'Ollama returned HTTP {upstream.status_code}'}) + '\n'
                    return
                for raw in upstream.iter_lines():
                    if raw:
                        yield raw.decode('utf-8', errors='replace') + '\n'
            _record_model_source(model_name, 'ollama', model_name, '', {'managed_by': 'direct_pull'})
        except requests.RequestException as exc:
            yield json.dumps({'error': f'Ollama pull failed: {exc}'}) + '\n'

    return Response(stream_pull(), mimetype='application/x-ndjson')

def _parse_modelfile_for_api(content, model_name):
    """Translate the local wizard/basic Modelfile syntax into /api/create JSON.

    This keeps host-service deployments independent of `docker exec ollama`.
    File-based FROM/ADAPTER imports are intentionally rejected here because a
    path inside the WebUI container is not a path visible to host Ollama.
    """
    req = {'model': model_name, 'parameters': {}, 'messages': [], 'stream': True}
    lines = (content or '').splitlines()
    i = 0

    def read_value(first):
        nonlocal i
        value = first.strip()
        if not value.startswith('"""'):
            return value.strip().strip('"')
        value = value[3:]
        if value.endswith('"""'):
            return value[:-3]
        parts = [value] if value else []
        while i < len(lines):
            part = lines[i]
            i += 1
            if '"""' in part:
                before, _, _ = part.partition('"""')
                parts.append(before)
                break
            parts.append(part)
        return '\n'.join(parts).strip('\n')

    while i < len(lines):
        raw = lines[i]
        i += 1
        line = raw.strip()
        if not line or line.startswith('#'):
            continue
        pieces = line.split(None, 1)
        command = pieces[0].upper()
        arg = pieces[1] if len(pieces) > 1 else ''

        if command == 'FROM':
            source = read_value(arg).strip()
            if not source:
                raise ValueError('Modelfile FROM is empty')
            if source.startswith(('/', './', '../')):
                raise ValueError('Host Ollama cannot use a file path inside the WebUI container. Import the file into Ollama first, then use the installed model name in FROM.')
            req['from'] = source
        elif command == 'PARAMETER':
            param_parts = arg.split(None, 1)
            if len(param_parts) != 2:
                raise ValueError(f'Invalid PARAMETER line: {line}')
            key, value = param_parts
            parsed = _parse_scalar(value)
            if key in req['parameters']:
                current = req['parameters'][key]
                req['parameters'][key] = current + [parsed] if isinstance(current, list) else [current, parsed]
            else:
                req['parameters'][key] = parsed
        elif command == 'SYSTEM':
            req['system'] = read_value(arg)
        elif command == 'TEMPLATE':
            req['template'] = read_value(arg)
        elif command == 'LICENSE':
            req['license'] = read_value(arg)
        elif command == 'MESSAGE':
            msg_parts = arg.split(None, 1)
            if len(msg_parts) != 2:
                raise ValueError(f'Invalid MESSAGE line: {line}')
            role, body = msg_parts
            req['messages'].append({'role': role, 'content': read_value(body)})
        elif command == 'ADAPTER':
            raise ValueError('ADAPTER file paths are not supported by the host-service wizard path. Import the adapter/model into Ollama first or use the dedicated import workflow.')
        else:
            raise ValueError(f'Unsupported Modelfile command in WebUI API mode: {command}')

    if not req.get('from'):
        raise ValueError('Modelfile requires FROM <installed-model>')
    if not req['parameters']:
        req.pop('parameters')
    if not req['messages']:
        req.pop('messages')
    return req


@app.route('/api/create-model', methods=['POST'])
def create_model():
    try:
        submitted = request.get_json(silent=True) or {}
        model_name = str(submitted.get('name') or '').strip()
        modelfile_path = str(submitted.get('path') or '').strip()
        if not (model_name and modelfile_path):
            return jsonify({'error': 'Missing name or path'}), 400
        try:
            content = _modelfile_path(modelfile_path).read_text(encoding='utf-8')
        except FileNotFoundError:
            return jsonify({'error': 'Modelfile not found'}), 404
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        try:
            create_payload = _parse_modelfile_for_api(content, model_name)
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400

        def generate():
            try:
                yield json.dumps({'status': f'Creating model {model_name} through the Ollama API...'}) + '\n'
                response = requests.post(
                    f'{OLLAMA_API}/api/create',
                    json=create_payload,
                    stream=True,
                    timeout=3600,
                )
                if not response.ok:
                    yield json.dumps({'error': response.text or f'Ollama returned HTTP {response.status_code}'}) + '\n'
                    return
                for line in response.iter_lines():
                    if not line:
                        continue
                    try:
                        obj = json.loads(line)
                    except Exception:
                        obj = {'status': line.decode('utf-8', errors='replace')}
                    yield json.dumps(obj) + '\n'
            except Exception as exc:
                yield json.dumps({'error': str(exc)}) + '\n'

        return Response(generate(), mimetype='application/x-ndjson')
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.get('/api/modelfile')
def read_custom_modelfile():
    try:
        path = _modelfile_path(request.args.get('path') or '')
        return jsonify({'content': path.read_text(encoding='utf-8'), 'path': str(path)})
    except FileNotFoundError:
        return jsonify({'error': 'Modelfile not found'}), 404
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except OSError:
        return jsonify({'error': 'Could not read Modelfile'}), 500

@app.post('/api/modelfile')
def create_custom_modelfile():
    try:
        payload = request.get_json(silent=True) or {}
        raw_name = str(payload.get('name') or '').strip()
        if not raw_name:
            return jsonify({'error': 'Missing name'}), 400
        directory_name = secure_filename(raw_name)
        if not directory_name:
            return jsonify({'error': 'Model name contains no usable characters'}), 400
        path = _custom_modelfile(_models_root() / 'custom' / directory_name / 'Modelfile', require_existing=False)
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            with path.open('x', encoding='utf-8') as handle:
                handle.write(str(payload.get('content') or ''))
        except FileExistsError:
            return jsonify({'error': 'Modelfile already exists'}), 409
        return jsonify({'success': True, 'path': str(path)})
    except (OSError, ValueError) as exc:
        app.logger.warning('Could not create custom Modelfile: %s', exc)
        return jsonify({'error': str(exc)}), 500

@app.put('/api/modelfile')
def update_custom_modelfile():
    try:
        payload = request.get_json(silent=True) or {}
        if not payload.get('path') or payload.get('content') is None:
            return jsonify({'error': 'Missing path or content'}), 400
        path = _custom_modelfile(payload['path'])
        path.write_text(str(payload['content']), encoding='utf-8')
        return jsonify({'success': True})
    except FileNotFoundError:
        return jsonify({'error': 'Modelfile not found'}), 404
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except OSError:
        return jsonify({'error': 'Could not update Modelfile'}), 500

@app.delete('/api/modelfile')
def delete_custom_modelfile():
    try:
        path = _custom_modelfile(request.args.get('path') or '')
        path.unlink()
        try:
            path.parent.rmdir()
        except OSError:
            pass
        return jsonify({'success': True})
    except FileNotFoundError:
        return jsonify({'error': 'Modelfile not found'}), 404
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except OSError:
        return jsonify({'error': 'Could not delete Modelfile'}), 500

@app.delete('/api/delete-model')
def remove_installed_model():
    model_name = str(request.args.get('name') or '').strip()
    if not model_name:
        return jsonify({'error': 'Missing model name'}), 400
    try:
        upstream = requests.delete(f'{OLLAMA_API}/api/delete', json={'model': model_name}, timeout=(3, 120))
    except requests.RequestException as exc:
        app.logger.warning('Ollama model removal failed: %s', exc)
        return jsonify({'error': 'Ollama is unavailable'}), 503
    if not upstream.ok:
        return jsonify({'error': upstream.text or 'Failed to delete model'}), upstream.status_code
    _delete_model_pref(model_name)
    _set_setting(_thinking_override_key(model_name), '')
    with _db() as conn:
        conn.execute('DELETE FROM model_sources WHERE model=?', (model_name,))
    return jsonify({'success': True})

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        data = request.json or {}
        model = data.get('model', 'llama3.2:1b')
        message = data.get('message', '')
        mtp_pref = _model_pref(model)
        options = {}
        if mtp_pref:
            options['draft_num_predict'] = mtp_pref['mtp_draft_n_max'] if mtp_pref['mtp_enabled'] else 0

        client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        start = time.perf_counter()
        live_id = _start_live_generation(model, '/api/chat', mtp_pref)

        # Stream response from Ollama while feeding the same live telemetry used
        # by the observable proxy, so the built-in Chat page appears on Dashboard.
        def generate():
            final = {}
            stream_error = None
            status_code = 500
            response = None
            try:
                payload = {'model': model, 'prompt': message, 'stream': True}
                if options:
                    payload['options'] = options
                response = requests.post(f'{OLLAMA_API}/api/generate', json=payload, stream=True, timeout=3600)
                status_code = response.status_code
                for line in response.iter_lines():
                    if line:
                        try:
                            obj = json.loads(line)
                            _update_live_generation(live_id, obj)
                            if obj.get('done'):
                                final.update(obj)
                        except Exception:
                            pass
                        yield line.decode('utf-8') + '\n'
            except Exception as e:
                stream_error = e
                yield json.dumps({'error': str(e)}) + '\n'
            finally:
                latency_ms = (time.perf_counter() - start) * 1000
                try:
                    _log_request('/api/chat', model, status_code, latency_ms, final, client_ip)
                finally:
                    _finish_live_generation(live_id, status_code, final, stream_error)
                    if response is not None:
                        response.close()

        return Response(generate(), mimetype='text/event-stream')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# --- Aperyn manager -------------------------------------------------

PARAMETER_DEFINITIONS = [
    {"name":"num_ctx","group":"Context & runtime","type":"int","type_label":"integer · tokens","default":4096,"min":512,"max":1048576,"step":512,"description":"Context window size. Higher values increase memory use."},
    {"name":"num_predict","group":"Generation","type":"int","type_label":"integer · tokens","default":-1,"min":-2,"max":1048576,"step":1,"description":"Maximum number of tokens to generate. -1 allows generation until a stop condition."},
    {"name":"draft_num_predict","group":"Generation","type":"int","type_label":"integer · speculative tokens","default":2,"min":0,"max":64,"step":1,"editor_hidden":True,"description":"Speculative draft depth. Managed by the dedicated MTP control when editing a model."},
    {"name":"temperature","group":"Sampling","type":"float","type_label":"decimal","default":0.8,"min":0,"max":2,"step":0.05,"slider":True,"description":"Randomness of generation. Lower is more focused; higher is more varied."},
    {"name":"top_k","group":"Sampling","type":"int","type_label":"integer","default":40,"min":0,"max":1000,"step":1,"slider":True,"description":"Restrict sampling to the K most likely next tokens."},
    {"name":"top_p","group":"Sampling","type":"float","type_label":"decimal","default":0.9,"min":0,"max":1,"step":0.01,"slider":True,"description":"Nucleus sampling probability threshold."},
    {"name":"min_p","group":"Sampling","type":"float","type_label":"decimal","default":0.0,"min":0,"max":1,"step":0.01,"slider":True,"description":"Minimum probability threshold relative to the most likely token."},
    {"name":"typical_p","group":"Sampling","type":"float","type_label":"decimal · advanced","default":1.0,"min":0,"max":1,"step":0.01,"slider":True,"description":"Locally typical sampling threshold."},
    {"name":"repeat_last_n","group":"Penalties","type":"int","type_label":"integer","default":64,"min":-1,"max":32768,"step":1,"description":"How far back to look when applying repetition penalties. -1 uses num_ctx."},
    {"name":"repeat_penalty","group":"Penalties","type":"float","type_label":"decimal","default":1.0,"min":0,"max":2,"step":0.05,"slider":True,"description":"Penalty applied to repeated tokens. 1.0 disables the penalty."},
    {"name":"presence_penalty","group":"Penalties","type":"float","type_label":"decimal · advanced","default":0.0,"min":-2,"max":2,"step":0.05,"slider":True,"description":"Penalises tokens based on whether they have already appeared."},
    {"name":"frequency_penalty","group":"Penalties","type":"float","type_label":"decimal · advanced","default":0.0,"min":-2,"max":2,"step":0.05,"slider":True,"description":"Penalises tokens based on how often they have appeared."},
    {"name":"penalize_newline","group":"Penalties","type":"bool","type_label":"boolean · advanced","default":False,"description":"Apply the repetition penalty to newline tokens as well."},
    {"name":"seed","group":"Generation","type":"int","type_label":"integer","default":0,"min":0,"max":2147483647,"step":1,"description":"Random seed. A fixed value makes comparable prompts more reproducible."},
    {"name":"stop","group":"Generation","type":"string-list","type_label":"text · comma-separated","default":"","description":"One or more stop strings. Separate multiple values with commas."},
    {"name":"num_keep","group":"Context & runtime","type":"int","type_label":"integer · advanced","default":4,"min":0,"max":262144,"step":1,"description":"Number of prompt tokens to retain when context shifting occurs."},
    {"name":"numa","group":"Hardware & loading","type":"bool","type_label":"boolean · advanced","default":False,"description":"Enable NUMA-aware runner behaviour on multi-socket systems."},
    {"name":"num_batch","group":"Hardware & loading","type":"int","type_label":"integer · advanced","default":512,"min":1,"max":8192,"step":1,"slider":True,"description":"Prompt processing batch size. Larger values can improve speed but use more memory."},
    {"name":"num_gpu","group":"Hardware & loading","type":"int","type_label":"integer · GPU layers","default":-1,"min":-1,"max":999,"step":1,"slider":True,"model_limited":True,"description":"Number of model layers to offload to GPU. -1 lets Ollama choose automatically; the editor clamps the upper bound to this model’s offloadable layers."},
    {"name":"main_gpu","group":"Hardware & loading","type":"int","type_label":"integer · GPU index","default":0,"min":0,"max":31,"step":1,"slider":True,"description":"Primary GPU index for model loading on multi-GPU systems."},
    {"name":"use_mmap","group":"Hardware & loading","type":"bool","type_label":"boolean · advanced","default":True,"description":"Use memory-mapped model files when supported."},
    {"name":"num_thread","group":"Hardware & loading","type":"int","type_label":"integer · CPU threads","default":0,"min":0,"max":512,"step":1,"slider":True,"description":"CPU threads used for computation. 0 leaves selection to Ollama."},
]

_live_lock = threading.Lock()
_active_generations = {}
_recent_generations = deque(maxlen=12)


def _db():
    os.makedirs(os.path.dirname(DATABASE_PATH) or '.', exist_ok=True)
    conn = sqlite3.connect(DATABASE_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA busy_timeout=30000')
    return conn


def _init_db():
    with _db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS request_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at TEXT NOT NULL,
            endpoint TEXT NOT NULL,
            model TEXT,
            status_code INTEGER NOT NULL,
            latency_ms REAL,
            prompt_tokens INTEGER DEFAULT 0,
            output_tokens INTEGER DEFAULT 0,
            eval_duration_ns INTEGER DEFAULT 0,
            client_ip TEXT
        )""")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_request_log_created ON request_log(created_at)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_request_log_model ON request_log(model)')
        conn.execute('PRAGMA journal_mode=WAL')
        conn.execute("""CREATE TABLE IF NOT EXISTS model_preferences (
            model TEXT PRIMARY KEY,
            mtp_enabled INTEGER NOT NULL DEFAULT 0,
            mtp_draft_n_max INTEGER NOT NULL DEFAULT 2,
            updated_at TEXT NOT NULL
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS conversations (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            model TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS conversation_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            conversation_id TEXT NOT NULL,
            role TEXT NOT NULL,
            content TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(conversation_id) REFERENCES conversations(id) ON DELETE CASCADE
        )""")
        # v1.7 chat metadata migrations (safe for existing databases).
        cols = {row[1] for row in conn.execute('PRAGMA table_info(conversation_messages)')}
        if 'thinking' not in cols:
            conn.execute("ALTER TABLE conversation_messages ADD COLUMN thinking TEXT NOT NULL DEFAULT ''")
        if 'attachments_json' not in cols:
            conn.execute("ALTER TABLE conversation_messages ADD COLUMN attachments_json TEXT NOT NULL DEFAULT '[]'")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_conversation_messages_chat ON conversation_messages(conversation_id, id)')
        conn.execute("""CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL
        )""")
        # v1.10 operations/control-plane migrations.
        request_cols = {row[1] for row in conn.execute('PRAGMA table_info(request_log)')}
        if 'client_name' not in request_cols:
            conn.execute("ALTER TABLE request_log ADD COLUMN client_name TEXT NOT NULL DEFAULT ''")
        if 'request_meta_json' not in request_cols:
            conn.execute("ALTER TABLE request_log ADD COLUMN request_meta_json TEXT NOT NULL DEFAULT '{}'")
        conv_cols = {row[1] for row in conn.execute('PRAGMA table_info(conversations)')}
        if 'parent_id' not in conv_cols:
            conn.execute("ALTER TABLE conversations ADD COLUMN parent_id TEXT")
        if 'branched_from_message_id' not in conv_cols:
            conn.execute("ALTER TABLE conversations ADD COLUMN branched_from_message_id INTEGER")
        msg_cols = {row[1] for row in conn.execute('PRAGMA table_info(conversation_messages)')}
        if 'tool_calls_json' not in msg_cols:
            conn.execute("ALTER TABLE conversation_messages ADD COLUMN tool_calls_json TEXT NOT NULL DEFAULT '[]'")
        if 'output_tokens' not in msg_cols:
            conn.execute("ALTER TABLE conversation_messages ADD COLUMN output_tokens INTEGER NOT NULL DEFAULT 0")
        if 'eval_duration_ns' not in msg_cols:
            conn.execute("ALTER TABLE conversation_messages ADD COLUMN eval_duration_ns INTEGER NOT NULL DEFAULT 0")
        conn.execute("""CREATE TABLE IF NOT EXISTS model_sources (
            model TEXT PRIMARY KEY, source TEXT NOT NULL DEFAULT 'unknown', remote_id TEXT NOT NULL DEFAULT '',
            revision TEXT NOT NULL DEFAULT '', extra_json TEXT NOT NULL DEFAULT '{}', updated_at TEXT NOT NULL
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS download_jobs (
            id TEXT PRIMARY KEY, kind TEXT NOT NULL, source TEXT NOT NULL, target_model TEXT NOT NULL,
            payload_json TEXT NOT NULL DEFAULT '{}', phase TEXT NOT NULL DEFAULT 'queued', status TEXT NOT NULL DEFAULT 'Queued',
            completed INTEGER NOT NULL DEFAULT 0, total INTEGER NOT NULL DEFAULT 0, done INTEGER NOT NULL DEFAULT 0,
            success INTEGER, error TEXT NOT NULL DEFAULT '', cancel_requested INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL, updated_at TEXT NOT NULL
        )""")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_download_jobs_created ON download_jobs(created_at)')
        conn.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL COLLATE NOCASE UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user' CHECK(role IN ('admin','user')),
            active INTEGER NOT NULL DEFAULT 1 CHECK(active IN (0,1)),
            must_change_password INTEGER NOT NULL DEFAULT 0 CHECK(must_change_password IN (0,1)),
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            last_login_at TEXT
        )""")
        conn.execute('CREATE UNIQUE INDEX IF NOT EXISTS idx_users_username ON users(username COLLATE NOCASE)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_users_active_role ON users(active, role)')
        conn.execute("""CREATE TABLE IF NOT EXISTS user_preferences (
            user_id INTEGER NOT NULL,
            key TEXT NOT NULL,
            value TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL,
            PRIMARY KEY(user_id, key),
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )""")
        conn.execute("""CREATE TABLE IF NOT EXISTS agent_sessions (
            id TEXT PRIMARY KEY,
            remote_id TEXT NOT NULL UNIQUE,
            user_id INTEGER NOT NULL,
            title TEXT NOT NULL DEFAULT 'New agent task',
            model TEXT NOT NULL DEFAULT '',
            agent_mode TEXT NOT NULL DEFAULT 'build' CHECK(agent_mode IN ('build','plan')),
            approval_mode TEXT NOT NULL DEFAULT 'ask' CHECK(approval_mode IN ('ask','safe','full')),
            status TEXT NOT NULL DEFAULT 'idle',
            workspace TEXT NOT NULL DEFAULT '/workspace',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            last_seen_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
        )""")
        agent_cols = {row[1] for row in conn.execute('PRAGMA table_info(agent_sessions)')}
        if 'approval_mode' not in agent_cols:
            conn.execute("ALTER TABLE agent_sessions ADD COLUMN approval_mode TEXT NOT NULL DEFAULT 'ask'")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_agent_sessions_user_updated ON agent_sessions(user_id, updated_at DESC)')
        # Managed Ollama hosts deliberately keep the remote helper credential on
        # the remote machine.  Aperyn stores only a one-way hash of its connector
        # token and communicates through an outbound, allow-listed connector.
        conn.execute("""CREATE TABLE IF NOT EXISTS managed_hosts (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL COLLATE NOCASE UNIQUE,
            endpoint TEXT NOT NULL DEFAULT '',
            connector_token_hash TEXT NOT NULL DEFAULT '',
            pairing_token_hash TEXT NOT NULL DEFAULT '',
            pairing_expires_at TEXT,
            state TEXT NOT NULL DEFAULT 'not_paired',
            last_seen_at TEXT,
            snapshot_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )""")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_managed_hosts_seen ON managed_hosts(last_seen_at)')
        conn.execute("""CREATE TABLE IF NOT EXISTS host_connector_actions (
            id TEXT PRIMARY KEY,
            host_id TEXT NOT NULL,
            operation TEXT NOT NULL,
            payload_json TEXT NOT NULL DEFAULT '{}',
            status TEXT NOT NULL DEFAULT 'queued' CHECK(status IN ('queued','running','completed','failed','expired')),
            result_json TEXT NOT NULL DEFAULT '{}',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            FOREIGN KEY(host_id) REFERENCES managed_hosts(id) ON DELETE CASCADE
        )""")
        conn.execute('CREATE INDEX IF NOT EXISTS idx_host_connector_actions_poll ON host_connector_actions(host_id,status,created_at)')
        now = datetime.now(timezone.utc).isoformat()
        user_count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        if user_count == 0:
            conn.execute("""INSERT INTO users(username,password_hash,role,active,must_change_password,created_at,updated_at)
                            VALUES(?,?,?,?,?,?,?)""",
                         ('admin', generate_password_hash('password', method='scrypt'), 'admin', 1, 1, now, now))
        # One-time migration: preserve the former global appearance for the
        # primary account, but only when that user has no per-user theme yet.
        primary_user = conn.execute('SELECT id FROM users ORDER BY id LIMIT 1').fetchone()
        if primary_user:
            for key in ('theme_accent','theme_background','theme_panel','theme_panel2','theme_glass'):
                legacy = conn.execute('SELECT value FROM app_settings WHERE key=?', (key,)).fetchone()
                exists = conn.execute('SELECT 1 FROM user_preferences WHERE user_id=? AND key=?', (primary_user[0], key)).fetchone()
                if legacy and legacy[0] and not exists:
                    conn.execute('INSERT INTO user_preferences(user_id,key,value,updated_at) VALUES(?,?,?,?)',
                                 (primary_user[0], key, legacy[0], now))



def _get_setting(key, default=''):
    try:
        with _db() as conn:
            row = conn.execute('SELECT value FROM app_settings WHERE key = ?', (key,)).fetchone()
        return row['value'] if row else default
    except sqlite3.Error:
        return default


def _set_setting(key, value):
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("""INSERT INTO app_settings(key, value, updated_at) VALUES(?,?,?)
                        ON CONFLICT(key) DO UPDATE SET value=excluded.value, updated_at=excluded.updated_at""",
                     (key, str(value or ''), now))


def _get_user_preference(user_id, key, default=''):
    if not user_id:
        return default
    with _db() as conn:
        row = conn.execute('SELECT value FROM user_preferences WHERE user_id=? AND key=?', (user_id, key)).fetchone()
    return row['value'] if row else default


def _set_user_preference(user_id, key, value):
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("""INSERT INTO user_preferences(user_id,key,value,updated_at) VALUES(?,?,?,?)
                        ON CONFLICT(user_id,key) DO UPDATE SET value=excluded.value,updated_at=excluded.updated_at""",
                     (user_id, key, str(value or ''), now))


_login_attempts = {}
_login_attempts_lock = threading.Lock()


def _csrf_token():
    if not session.get('csrf_token'):
        session['csrf_token'] = secrets.token_urlsafe(32)
    return session['csrf_token']


def _current_user(refresh=False):
    user_id = session.get('user_id')
    if not user_id:
        return None
    if not refresh and getattr(g, 'current_user', None):
        return g.current_user
    with _db() as conn:
        row = conn.execute('SELECT id,username,role,active,must_change_password,created_at,last_login_at FROM users WHERE id=?', (user_id,)).fetchone()
    user = dict(row) if row else None
    if not user or not user['active']:
        session.clear()
        return None
    g.current_user = user
    return user


def _auth_json_or_redirect(status=401, message='Authentication required'):
    if request.path.startswith('/api/'):
        return jsonify({'error': message}), status
    return redirect(url_for('login', next=request.full_path if request.query_string else request.path))


@app.context_processor
def _auth_template_context():
    return {'current_user': _current_user(), 'csrf_token': _csrf_token}


@app.before_request
def _authentication_and_csrf():
    if request.is_secure:
        app.config['SESSION_COOKIE_SECURE'] = True
    endpoint = request.endpoint or ''
    public = endpoint in {'login', 'static', 'health', 'service_worker', 'host_connector_register', 'host_connector_poll', 'host_connector_action_result'} or request.path.startswith('/ollama/')
    if not public:
        user = _current_user()
        if not user:
            return _auth_json_or_redirect()
        if user['must_change_password'] and endpoint not in {'change_password', 'logout', 'auth_session'}:
            return _auth_json_or_redirect(403, 'Password change required') if request.path.startswith('/api/') else redirect(url_for('change_password'))
    if request.method in {'POST','PUT','PATCH','DELETE'} and endpoint not in {'login', 'host_connector_register', 'host_connector_poll', 'host_connector_action_result'} and not request.path.startswith('/ollama/'):
        supplied = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token')
        expected = session.get('csrf_token')
        if not expected or not supplied or not secrets.compare_digest(str(supplied), str(expected)):
            return jsonify({'error': 'Invalid or missing CSRF token'}), 400 if request.path.startswith('/api/') else 403


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        if _current_user():
            return redirect('/chat')
        return render_template('login.html', error=None)
    username = str(request.form.get('username') or '').strip()
    password = str(request.form.get('password') or '')
    key = (request.remote_addr or 'unknown', username.casefold())
    now_mono = time.monotonic()
    with _login_attempts_lock:
        attempts = [t for t in _login_attempts.get(key, []) if now_mono - t < 900]
        _login_attempts[key] = attempts
    if len(attempts) >= 5:
        retry = min(300, 2 ** min(len(attempts) - 4, 8))
        newest = attempts[-1]
        if now_mono - newest < retry:
            return render_template('login.html', error=f'Too many attempts. Try again in {int(retry - (now_mono-newest)) + 1} seconds.'), 429
    with _db() as conn:
        row = conn.execute('SELECT * FROM users WHERE username=? COLLATE NOCASE', (username,)).fetchone()
        valid = bool(row and row['active'] and check_password_hash(row['password_hash'], password))
        if valid:
            now = datetime.now(timezone.utc).isoformat()
            conn.execute('UPDATE users SET last_login_at=?,updated_at=? WHERE id=?', (now, now, row['id']))
    if not valid:
        with _login_attempts_lock:
            _login_attempts.setdefault(key, []).append(time.monotonic())
        time.sleep(min(0.6, 0.12 * (len(attempts) + 1)))
        return render_template('login.html', error='Invalid username or password.'), 401
    with _login_attempts_lock:
        _login_attempts.pop(key, None)
    session.clear()
    session.permanent = True
    session['user_id'] = row['id']
    session['csrf_token'] = secrets.token_urlsafe(32)
    next_path = str(request.args.get('next') or '/chat')
    if not next_path.startswith('/') or next_path.startswith('//'):
        next_path = '/chat'
    return redirect(url_for('change_password') if row['must_change_password'] else next_path)


@app.post('/logout')
def logout():
    session.clear()
    return redirect('/login')


@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    user = _current_user()
    error = None
    if request.method == 'POST':
        current = str(request.form.get('current_password') or '')
        new = str(request.form.get('new_password') or '')
        confirm = str(request.form.get('confirm_password') or '')
        with _db() as conn:
            row = conn.execute('SELECT password_hash FROM users WHERE id=?', (user['id'],)).fetchone()
            if not row or not check_password_hash(row['password_hash'], current):
                error = 'Current password is incorrect.'
            elif len(new) < 10:
                error = 'Use at least 10 characters for the new password.'
            elif new != confirm:
                error = 'New passwords do not match.'
            else:
                now = datetime.now(timezone.utc).isoformat()
                conn.execute('UPDATE users SET password_hash=?,must_change_password=0,updated_at=? WHERE id=?',
                             (generate_password_hash(new, method='scrypt'), now, user['id']))
                session['csrf_token'] = secrets.token_urlsafe(32)
                return redirect('/chat')
    return render_template('change_password.html', error=error)


@app.get('/api/auth/session')
def auth_session():
    user = _current_user()
    return jsonify({'user': user, 'csrf_token': _csrf_token()})


def _require_admin():
    user = _current_user()
    return user if user and user['role'] == 'admin' else None


def _valid_ollama_endpoint(value):
    endpoint = str(value or '').strip().rstrip('/')
    if not endpoint or not re.fullmatch(r'https?://[^\s/]+(?::\d+)?(?:/[^\s]*)?', endpoint, re.I):
        raise ValueError('Enter a valid http:// or https:// Ollama endpoint.')
    return endpoint


@app.route('/api/hosts', methods=['GET', 'POST'])
def managed_hosts_api():
    if not _require_admin():
        return jsonify({'error': 'Administrator access required'}), 403
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        name = str(data.get('name') or '').strip()
        if not re.fullmatch(r'[A-Za-z0-9][A-Za-z0-9 ._-]{1,62}', name):
            return jsonify({'error': 'Host name must be 2–63 letters, numbers, spaces, dots, dashes, or underscores.'}), 400
        try:
            endpoint = _valid_ollama_endpoint(data.get('endpoint'))
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        host_id = uuid.uuid4().hex
        now = datetime.now(timezone.utc).isoformat()
        try:
            with _db() as conn:
                conn.execute('INSERT INTO managed_hosts(id,name,endpoint,created_at,updated_at) VALUES(?,?,?,?,?)', (host_id, name, endpoint, now, now))
        except sqlite3.IntegrityError:
            return jsonify({'error': 'A managed host already uses that name.'}), 409
        return jsonify({'host': _host_summary(_managed_host_row(host_id))}), 201
    active = str(_get_setting('active_managed_host_id', '') or '')
    with _db() as conn:
        rows = conn.execute('SELECT * FROM managed_hosts ORDER BY name COLLATE NOCASE').fetchall()
    return jsonify({'active_host_id': active or 'local', 'hosts': [_host_summary(row) for row in rows]})


@app.post('/api/hosts/<host_id>/activate')
def activate_managed_host(host_id):
    if not _require_admin():
        return jsonify({'error': 'Administrator access required'}), 403
    if host_id == 'local':
        _set_setting('active_managed_host_id', '')
        return jsonify({'active_host_id': 'local'})
    row = _managed_host_row(host_id)
    if not row:
        return jsonify({'error': 'Managed host not found'}), 404
    _set_setting('active_managed_host_id', host_id)
    return jsonify({'active_host_id': host_id, 'host': _host_summary(row)})


@app.post('/api/hosts/<host_id>/pairing')
def create_host_pairing(host_id):
    if not _require_admin():
        return jsonify({'error': 'Administrator access required'}), 403
    row = _managed_host_row(host_id)
    if not row:
        return jsonify({'error': 'Managed host not found'}), 404
    pairing_token = secrets.token_urlsafe(32)
    expiry = datetime.now(timezone.utc) + timedelta(minutes=15)
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("UPDATE managed_hosts SET pairing_token_hash=?,pairing_expires_at=?,state='pairing',updated_at=? WHERE id=?",
                     (_connector_token_hash(pairing_token), expiry.isoformat(), now, host_id))
    return jsonify({'host_id': host_id, 'pairing_token': pairing_token, 'expires_at': expiry.isoformat(),
                    'server_url': request.url_root.rstrip('/'), 'endpoint': row['endpoint']})


@app.delete('/api/hosts/<host_id>')
def delete_managed_host(host_id):
    if not _require_admin():
        return jsonify({'error': 'Administrator access required'}), 403
    with _db() as conn:
        row = conn.execute('SELECT id FROM managed_hosts WHERE id=?', (host_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Managed host not found'}), 404
        conn.execute('DELETE FROM managed_hosts WHERE id=?', (host_id,))
    if _get_setting('active_managed_host_id', '') == host_id:
        _set_setting('active_managed_host_id', '')
    return jsonify({'deleted': True})


@app.post('/api/host-connector/register')
def host_connector_register():
    data = request.get_json(silent=True) or {}
    host_id = str(data.get('host_id') or '').strip()
    token = str(data.get('pairing_token') or '').strip()
    row = _managed_host_row(host_id)
    if not row or not token or not row.get('pairing_token_hash') or not hmac.compare_digest(row['pairing_token_hash'], _connector_token_hash(token)):
        return jsonify({'error': 'Invalid or expired pairing token'}), 401
    try:
        valid_until = datetime.fromisoformat(str(row.get('pairing_expires_at') or '').replace('Z', '+00:00'))
    except Exception:
        valid_until = datetime.now(timezone.utc) - timedelta(seconds=1)
    if valid_until <= datetime.now(timezone.utc):
        return jsonify({'error': 'Invalid or expired pairing token'}), 401
    connector_token = secrets.token_urlsafe(48)
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("UPDATE managed_hosts SET connector_token_hash=?,pairing_token_hash='',pairing_expires_at=NULL,state='connected',last_seen_at=?,updated_at=? WHERE id=?",
                     (_connector_token_hash(connector_token), now, now, host_id))
    return jsonify({'host_id': host_id, 'connector_token': connector_token, 'poll_seconds': 2,
                    'allowed_operations': ['helper.status', 'helper.gpu', 'helper.apply']})


@app.post('/api/host-connector/poll')
def host_connector_poll():
    row = _connector_identity()
    if not row:
        return jsonify({'error': 'Connector authentication failed'}), 401
    data = request.get_json(silent=True) or {}
    snapshot = data.get('snapshot') if isinstance(data.get('snapshot'), dict) else {}
    snapshot_json = json.dumps(snapshot, separators=(',', ':'))[:262144]
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("UPDATE managed_hosts SET state='connected',last_seen_at=?,snapshot_json=?,updated_at=? WHERE id=?", (now, snapshot_json, now, row['id']))
        action = conn.execute("SELECT * FROM host_connector_actions WHERE host_id=? AND status='queued' ORDER BY created_at LIMIT 1", (row['id'],)).fetchone()
        if action:
            conn.execute("UPDATE host_connector_actions SET status='running',updated_at=? WHERE id=?", (now, action['id']))
    if not action:
        return jsonify({'action': None, 'poll_seconds': 2})
    try:
        payload = json.loads(action['payload_json'] or '{}')
    except Exception:
        payload = {}
    return jsonify({'action': {'id': action['id'], 'operation': action['operation'], 'payload': payload}, 'poll_seconds': 1})


@app.post('/api/host-connector/actions/<action_id>/result')
def host_connector_action_result(action_id):
    row = _connector_identity()
    if not row:
        return jsonify({'error': 'Connector authentication failed'}), 401
    data = request.get_json(silent=True) or {}
    result = data.get('result') if isinstance(data.get('result'), dict) else {}
    status = 'completed' if data.get('ok') is True else 'failed'
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        action = conn.execute('SELECT host_id,status FROM host_connector_actions WHERE id=?', (action_id,)).fetchone()
        if not action or action['host_id'] != row['id'] or action['status'] not in {'running', 'queued'}:
            return jsonify({'error': 'Unknown action'}), 404
        conn.execute('UPDATE host_connector_actions SET status=?,result_json=?,updated_at=? WHERE id=?',
                     (status, json.dumps(result, separators=(',', ':'))[:262144], now, action_id))
    return jsonify({'ok': True})


@app.route('/api/auth/users', methods=['GET', 'POST'])
def auth_users():
    admin = _require_admin()
    if not admin:
        return jsonify({'error': 'Administrator access required'}), 403
    if request.method == 'GET':
        with _db() as conn:
            rows = conn.execute('SELECT id,username,role,active,must_change_password,created_at,last_login_at FROM users ORDER BY id').fetchall()
        return jsonify({'users': [dict(x) for x in rows], 'primary_user_id': rows[0]['id'] if rows else None})
    data = request.get_json(silent=True) or {}
    username = str(data.get('username') or '').strip()
    password = str(data.get('password') or '')
    role = 'admin' if data.get('role') == 'admin' else 'user'
    if not re.fullmatch(r'[A-Za-z0-9_.-]{3,64}', username):
        return jsonify({'error': 'Username must be 3–64 letters, numbers, dots, dashes, or underscores.'}), 400
    if len(password) < 10:
        return jsonify({'error': 'Temporary password must be at least 10 characters.'}), 400
    now = datetime.now(timezone.utc).isoformat()
    try:
        with _db() as conn:
            cur = conn.execute('INSERT INTO users(username,password_hash,role,active,must_change_password,created_at,updated_at) VALUES(?,?,?,?,?,?,?)',
                               (username, generate_password_hash(password, method='scrypt'), role, 1, 1, now, now))
        return jsonify({'ok': True, 'id': cur.lastrowid}), 201
    except sqlite3.IntegrityError:
        return jsonify({'error': 'That username already exists.'}), 409


@app.route('/api/auth/users/<int:user_id>', methods=['PATCH', 'DELETE'])
def auth_user(user_id):
    actor = _current_user()
    admin = actor if actor and actor['role'] == 'admin' else None
    if not admin and (not actor or actor['id'] != user_id or request.method == 'DELETE'):
        return jsonify({'error': 'Administrator access required'}), 403
    with _db() as conn:
        target = conn.execute('SELECT * FROM users WHERE id=?', (user_id,)).fetchone()
        primary = conn.execute('SELECT id FROM users ORDER BY id LIMIT 1').fetchone()
        if not target:
            return jsonify({'error': 'User not found'}), 404
        if request.method == 'DELETE':
            if primary and user_id == primary['id']:
                return jsonify({'error': 'The primary account cannot be deleted.'}), 409
            if target['role'] == 'admin' and target['active']:
                active_admins = conn.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND active=1").fetchone()[0]
                if active_admins <= 1:
                    return jsonify({'error': 'The final active administrator cannot be deleted.'}), 409
            conn.execute('DELETE FROM users WHERE id=?', (user_id,))
            return jsonify({'ok': True})
        data = request.get_json(silent=True) or {}
        if not admin and any(key in data for key in ('role', 'active')):
            return jsonify({'error': 'Only administrators can change roles or account status.'}), 403
        updates, values = [], []
        if 'username' in data:
            username = str(data.get('username') or '').strip()
            if not re.fullmatch(r'[A-Za-z0-9_.-]{3,64}', username):
                return jsonify({'error': 'Invalid username format.'}), 400
            updates.append('username=?'); values.append(username)
        if 'password' in data:
            password = str(data.get('password') or '')
            if actor and actor['id'] == user_id:
                return jsonify({'error': 'Use the change-password screen so the current password can be verified.'}), 400
            if len(password) < 10:
                return jsonify({'error': 'Password must be at least 10 characters.'}), 400
            updates.extend(['password_hash=?', f"must_change_password={'0' if actor and actor['id'] == user_id else '1'}"]); values.append(generate_password_hash(password, method='scrypt'))
        new_role = 'admin' if data.get('role') == 'admin' else ('user' if 'role' in data else target['role'])
        new_active = 1 if data.get('active') else (0 if 'active' in data else target['active'])
        losing_admin = target['role'] == 'admin' and target['active'] and (new_role != 'admin' or not new_active)
        if losing_admin:
            active_admins = conn.execute("SELECT COUNT(*) FROM users WHERE role='admin' AND active=1").fetchone()[0]
            if active_admins <= 1:
                return jsonify({'error': 'The final active administrator cannot be disabled or demoted.'}), 409
        if primary and user_id == primary['id'] and not new_active:
            return jsonify({'error': 'The primary account cannot be disabled.'}), 409
        if 'role' in data: updates.append('role=?'); values.append(new_role)
        if 'active' in data: updates.append('active=?'); values.append(new_active)
        if not updates:
            return jsonify({'error': 'No supported changes supplied.'}), 400
        updates.append('updated_at=?'); values.append(datetime.now(timezone.utc).isoformat()); values.append(user_id)
        try:
            conn.execute(f"UPDATE users SET {','.join(updates)} WHERE id=?", values)
        except sqlite3.IntegrityError:
            return jsonify({'error': 'That username already exists.'}), 409
    return jsonify({'ok': True})



def _thinking_override_key(model):
    return 'thinking_override:' + str(model or '').strip()


def _thinking_override(model):
    model = str(model or '').strip()
    if not model:
        return None
    raw = _get_setting(_thinking_override_key(model), '')
    if not raw:
        return None
    try:
        data = json.loads(raw)
    except Exception:
        return None
    if not isinstance(data, dict) or not data.get('enabled'):
        return None
    levels = []
    for level in (data.get('levels') or []):
        level = str(level).strip().lower()
        if level in ('off','on','low','medium','high','xhigh','max') and level not in levels:
            levels.append(level)
    if not levels:
        return None
    method = str(data.get('method') or 'native').strip().lower()
    if method not in ('native','xhigh-default','system-directive'):
        method = 'native'
    default = str(data.get('default') or '').strip().lower()
    if default not in levels:
        default = levels[0]
    return {'enabled': True, 'levels': levels, 'method': method, 'default': default}


def _save_thinking_override(model, override):
    model = str(model or '').strip()
    if not model:
        return
    if not isinstance(override, dict) or not override.get('enabled'):
        _set_setting(_thinking_override_key(model), '')
    else:
        levels = []
        for level in (override.get('levels') or []):
            level = str(level).strip().lower()
            if level in ('off','on','low','medium','high','xhigh','max') and level not in levels:
                levels.append(level)
        if not levels:
            raise ValueError('Manual thinking override requires at least one enabled level.')
        method = str(override.get('method') or 'native').strip().lower()
        if method not in ('native','xhigh-default','system-directive'):
            raise ValueError('Unsupported thinking override control method.')
        default = str(override.get('default') or '').strip().lower()
        if default not in levels:
            raise ValueError('Default thinking level must be one of the enabled levels.')
        payload = {'enabled': True, 'levels': levels, 'method': method, 'default': default}
        _set_setting(_thinking_override_key(model), json.dumps(payload, separators=(',',':')))
    with _THINKING_PROFILE_LOCK:
        _THINKING_PROFILE_CACHE.pop(model, None)


def _manual_thinking_profile(model, override, detected=None):
    levels = list(override.get('levels') or [])
    method = override.get('method') or 'native'
    labels = {'off':'Off','on':'On','low':'Low','medium':'Medium','high':'High','xhigh':'XHigh','max':'Max'}
    options = []
    for level in levels:
        option = {'value': level, 'label': f'Thinking: {labels.get(level, level.title())}'}
        if level == 'off':
            option['wire_value'] = False
        elif level == 'on':
            option['wire_value'] = True
        elif method == 'system-directive' and level in ('low','medium','high','xhigh','max'):
            option['wire_value'] = None
            option['system_instruction'] = f'Reasoning effort: {level}'
        elif level == 'xhigh':
            option['wire_value'] = None if method == 'xhigh-default' else 'max'
        elif level == 'max':
            option['wire_value'] = 'max'
        else:
            option['wire_value'] = level
        options.append(option)
    method_summary = {
        'native': 'Native Ollama think mapping; XHigh is sent as max.',
        'xhigh-default': 'Native Ollama think mapping, except XHigh omits think so the model template can use its own XHigh default.',
        'system-directive': 'Template-level system directive mapping (Reasoning effort: <level>); useful for Muse/custom GGUF templates.',
    }[method]
    evidence = ['manual per-model override', f'control method: {method}', 'manual levels: ' + ', '.join(levels)]
    if detected and detected.get('summary'):
        evidence.append('auto-detection was overridden: ' + str(detected.get('summary')))
    return {
        'supported': True,
        'mode': 'manual-override',
        'options': options,
        'default': override.get('default') if override.get('default') in levels else levels[0],
        'confidence': 'manual',
        'summary': method_summary,
        'evidence': evidence,
        'manual_override': True,
        'override_config': override,
        'detected_profile': detected,
    }


def _effective_thinking_profile(show_data, model_name=''):
    detected = _thinking_profile(show_data, model_name)
    override = _thinking_override(model_name)
    return _manual_thinking_profile(model_name, override, detected) if override else detected


def _configured_upstream():
    host_id = str(_get_setting('active_managed_host_id', '') or '').strip()
    if host_id:
        try:
            with _db() as conn:
                row = conn.execute('SELECT endpoint FROM managed_hosts WHERE id=?', (host_id,)).fetchone()
            if row and str(row['endpoint'] or '').strip():
                return str(row['endpoint']).strip().rstrip('/')
        except sqlite3.Error:
            pass
    return (_get_setting('ollama_endpoint', '') or '').strip().rstrip('/')


def _managed_host_row(host_id):
    if not host_id:
        return None
    with _db() as conn:
        row = conn.execute('SELECT * FROM managed_hosts WHERE id=?', (host_id,)).fetchone()
    return dict(row) if row else None


def _active_managed_host():
    return _managed_host_row(str(_get_setting('active_managed_host_id', '') or '').strip())


def _host_summary(row):
    row = dict(row)
    try:
        snapshot = json.loads(row.get('snapshot_json') or '{}')
    except Exception:
        snapshot = {}
    return {
        'id': row['id'], 'name': row['name'], 'endpoint': row['endpoint'],
        'state': row['state'], 'last_seen_at': row['last_seen_at'],
        'paired': bool(row.get('connector_token_hash')), 'snapshot': snapshot if isinstance(snapshot, dict) else {},
    }


def _connector_token_hash(token):
    return hashlib.sha256(str(token).encode('utf-8')).hexdigest()


def _connector_identity():
    host_id = str(request.headers.get('X-Aperyn-Host-ID') or '').strip()
    auth = str(request.headers.get('Authorization') or '')
    token = auth[7:].strip() if auth.startswith('Bearer ') else ''
    if not host_id or not token:
        return None
    row = _managed_host_row(host_id)
    if not row or not row.get('connector_token_hash'):
        return None
    if not hmac.compare_digest(row['connector_token_hash'], _connector_token_hash(token)):
        return None
    return row


def _queue_host_action(host_id, operation, payload=None):
    if operation not in {'helper.status', 'helper.gpu', 'helper.apply'}:
        raise ValueError('Unsupported remote host operation')
    now = datetime.now(timezone.utc).isoformat()
    action_id = uuid.uuid4().hex
    with _db() as conn:
        conn.execute('INSERT INTO host_connector_actions(id,host_id,operation,payload_json,status,result_json,created_at,updated_at) VALUES(?,?,?,?,?,?,?,?)',
                     (action_id, host_id, operation, json.dumps(payload or {}), 'queued', '{}', now, now))
    return action_id


def _wait_for_host_action(action_id, timeout=25):
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        with _db() as conn:
            row = conn.execute('SELECT status,result_json FROM host_connector_actions WHERE id=?', (action_id,)).fetchone()
        if row and row['status'] in {'completed', 'failed'}:
            try:
                result = json.loads(row['result_json'] or '{}')
            except Exception:
                result = {}
            if row['status'] == 'failed':
                raise RuntimeError(str(result.get('error') or 'Remote host action failed'))
            return result if isinstance(result, dict) else {}
        time.sleep(0.2)
    raise RuntimeError('Remote host did not respond. Check that its Aperyn Host Connector is running.')


def _remote_host_action(host_id, operation, payload=None, timeout=25):
    row = _managed_host_row(host_id)
    if not row or not row.get('connector_token_hash'):
        raise RuntimeError('This Ollama host has not been paired with an Aperyn Host Connector.')
    if row.get('state') != 'connected':
        raise RuntimeError('This Ollama host connector is offline.')
    return _wait_for_host_action(_queue_host_action(host_id, operation, payload), timeout)


def _parse_scalar(value):
    value = value.strip().strip('"')
    low = value.lower()
    if low in ('true','false'):
        return low == 'true'
    try:
        if any(c in value for c in '.eE'):
            return float(value)
        return int(value)
    except ValueError:
        return value


def _parse_parameters(text):
    parsed = {}
    for raw in (text or '').splitlines():
        line = raw.strip()
        if not line:
            continue
        parts = line.split(None, 1)
        if len(parts) != 2:
            continue
        key, value = parts
        val = _parse_scalar(value)
        if key in parsed:
            if not isinstance(parsed[key], list):
                parsed[key] = [parsed[key]]
            parsed[key].append(val)
        else:
            parsed[key] = val
    return parsed


def _model_pref(model):
    with _db() as conn:
        row = conn.execute('SELECT model, mtp_enabled, mtp_draft_n_max, updated_at FROM model_preferences WHERE model = ?', (model,)).fetchone()
    if not row:
        return None
    return {
        'model': row['model'],
        'mtp_enabled': bool(row['mtp_enabled']),
        'mtp_draft_n_max': int(row['mtp_draft_n_max']),
        'updated_at': row['updated_at'],
        'source': 'manager',
    }


def _save_model_pref(model, mtp_enabled, mtp_draft_n_max):
    now = datetime.now(timezone.utc).isoformat()
    with _db() as conn:
        conn.execute("""INSERT INTO model_preferences(model, mtp_enabled, mtp_draft_n_max, updated_at)
                        VALUES(?,?,?,?)
                        ON CONFLICT(model) DO UPDATE SET mtp_enabled=excluded.mtp_enabled,
                        mtp_draft_n_max=excluded.mtp_draft_n_max, updated_at=excluded.updated_at""",
                     (model, 1 if mtp_enabled else 0, int(mtp_draft_n_max), now))


def _delete_model_pref(model):
    with _db() as conn:
        conn.execute('DELETE FROM model_preferences WHERE model = ?', (model,))


def _model_limits(show_data):
    """Derive safe model-specific editor ceilings from /api/show --verbose.

    Ollama models expose architecture metadata such as `<arch>.block_count` and
    `<arch>.context_length`. Ollama's layer allocator treats BlockCount()+1 as
    the full-offload boundary (the repeating blocks plus the output layer), so
    the UI can present an exact num_gpu ceiling rather than an arbitrary 999.
    """
    info = show_data.get('model_info') or {}
    details = show_data.get('details') or {}
    architecture = ''
    if isinstance(info, dict):
        architecture = str(info.get('general.architecture') or '').strip()
    if not architecture:
        architecture = str(details.get('family') or '').strip()

    block_count = None
    context_length = None
    block_key = None
    context_key = None
    if isinstance(info, dict):
        preferred_block = f'{architecture}.block_count' if architecture else None
        preferred_ctx = f'{architecture}.context_length' if architecture else None
        candidates = list(info.items())
        if preferred_block and preferred_block in info:
            candidates = [(preferred_block, info[preferred_block])] + [(k,v) for k,v in candidates if k != preferred_block]
        for key, value in candidates:
            key_l = str(key).lower()
            if block_count is None and (key_l == 'block_count' or key_l.endswith('.block_count')):
                try:
                    n = int(value)
                    if n > 0:
                        block_count, block_key = n, str(key)
                except (TypeError, ValueError):
                    pass
            if context_length is None and (key_l == 'context_length' or key_l.endswith('.context_length')):
                try:
                    n = int(value)
                    if n > 0:
                        context_length, context_key = n, str(key)
                except (TypeError, ValueError):
                    pass

    gpu_max = block_count + 1 if block_count else None
    return {
        'architecture': architecture or None,
        'block_count': block_count,
        'block_count_key': block_key,
        'num_gpu': {
            'min': -1,
            'max': gpu_max,
            'full_offload': gpu_max,
            'summary': (f'{gpu_max} offloadable layers ({block_count} repeating blocks + output layer)' if gpu_max else 'Model layer count was not exposed by Ollama.'),
        },
        'num_ctx': {
            'min': 512,
            'max': context_length,
            'training_context': context_length,
            'context_key': context_key,
            'summary': (f'Model training context: {context_length:,} tokens' if context_length else 'Model training context was not exposed by Ollama.'),
        },
    }


_THINKING_PROFILE_CACHE = {}
_THINKING_PROFILE_LOCK = threading.Lock()
_THINKING_PROFILE_TTL = 300


def _thinking_profile(show_data, model_name=''):
    """Model-aware thinking controls with template-specific wire mappings.

    Ollama's public API accepts a small wire vocabulary (boolean, low/medium/high/max),
    but individual GGUF chat templates can expose different model-facing labels and
    defaults.  In particular, some Qwen3.8 templates accept literal ``xhigh`` inside
    the template yet reject Ollama's ``max`` value, while other Qwen3.8 conversions
    accept ``max`` and may default to ``medium``.  The UI therefore keeps a separate
    display value and wire value for every installed model.
    """
    show_data = show_data or {}
    caps = [str(x).strip().lower() for x in (show_data.get('capabilities') or [])]
    info = show_data.get('model_info') or {}
    details = show_data.get('details') or {}
    template = str(show_data.get('template') or '')
    template_l = template.lower()
    renderer = str(show_data.get('renderer') or '').lower()

    family_bits = [str(model_name or ''), str(details.get('family') or ''), renderer]
    families = details.get('families') or []
    if isinstance(families, list):
        family_bits.extend(str(x) for x in families)
    if isinstance(info, dict):
        family_bits.extend([str(info.get('general.architecture') or ''), str(info.get('general.name') or '')])
    identity = ' '.join(family_bits).lower().replace('_', '-').replace(' ', '-')

    capability_reported = 'thinking' in caps
    template_markers = ('.thinking', '.think', 'isthinkset', '<think>', '/think', '/no_think', 'reasoning_effort', 'thinking_effort', 'reasoning_strength', 'enable_thinking')
    template_evidence = any(marker in template_l for marker in template_markers)
    known_thinking_family = any(x in identity for x in ('qwen3', 'deepseek-r1', 'deepseek-v3.1', 'deepseek-v31', 'gpt-oss'))
    supported = capability_reported or template_evidence or known_thinking_family

    if not supported:
        return {'supported': False, 'mode': 'none', 'options': [], 'default': None,
                'confidence': 'reported' if caps else 'unknown',
                'summary': 'This model does not advertise Ollama thinking support.',
                'evidence': ['capabilities: ' + ', '.join(caps)] if caps else []}

    if 'gpt-oss' in identity or 'gptoss' in identity:
        return {'supported': True, 'mode': 'levels-only',
                'options': [
                    {'value': 'low', 'wire_value': 'low', 'label': 'Thinking: Low'},
                    {'value': 'medium', 'wire_value': 'medium', 'label': 'Thinking: Medium'},
                    {'value': 'high', 'wire_value': 'high', 'label': 'Thinking: High'},
                ], 'default': 'medium', 'confidence': 'reported' if capability_reported else 'family',
                'summary': 'GPT-OSS uses low / medium / high reasoning effort; boolean on/off is not offered.',
                'evidence': ['family: gpt-oss'] + (['capability: thinking'] if capability_reported else [])}

    def _template_levels_and_default():
        """Return (ordered levels, default level, evidence strings)."""
        found = []
        evidence = []
        default_level = None

        # Strongest signal: templates which spell out their accepted values in an
        # exception message, e.g. "Supported types are xhigh (default), medium, and low".
        for m in re.finditer(r'supported\s+(?:types|values|reasoning(?:\s+efforts?)?)\s+(?:are|:)\s*([^\n.;}]+)', template_l):
            chunk = m.group(1)
            levels = re.findall(r'\b(xhigh|max|high|medium|low)\b', chunk)
            if levels:
                for level in levels:
                    if level not in found:
                        found.append(level)
                dm = re.search(r'\b(xhigh|max|high|medium|low)\b\s*\(\s*default\s*\)', chunk)
                if dm:
                    default_level = dm.group(1)
                evidence.append('template supported-types declaration: ' + ', '.join(levels))

        # Common Jinja assignments/default filters expose the real template default.
        default_patterns = [
            r'(?:reasoning_effort|thinking_effort|reasoning_strength)\s*=\s*(?:reasoning_effort|thinking_effort|reasoning_strength)\s+or\s+["\'](xhigh|max|high|medium|low)["\']',
            r'(?:reasoning_effort|thinking_effort|reasoning_strength)\s*\|\s*default\s*\(\s*["\'](xhigh|max|high|medium|low)["\']',
            r'(?:reasoning_effort|thinking_effort|reasoning_strength)[^\n]{0,120}\bdefault[^\n]{0,60}["\'](xhigh|max|high|medium|low)["\']',
            r'["\'](xhigh|max|high|medium|low)["\']\s*\(\s*default\s*\)',
        ]
        for pattern in default_patterns:
            dm = re.search(pattern, template_l)
            if dm:
                default_level = dm.group(1)
                break

        # Broader level discovery for templates that use comparisons/lists but do not
        # contain an explicit Supported-types sentence.
        for level in ('low', 'medium', 'high', 'xhigh', 'max'):
            quoted = re.search(r'["\']' + re.escape(level) + r'["\']', template_l)
            controlled = re.search(r'(?:eq|ne|contains|in\s*\[|reasoning_effort|thinking_effort|reasoning_strength|think)[^\n]{0,140}["\']' + re.escape(level) + r'["\']', template_l)
            if quoted and controlled and level not in found:
                found.append(level)

        return found, default_level, evidence

    level_hits, template_default, level_evidence = _template_levels_and_default()

    # Muse Glimmer uses a template-level `reasoning_strength` control rather than
    # Ollama's normal boolean/effort `think` switch.  The canonical model supports
    # low / medium / high / xhigh and defaults to high.  Ollama's native /api/chat
    # does not expose arbitrary chat_template_kwargs, so the WebUI uses the
    # model-supported `Reasoning effort: <level>` system directive, which current
    # Muse GGUF templates normalize to their `Reasoning strength` control.
    is_muse_glimmer = any(x in identity for x in ('muse-glimmer', 'museglimmer')) or 'reasoning_strength' in template_l
    if is_muse_glimmer:
        muse_levels = [x for x in ('low', 'medium', 'high', 'xhigh') if x in level_hits]
        if len(muse_levels) < 2:
            muse_levels = ['low', 'medium', 'high', 'xhigh']
        muse_default = template_default if template_default in muse_levels else 'high'
        options = [
            {'value': 'auto', 'wire_value': None, 'label': f'Thinking: Default ({muse_default.title()})'}
        ]
        for level in muse_levels:
            options.append({
                'value': level,
                'wire_value': None,
                'label': f"Thinking: {'XHigh' if level == 'xhigh' else level.title()}",
                'system_instruction': f'Reasoning effort: {level}',
            })
        return {
            'supported': True,
            'mode': 'reasoning-strength',
            'options': options,
            'default': 'auto',
            'confidence': 'template' if 'reasoning_strength' in template_l else 'family',
            'summary': f'Muse-style reasoning strength detected. Levels: {" / ".join(muse_levels)}; template default: {muse_default}.',
            'evidence': (['capability: thinking'] if capability_reported else []) + (['template variable: reasoning_strength'] if 'reasoning_strength' in template_l else ['family/model identity: Muse Glimmer']) + level_evidence + [f'template levels: {", ".join(muse_levels)}', f'template default: {muse_default}'],
            'template_default': muse_default,
            'template_levels': muse_levels,
            'request_strategy': 'system-reasoning-effort',
        }

    is_qwen38 = any(x in identity for x in ('qwen3.8', 'qwen38')) or ('qwen3.8' in template_l) or ('xhigh' in template_l and 'reasoning_effort' in template_l)
    can_disable = any(x in template_l for x in ('enable_thinking', '/no_think', 'isthinkset', 'thinking=false', 'thinking == false')) or is_qwen38

    # For Qwen3.8 conversions, trust the installed template over family assumptions.
    # HauHau-style templates commonly accept literal xhigh/medium/low and reject max;
    # in that case xhigh can only be requested safely by omitting the wire level when
    # xhigh is the template default.  Other conversions accept max, so XHigh maps to max.
    if is_qwen38 and level_hits:
        options = []
        if can_disable:
            options.append({'value': 'off', 'wire_value': False, 'label': 'Thinking: Off'})
        if 'low' in level_hits:
            options.append({'value': 'low', 'wire_value': 'low', 'label': 'Thinking: Low'})
        if 'medium' in level_hits:
            options.append({'value': 'medium', 'wire_value': 'medium', 'label': 'Thinking: Medium'})
        if 'high' in level_hits:
            options.append({'value': 'high', 'wire_value': 'high', 'label': 'Thinking: High'})

        top_note = None
        if 'max' in level_hits:
            options.append({'value': 'xhigh', 'wire_value': 'max', 'label': 'Thinking: XHigh'})
            top_note = 'XHigh is sent to Ollama as max because this installed template accepts max.'
        elif 'xhigh' in level_hits and template_default == 'xhigh':
            options.append({'value': 'xhigh', 'wire_value': None, 'label': 'Thinking: XHigh'})
            top_note = 'XHigh uses this template’s default by omitting the Ollama think level; this template rejects max.'
        elif 'xhigh' in level_hits:
            # Ollama does not accept literal xhigh on the native API.  If xhigh is not
            # the template default and max is not accepted, there is no safe wire value.
            top_note = 'This template advertises xhigh but neither defaults to it nor accepts Ollama max, so XHigh cannot be selected safely through the native API.'

        if not options:
            options = [
                {'value': 'auto', 'wire_value': None, 'label': 'Thinking: Default'},
                {'value': 'off', 'wire_value': False, 'label': 'Thinking: Off'},
                {'value': 'on', 'wire_value': True, 'label': 'Thinking: On'},
            ]
        preferred = template_default if any(o['value'] == template_default for o in options) else None
        if preferred is None and template_default == 'max' and any(o['value'] == 'xhigh' for o in options):
            preferred = 'xhigh'
        if preferred is None:
            preferred = 'medium' if any(o['value'] == 'medium' for o in options) else options[0]['value']
        evidence = (['capability: thinking'] if capability_reported else []) + ['family/template: qwen3.8'] + level_evidence
        if level_hits:
            evidence.append('template levels: ' + ', '.join(level_hits))
        if template_default:
            evidence.append('template default: ' + template_default)
        if top_note:
            evidence.append(top_note)
        return {'supported': True, 'mode': 'qwen38-template', 'options': options, 'default': preferred,
                'confidence': 'template',
                'summary': top_note or 'Qwen3.8 reasoning levels were derived from this installed model template.',
                'evidence': evidence,
                'template_default': template_default,
                'template_levels': level_hits}

    if len(level_hits) >= 2:
        options = []
        if can_disable:
            options.append({'value': 'off', 'wire_value': False, 'label': 'Thinking: Off'})
        labels = {'low':'Low','medium':'Medium','high':'High','xhigh':'XHigh','max':'Max'}
        for level in level_hits:
            if level == 'xhigh':
                # Native Ollama cannot send literal xhigh.  Only expose it when the
                # template's default gives us a safe omission path, or when max is also
                # accepted and can act as the wire equivalent.
                if 'max' in level_hits:
                    options.append({'value':'xhigh','wire_value':'max','label':'Thinking: XHigh'})
                elif template_default == 'xhigh':
                    options.append({'value':'xhigh','wire_value':None,'label':'Thinking: XHigh'})
            else:
                options.append({'value':level,'wire_value':level,'label':f'Thinking: {labels[level]}'})
        unique=[]; seen=set()
        for option in options:
            if option['value'] not in seen:
                seen.add(option['value']); unique.append(option)
        options=unique
        if not options:
            options = [{'value':'auto','wire_value':None,'label':'Thinking: Default'}]
        default = template_default if any(o['value']==template_default for o in options) else ('auto' if any(o['value']=='auto' for o in options) else options[0]['value'])
        return {'supported': True, 'mode': 'levels', 'options': options, 'default': default,
                'confidence': 'template',
                'summary': "Named reasoning levels were detected in this model's installed chat template.",
                'evidence': (['capability: thinking'] if capability_reported else []) + level_evidence + [f'template levels: {", ".join(level_hits)}'] + ([f'template default: {template_default}'] if template_default else []),
                'template_default': template_default,
                'template_levels': level_hits}

    evidence=[]
    if capability_reported: evidence.append('capability: thinking')
    if template_evidence: evidence.append('thinking controls detected in chat template')
    if known_thinking_family and not capability_reported: evidence.append('known thinking model family')
    return {'supported': True, 'mode': 'boolean',
            'options': [
                {'value': 'auto', 'wire_value': None, 'label': 'Thinking: Default'},
                {'value': 'off', 'wire_value': False, 'label': 'Thinking: Off'},
                {'value': 'on', 'wire_value': True, 'label': 'Thinking: On'},
            ], 'default': 'auto',
            'confidence': 'reported' if capability_reported else ('template' if template_evidence else 'family'),
            'summary': 'Thinking is supported, but no model-specific effort levels were proven; only default / off / on are shown.',
            'evidence': evidence}


def _thinking_profile_for_model(model_name, refresh=False):
    model_name = str(model_name or '').strip()
    if not model_name:
        raise ValueError('model is required')
    now = time.time()
    if not refresh:
        with _THINKING_PROFILE_LOCK:
            cached = _THINKING_PROFILE_CACHE.get(model_name)
            if cached and now - cached['at'] < _THINKING_PROFILE_TTL:
                return dict(cached['profile'])
    show = _safe_ollama_json('POST', '/api/show', json={'model': model_name, 'verbose': True}, timeout=30)
    profile = _effective_thinking_profile(show, model_name)
    profile['model'] = model_name
    with _THINKING_PROFILE_LOCK:
        _THINKING_PROFILE_CACHE[model_name] = {'at': now, 'profile': dict(profile)}
    return profile


def _think_value_allowed(profile, value):
    if value is None:
        return True
    normalized = 'on' if value is True else ('off' if value is False else str(value).strip().lower())
    return normalized in {str(x.get('value')) for x in (profile.get('options') or [])}


def _thinking_wire_value(profile, selected):
    """Translate a model/UI reasoning choice into Ollama's native think value."""
    if selected is None:
        return None
    normalized = 'on' if selected is True else ('off' if selected is False else str(selected).strip().lower())
    for option in (profile.get('options') or []):
        if str(option.get('value')) == normalized:
            if 'wire_value' in option:
                return option.get('wire_value')
            if normalized == 'auto': return None
            if normalized == 'on': return True
            if normalized == 'off': return False
            if normalized == 'xhigh': return 'max'
            return normalized
    return selected


def _thinking_system_instruction(profile, selected):
    """Return an optional system directive for template-level reasoning controls."""
    if selected is None:
        return None
    normalized = 'on' if selected is True else ('off' if selected is False else str(selected).strip().lower())
    for option in (profile.get('options') or []):
        if str(option.get('value')) == normalized:
            instruction = option.get('system_instruction')
            return str(instruction).strip() if instruction else None
    return None


def _mtp_capability(show_data):
    """Mirror Ollama's embedded-MTP detection as closely as /api/show allows.

    Ollama 0.32.14+ enables embedded MTP when nextn_predict_layers > 0, or for
    legacy qwen35/qwen35moe GGUFs when tensors with the `mtp.` prefix exist.
    /api/show --verbose exposes the metadata/tensor names needed for a safe
    best-effort UI check without attempting to start the model.
    """
    evidence = []
    model_info = show_data.get('model_info') or {}
    details = show_data.get('details') or {}

    architecture = ''
    if isinstance(model_info, dict):
        architecture = str(model_info.get('general.architecture') or '')
        for key, value in model_info.items():
            key_l = str(key).lower()
            if key_l == 'nextn_predict_layers' or key_l.endswith('.nextn_predict_layers'):
                try:
                    layers = int(value)
                except (TypeError, ValueError):
                    layers = 0
                if layers > 0:
                    evidence.append(f'metadata: {key}={value}')
    if not architecture:
        architecture = str(details.get('family') or '').lower()

    tensors = show_data.get('tensors')
    tensor_list_present = isinstance(tensors, list) and len(tensors) > 0
    if architecture.lower() in ('qwen35', 'qwen35moe') and isinstance(tensors, list):
        for tensor in tensors:
            name = str(tensor.get('name') or '') if isinstance(tensor, dict) else str(tensor)
            if name.lower().startswith('mtp.'):
                evidence.append(f'legacy {architecture} tensor: {name}')
                if len(evidence) >= 8:
                    break

    if evidence:
        status = 'supported'
        summary = 'Embedded MTP layers detected using Ollama-compatible NextN/MTP checks.'
    elif tensor_list_present:
        status = 'unsupported'
        summary = 'Verbose tensor metadata was available, but Ollama-compatible embedded MTP evidence was not found.'
    else:
        status = 'unknown'
        summary = 'Ollama did not return enough verbose tensor metadata to prove whether embedded MTP layers are present.'

    return {
        'status': status,
        'summary': summary,
        'evidence': evidence[:8],
        'tensor_count': len(tensors) if isinstance(tensors, list) else None,
        'architecture': architecture or None,
        'detection': 'nextn_predict_layers > 0; legacy qwen35/qwen35moe mtp.* tensors',
    }


def _effective_mtp_pref(model, parsed_parameters=None):
    stored = _model_pref(model)
    if stored:
        return stored
    parsed_parameters = parsed_parameters or {}
    raw = parsed_parameters.get('draft_num_predict')
    if isinstance(raw, list):
        raw = raw[0] if raw else 0
    try:
        depth = int(raw or 0)
    except (TypeError, ValueError):
        depth = 0
    return {
        'model': model,
        'mtp_enabled': depth > 0,
        'mtp_draft_n_max': depth if depth > 0 else 2,
        'updated_at': None,
        'source': 'model' if raw is not None else 'default',
    }


def _log_request(endpoint, model, status_code, latency_ms, final=None, client_ip=None, request_meta=None):
    final = final or {}
    request_meta = request_meta or {}
    prompt_tokens = int(final.get('prompt_eval_count') or 0)
    output_tokens = int(final.get('eval_count') or 0)
    eval_duration = int(final.get('eval_duration') or 0)
    client_name = str(request_meta.get('client_name') or '')[:120]
    safe_meta = {k: v for k, v in request_meta.items() if k not in ('prompt','messages','input','content')}
    try:
        meta_json = json.dumps(safe_meta, ensure_ascii=False)[:24000]
    except Exception:
        meta_json = '{}'
    with _db() as conn:
        conn.execute('INSERT INTO request_log(created_at, endpoint, model, status_code, latency_ms, prompt_tokens, output_tokens, eval_duration_ns, client_ip, client_name, request_meta_json) VALUES(?,?,?,?,?,?,?,?,?,?,?)',
                     (datetime.now(timezone.utc).isoformat(), endpoint, model, int(status_code), float(latency_ms), prompt_tokens, output_tokens, eval_duration, client_ip, client_name, meta_json))


def _safe_ollama_json(method, path, **kwargs):
    response = requests.request(method, f'{OLLAMA_API}{path}', timeout=kwargs.pop('timeout', 15), **kwargs)
    response.raise_for_status()
    return response.json() if response.content else {}


def _start_live_generation(model, endpoint, mtp_pref):
    ident = uuid.uuid4().hex[:12]
    now = datetime.now(timezone.utc).isoformat()
    entry = {
        'id': ident,
        'model': model or 'unknown',
        'endpoint': endpoint,
        'started_at': now,
        'phase': 'starting',
        'estimated_tokens': 0,
        'stream_chunks': 0,
        'characters': 0,
        'prompt_tokens': None,
        'output_tokens': None,
        'exact_tps': None,
        'mtp_enabled': bool((mtp_pref or {}).get('mtp_enabled')),
        'mtp_draft_n_max': int((mtp_pref or {}).get('mtp_draft_n_max') or 0),
        '_start_perf': time.perf_counter(),
        '_first_output_perf': None,
        '_chunk_times': [],
    }
    with _live_lock:
        _active_generations[ident] = entry
    return ident


def _stream_text_size(obj):
    if not isinstance(obj, dict):
        return 0
    text = obj.get('response') or obj.get('thinking') or ''
    message = obj.get('message')
    if isinstance(message, dict):
        text += str(message.get('content') or '')
        text += str(message.get('thinking') or '')
    return len(text)


def _update_live_generation(ident, obj):
    now_perf = time.perf_counter()
    with _live_lock:
        entry = _active_generations.get(ident)
        if not entry:
            return
        if obj.get('done'):
            entry['phase'] = 'finishing'
            return
        chars = _stream_text_size(obj)
        if chars > 0:
            if entry['_first_output_perf'] is None:
                entry['_first_output_perf'] = now_perf
            entry['phase'] = 'generating'
            entry['stream_chunks'] += 1
            entry['characters'] += chars
            entry['_chunk_times'].append(now_perf)
            if len(entry['_chunk_times']) > 256:
                entry['_chunk_times'] = entry['_chunk_times'][-256:]
            # Ollama only reports eval_count on the final response. Streaming
            # chunks are therefore deliberately labelled as an estimate.
            entry['estimated_tokens'] = entry['stream_chunks']


def _finish_live_generation(ident, status_code, final=None, error=None):
    final = final or {}
    now_perf = time.perf_counter()
    with _live_lock:
        entry = _active_generations.pop(ident, None)
        if not entry:
            return
        elapsed_s = max(0.0, now_perf - entry['_start_perf'])
        eval_count = int(final.get('eval_count') or 0)
        eval_duration_ns = int(final.get('eval_duration') or 0)
        exact_tps = eval_count / (eval_duration_ns / 1_000_000_000) if eval_duration_ns > 0 else None
        entry.update({
            'phase': 'error' if error or int(status_code) >= 400 else 'complete',
            'finished_at': datetime.now(timezone.utc).isoformat(),
            'elapsed_ms': elapsed_s * 1000,
            'prompt_tokens': int(final.get('prompt_eval_count') or 0),
            'output_tokens': eval_count,
            'exact_tps': exact_tps,
            'status_code': int(status_code),
            'error': str(error) if error else None,
        })
        entry.pop('_start_perf', None)
        entry.pop('_first_output_perf', None)
        entry.pop('_chunk_times', None)
        _recent_generations.appendleft(entry)


def _live_snapshot(entry, now_perf=None):
    now_perf = now_perf or time.perf_counter()
    out = {k: v for k, v in entry.items() if not k.startswith('_')}
    start_perf = entry.get('_start_perf')
    first_perf = entry.get('_first_output_perf')
    out['elapsed_ms'] = max(0.0, (now_perf - start_perf) * 1000) if start_perf else out.get('elapsed_ms')
    if first_perf:
        generation_s = max(0.001, now_perf - first_perf)
        out['estimated_tps'] = entry.get('estimated_tokens', 0) / generation_s
        recent = [t for t in (entry.get('_chunk_times') or []) if now_perf - t <= 3.0]
        if len(recent) >= 2:
            out['estimated_current_tps'] = (len(recent) - 1) / max(0.001, recent[-1] - recent[0])
        else:
            out['estimated_current_tps'] = out['estimated_tps']
    else:
        out['estimated_tps'] = None
        out['estimated_current_tps'] = None
    return out


def _version_tuple(version):
    parts = []
    for part in str(version or '').lstrip('vV').split('.'):
        digits = ''
        for ch in part:
            if not ch.isdigit():
                break
            digits += ch
        if not digits:
            break
        parts.append(int(digits))
        if len(parts) == 3:
            break
    return tuple((parts + [0, 0, 0])[:3])


_init_db()

# External inference connections are encrypted in the persistent data store.
# Only masked summaries are exposed to the browser.
from provider_store import ProviderError, ProviderStore
provider_store = ProviderStore(DATABASE_PATH)


def _external_model_options():
    return provider_store.model_options()


@app.get('/api/providers')
def provider_connections_api():
    return jsonify({'providers': provider_store.summaries()})


@app.put('/api/providers/<provider>')
def provider_connection_save_api(provider):
    user = _current_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'error': 'Administrator access is required'}), 403
    data = request.get_json(silent=True) or {}
    try:
        value = provider_store.save(provider, data.get('api_key'), data.get('models') or [])
    except ProviderError as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify({'provider': value, 'agent_reload': 'automatic'})


@app.delete('/api/providers/<provider>')
def provider_connection_delete_api(provider):
    user = _current_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'error': 'Administrator access is required'}), 403
    try:
        provider_store.delete(provider)
    except ProviderError as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify({'success': True})


@app.post('/api/providers/<provider>/test')
def provider_connection_test_api(provider):
    user = _current_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'error': 'Administrator access is required'}), 403
    data = request.get_json(silent=True) or {}
    try:
        models = provider_store.discover_models(provider, data.get('api_key'))
    except (ProviderError, requests.RequestException) as exc:
        return jsonify({'error': str(exc)}), 400
    return jsonify({'connected': True, 'models': models})


@app.get('/api/inference/models')
def inference_models_api():
    local = []
    try:
        response = requests.get(f'{OLLAMA_API}/api/tags', timeout=(3, 20))
        response.raise_for_status()
        for item in response.json().get('models') or []:
            name = str(item.get('name') or item.get('model') or '').strip()
            if name:
                local.append({'provider': 'ollama', 'provider_name': 'Ollama', 'name': name, 'value': f'ollama:{name}'})
    except (requests.RequestException, ValueError):
        pass
    return jsonify({'models': provider_store.model_options(local)})

# The coding-agent engine remains a private sidecar. This blueprint is the
# authenticated, ownership-checking boundary exposed to the Aperyn frontend.
from agent_gateway import create_agent_blueprint
app.register_blueprint(create_agent_blueprint(DATABASE_PATH, _current_user, OLLAMA_API, _external_model_options))


@app.route('/api/manager/status')
def manager_status():
    try:
        info = _safe_ollama_json('GET', '/api/version', timeout=4)
        version = info.get('version')
        native_verified = _version_tuple(version) >= (0, 32, 14)
        return jsonify({
            'online': True,
            'version': version,
            'url': OLLAMA_API,
            'proxy_port': PROXY_PUBLIC_PORT,
            'proxy_internal_url': PROXY_INTERNAL_URL,
            'mtp_native': {
                'verified': native_verified,
                'minimum_verified_version': '0.32.14',
                'summary': ('Native per-model embedded-MTP auto-detection is verified for this Ollama version.'
                            if native_verified else
                            'This build only verifies native per-model embedded-MTP behaviour on Ollama 0.32.14 or newer.'),
            },
        })
    except Exception:
        return jsonify({'online': False, 'version': None, 'url': OLLAMA_API, 'proxy_port': PROXY_PUBLIC_PORT, 'proxy_internal_url': PROXY_INTERNAL_URL, 'mtp_native': {'verified': False, 'minimum_verified_version': '0.32.14'}})




GLOBAL_PERF_KEYS = {
    'flash_attention':'OLLAMA_FLASH_ATTENTION',
    'kv_cache_type':'OLLAMA_KV_CACHE_TYPE',
    'num_parallel':'OLLAMA_NUM_PARALLEL',
    'max_loaded_models':'OLLAMA_MAX_LOADED_MODELS',
    'max_queue':'OLLAMA_MAX_QUEUE',
    'keep_alive':'OLLAMA_KEEP_ALIVE',
    'sched_spread':'OLLAMA_SCHED_SPREAD',
    'context_length':'OLLAMA_CONTEXT_LENGTH',
    'llama_fit':'LLAMA_ARG_FIT',
    'llama_fit_target_mib':'LLAMA_ARG_FIT_TARGET',
    'load_timeout':'OLLAMA_LOAD_TIMEOUT',
    'max_transfer_streams':'OLLAMA_MAX_TRANSFER_STREAMS',
}

def _helper_relay(method, path, payload=None, timeout=65):
    if not HELPER_TOKEN:
        raise RuntimeError('Host performance helper is not configured. Run make install-helper, then rebuild/restart Aperyn.')
    url=f"{PROXY_INTERNAL_URL}/__ollama_control/helper/{path.lstrip('/')}"
    r=requests.request(method,url,json=payload,headers={'X-Ollama-Control-Helper-Token':HELPER_TOKEN},timeout=timeout)
    try: data=r.json()
    except Exception: data={'error':r.text or f'helper returned {r.status_code}'}
    if not r.ok: raise RuntimeError(data.get('error') or f'helper returned {r.status_code}')
    return data

def _global_perf_from_env(env):
    env=env or {}
    def val(name):
        v=env.get(name)
        return '' if v is None else str(v)
    def boolval(name):
        v=val(name).lower()
        if v in ('true','1','on','yes'): return 'on'
        if v in ('false','0','off','no'): return 'off'
        return 'default'
    overhead=val('OLLAMA_GPU_OVERHEAD')
    try: overhead_mib=round(int(overhead)/(1024*1024),3) if overhead else None
    except Exception: overhead_mib=None
    return {
        'flash_attention':boolval('OLLAMA_FLASH_ATTENTION'),
        'kv_cache_type':val('OLLAMA_KV_CACHE_TYPE') or 'default',
        'num_parallel':val('OLLAMA_NUM_PARALLEL'),
        'max_loaded_models':val('OLLAMA_MAX_LOADED_MODELS'),
        'max_queue':val('OLLAMA_MAX_QUEUE'),
        'keep_alive':val('OLLAMA_KEEP_ALIVE'),
        'sched_spread':boolval('OLLAMA_SCHED_SPREAD'),
        'gpu_overhead_mib':overhead_mib,
        'context_length':val('OLLAMA_CONTEXT_LENGTH'),
        'llama_fit':boolval('LLAMA_ARG_FIT'),
        'llama_fit_target_mib':val('LLAMA_ARG_FIT_TARGET'),
        'load_timeout':val('OLLAMA_LOAD_TIMEOUT'),
        'max_transfer_streams':val('OLLAMA_MAX_TRANSFER_STREAMS'),
    }

def _global_perf_payload(data):
    out={}
    def tri(ui, env):
        v=str(data.get(ui,'default') or 'default').strip().lower()
        if v=='default': out[env]=None
        elif v in ('on','true','1'): out[env]='true' if env!='LLAMA_ARG_FIT' else 'on'
        elif v in ('off','false','0'): out[env]='false' if env!='LLAMA_ARG_FIT' else 'off'
        else: raise ValueError(f'{ui} must be default, on, or off')
    tri('flash_attention','OLLAMA_FLASH_ATTENTION')
    tri('sched_spread','OLLAMA_SCHED_SPREAD')
    tri('llama_fit','LLAMA_ARG_FIT')
    kv=str(data.get('kv_cache_type','default') or 'default').strip().lower()
    if kv=='default': out['OLLAMA_KV_CACHE_TYPE']=None
    elif kv in ('f16','q8_0','q4_0'): out['OLLAMA_KV_CACHE_TYPE']=kv
    else: raise ValueError('kv_cache_type must be default, f16, q8_0, or q4_0')
    ranges={
      'num_parallel':('OLLAMA_NUM_PARALLEL',1,256),
      'max_loaded_models':('OLLAMA_MAX_LOADED_MODELS',0,256),
      'max_queue':('OLLAMA_MAX_QUEUE',1,65536),
      'context_length':('OLLAMA_CONTEXT_LENGTH',0,16777216),
      'llama_fit_target_mib':('LLAMA_ARG_FIT_TARGET',0,1048576),
      'max_transfer_streams':('OLLAMA_MAX_TRANSFER_STREAMS',1,256),
    }
    for ui,(env,lo,hi) in ranges.items():
        raw=data.get(ui,'')
        if raw is None or str(raw).strip()=='': out[env]=None; continue
        try: n=int(raw)
        except Exception: raise ValueError(f'{ui} must be an integer')
        if n<lo or n>hi: raise ValueError(f'{ui} must be between {lo} and {hi}')
        out[env]=str(n)
    raw=data.get('gpu_overhead_mib','')
    if raw is None or str(raw).strip()=='': out['OLLAMA_GPU_OVERHEAD']=None
    else:
        try: mib=float(raw)
        except Exception: raise ValueError('gpu_overhead_mib must be a number')
        if mib<0 or mib>1048576: raise ValueError('gpu_overhead_mib is out of range')
        out['OLLAMA_GPU_OVERHEAD']=str(int(mib*1024*1024))
    for ui,env in [('keep_alive','OLLAMA_KEEP_ALIVE'),('load_timeout','OLLAMA_LOAD_TIMEOUT')]:
        raw=str(data.get(ui,'') or '').strip()
        if not raw: out[env]=None
        elif re.fullmatch(r'-?\d+(?:\.\d+)?(?:ns|us|µs|ms|s|m|h)?',raw): out[env]=raw
        else: raise ValueError(f'{ui} must be a duration such as 5m, 1h, 0, or -1')
    return out

@app.route('/api/performance/global', methods=['GET','POST'])
def global_performance_api():
    remote_host = _active_managed_host()
    if remote_host:
        if request.method == 'GET':
            snapshot = _host_summary(remote_host).get('snapshot') or {}
            helper = snapshot.get('helper') if isinstance(snapshot.get('helper'), dict) else {}
            state = 'connected' if remote_host.get('state') == 'connected' and helper else ('unreachable' if remote_host.get('connector_token_hash') else 'not_installed')
            effective_env = helper.get('effective_env') if isinstance(helper.get('effective_env'), dict) else {}
            return jsonify({'helper_ready': state == 'connected', 'helper_state': state, 'remote_host': _host_summary(remote_host),
                            'install_command': 'Install and pair the Aperyn Host Connector on this Ollama host.',
                            'service_state': helper.get('service_state', 'unknown'), 'effective': _global_perf_from_env(effective_env),
                            'managed': helper.get('managed') or {}, 'conflicting_mtp_globals': helper.get('conflicting_mtp_globals') or {}}), 200 if state == 'connected' else 503
        data = request.get_json(silent=True) or {}
        try:
            settings = _global_perf_payload(data)
            result = _remote_host_action(remote_host['id'], 'helper.apply', {'settings': settings, 'clear_mtp_globals': bool(data.get('clear_mtp_globals'))})
        except ValueError as exc:
            return jsonify({'error': str(exc)}), 400
        except Exception as exc:
            return jsonify({'error': str(exc)}), 503
        if settings.get('OLLAMA_KV_CACHE_TYPE'):
            _set_setting('estimator_kv_cache_type', 'auto')
        return jsonify({'ok': True, 'message': f"Ollama performance settings applied on {remote_host['name']}.",
                        'service_state': result.get('service_state'), 'effective': _global_perf_from_env(result.get('effective_env') or {}),
                        'conflicting_mtp_globals': result.get('conflicting_mtp_globals') or {}})
    seed=_seed_hardware(); seed_env=seed.get('ollama_env') if isinstance(seed.get('ollama_env'),dict) else {}
    if request.method=='GET':
        try:
            probe = requests.get(f"{PROXY_INTERNAL_URL}/__ollama_control/helper-state", timeout=4)
            probe_data = probe.json()
        except Exception as exc:
            probe_data = {'state':'unreachable','detail':str(exc)}
        if probe_data.get('state') != 'connected':
            return jsonify({'helper_ready':False,'helper_state':probe_data.get('state','unreachable'),'install_command':'./ollama-control up','detail':probe_data.get('detail'),'service_state':'unknown','effective':_global_perf_from_env(seed_env),'managed':{},'conflicting_mtp_globals':{k:seed_env.get(k) for k in ('LLAMA_ARG_SPEC_TYPE','LLAMA_ARG_SPEC_DRAFT_N_MAX') if seed_env.get(k)}}), 503 if probe_data.get('state') not in ('not_installed','authentication_mismatch') else 200
        try:
            status=_helper_relay('GET','status')
            return jsonify({'helper_ready':True,'helper_state':'connected','install_command':'./ollama-control up','service_state':status.get('service_state'),'effective':_global_perf_from_env(status.get('effective_env')),'managed':status.get('managed') or {},'conflicting_mtp_globals':status.get('conflicting_mtp_globals') or {},'dropin':status.get('dropin')})
        except Exception as exc:
            state='authentication_mismatch' if '401' in str(exc) or 'unauthorized' in str(exc).lower() else 'unreachable'
            return jsonify({'helper_ready':False,'helper_state':state,'install_command':'./ollama-control up','error':str(exc),'effective':_global_perf_from_env(seed_env),'managed':{},'conflicting_mtp_globals':{}}),503
    data=request.get_json(silent=True) or {}
    try: settings=_global_perf_payload(data)
    except ValueError as exc: return jsonify({'error':str(exc)}),400
    try:
        result=_helper_relay('POST','apply',{'settings':settings,'clear_mtp_globals':bool(data.get('clear_mtp_globals'))})
    except Exception as exc:
        return jsonify({'error':str(exc),'install_command':'make install-helper'}),503
    # Refresh the local host snapshot on the next preflight; meanwhile make estimator KV follow the actual setting immediately.
    if settings.get('OLLAMA_KV_CACHE_TYPE'):
        _set_setting('estimator_kv_cache_type','auto')
    try:
        seed_path=Path(os.environ.get('HARDWARE_SEED_PATH','/data/hardware.json'))
        seed_data={}
        if seed_path.exists(): seed_data=json.loads(seed_path.read_text())
        if not isinstance(seed_data,dict): seed_data={}
        seed_data['ollama_env']={k:v for k,v in (result.get('effective_env') or {}).items() if v is not None}
        seed_path.parent.mkdir(parents=True,exist_ok=True); seed_path.write_text(json.dumps(seed_data,indent=2))
    except Exception:
        pass
    return jsonify({'ok':True,'message':'Ollama performance settings applied and service restarted.','service_state':result.get('service_state'),'effective':_global_perf_from_env(result.get('effective_env')),'conflicting_mtp_globals':result.get('conflicting_mtp_globals') or {}})


@app.route('/api/settings', methods=['GET', 'POST'])
def app_settings_api():
    default_upstream = os.environ.get('DEFAULT_OLLAMA_UPSTREAM', 'http://127.0.0.1:11434').rstrip('/')
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        if 'ollama_endpoint' in data:
            endpoint = str(data.get('ollama_endpoint') or '').strip().rstrip('/')
            if endpoint and not re.fullmatch(r'https?://[^\s/]+(?::\d+)?(?:/[^\s]*)?', endpoint, re.I):
                return jsonify({'error': 'Enter a valid http:// or https:// Ollama endpoint, or leave it blank for the default.'}), 400
            _set_setting('ollama_endpoint', endpoint)
            _set_setting('active_managed_host_id', '')
        if 'default_chat_model' in data:
            default_chat_model = str(data.get('default_chat_model') or '').strip()
            _set_setting('default_chat_model', default_chat_model)
        if 'assistant_name' in data:
            assistant_name = str(data.get('assistant_name') or '').strip()
            if len(assistant_name) > 40:
                return jsonify({'error': 'AI / assistant name must be 40 characters or fewer.'}), 400
            _set_setting('assistant_name', assistant_name)
        if 'theme' in data:
            theme = data.get('theme')
            if not isinstance(theme, dict):
                return jsonify({'error': 'theme must be an object'}), 400
            theme_keys = {
                'accent': 'theme_accent',
                'background': 'theme_background',
                'panel': 'theme_panel',
                'panel2': 'theme_panel2',
            }
            for ui_key, db_key in theme_keys.items():
                if ui_key not in theme:
                    continue
                value = str(theme.get(ui_key) or '').strip()
                if value and not re.fullmatch(r'#[0-9a-fA-F]{6}', value):
                    return jsonify({'error': f'{ui_key} must be a 6-digit hex colour such as #78a9ff'}), 400
                _set_user_preference(_current_user()['id'], db_key, value.lower())
            if 'glass' in theme:
                glass = str(theme.get('glass') or 'subtle').strip().lower()
                if glass not in ('off', 'subtle', 'full'):
                    return jsonify({'error': 'glass must be off, subtle, or full'}), 400
                _set_user_preference(_current_user()['id'], 'theme_glass', glass)
        if 'estimator_kv_cache_type' in data:
            kv = _normalize_kv_cache_type(data.get('estimator_kv_cache_type') or 'auto', allow_auto=True)
            if kv not in ('auto','f16','q8_0','q4_0'):
                return jsonify({'error':'Estimator KV cache type must be auto, f16, q8_0, or q4_0'}), 400
            _set_setting('estimator_kv_cache_type', kv)
        if 'agent_context_limit' in data:
            try:
                agent_context_limit = int(data.get('agent_context_limit') or 98304)
            except (TypeError, ValueError):
                return jsonify({'error': 'Agent context target must be a whole number'}), 400
            if agent_context_limit < 4096 or agent_context_limit > 1048576:
                return jsonify({'error': 'Agent context target must be between 4,096 and 1,048,576'}), 400
            _set_setting('agent_context_limit', str(agent_context_limit))
        for key in ('gpu_vram_override_gb','system_ram_override_gb'):
            if key in data:
                try: value = max(0.0, float(data.get(key) or 0))
                except (TypeError, ValueError): return jsonify({'error': f'{key} must be a number'}), 400
                _set_setting(key, str(value))
    configured = _configured_upstream()
    seed = _seed_hardware()
    service_env = seed.get('ollama_env') if isinstance(seed.get('ollama_env'), dict) else {}
    service_kv = _normalize_kv_cache_type(service_env.get('OLLAMA_KV_CACHE_TYPE'))
    estimator_kv = str(_get_setting('estimator_kv_cache_type','auto') or 'auto').strip().lower()
    effective_kv = service_kv if estimator_kv == 'auto' and service_kv else ('f16' if estimator_kv == 'auto' else estimator_kv)
    active_host = _active_managed_host()
    with _db() as conn:
        managed_rows = conn.execute('SELECT * FROM managed_hosts ORDER BY name COLLATE NOCASE').fetchall()
    return jsonify({
        'ollama_endpoint': configured,
        'default_ollama_endpoint': default_upstream,
        'effective_ollama_endpoint': configured or default_upstream,
        'active_host_id': active_host['id'] if active_host else 'local',
        'active_host_name': active_host['name'] if active_host else 'Local / default host',
        'managed_hosts': [_host_summary(row) for row in managed_rows],
        'proxy_port': PROXY_PUBLIC_PORT,
        'default_chat_model': str(_get_setting('default_chat_model','') or '').strip(),
        'assistant_name': (lambda value: 'Nym' if value in ('', 'AI') else value)(str(_get_setting('assistant_name','') or '').strip()),
        'theme': {
            'accent': str(_get_user_preference(_current_user()['id'], 'theme_accent','') or '').strip(),
            'background': str(_get_user_preference(_current_user()['id'], 'theme_background','') or '').strip(),
            'panel': str(_get_user_preference(_current_user()['id'], 'theme_panel','') or '').strip(),
            'panel2': str(_get_user_preference(_current_user()['id'], 'theme_panel2','') or '').strip(),
            'glass': str(_get_user_preference(_current_user()['id'], 'theme_glass','subtle') or 'subtle').strip(),
        },
        'estimator_kv_cache_type': estimator_kv,
        'detected_service_kv_cache_type': service_kv or None,
        'effective_estimator_kv_cache_type': effective_kv,
        'gpu_vram_override_gb': float(_get_setting('gpu_vram_override_gb','0') or 0),
        'system_ram_override_gb': float(_get_setting('system_ram_override_gb','0') or 0),
        'agent_context_limit': int(_get_setting('agent_context_limit','98304') or 98304),
    })


def _read_meminfo():
    path = '/host/proc/meminfo' if os.path.exists('/host/proc/meminfo') else '/proc/meminfo'
    out = {}
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
            for line in fh:
                if ':' not in line:
                    continue
                key, rest = line.split(':', 1)
                parts = rest.strip().split()
                if not parts:
                    continue
                try:
                    value = int(parts[0])
                except ValueError:
                    continue
                if len(parts) > 1 and parts[1].lower() == 'kb':
                    value *= 1024
                out[key] = value
    except OSError:
        pass
    return out


def _host_cpu_count():
    path = '/host/proc/cpuinfo' if os.path.exists('/host/proc/cpuinfo') else '/proc/cpuinfo'
    try:
        with open(path, 'r', encoding='utf-8', errors='ignore') as fh:
            count = sum(1 for line in fh if line.lower().startswith('processor'))
        if count:
            return count
    except OSError:
        pass
    return os.cpu_count() or 0


def _seed_hardware():
    path = os.environ.get('HARDWARE_SEED_PATH', '/data/hardware.json')
    try:
        with open(path, 'r', encoding='utf-8') as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else {}
    except (OSError, ValueError, TypeError):
        return {}


def _nvidia_smi_gpus():
    exe = shutil.which('nvidia-smi')
    if not exe:
        return []
    try:
        proc = subprocess.run(
            [exe, '--query-gpu=name,memory.total,memory.used,utilization.gpu,temperature.gpu', '--format=csv,noheader,nounits'],
            stdout=subprocess.PIPE, stderr=subprocess.DEVNULL, text=True, timeout=2, check=True,
        )
    except Exception:
        return []
    rows = []
    for line in proc.stdout.splitlines():
        parts = [x.strip() for x in line.split(',')]
        if len(parts) < 5:
            continue
        try:
            total = int(float(parts[1])) * 1024 * 1024
            used = int(float(parts[2])) * 1024 * 1024
            util = float(parts[3])
            temp = float(parts[4])
        except ValueError:
            continue
        rows.append({'name': parts[0], 'total_bytes': total, 'used_bytes': used, 'utilization': util, 'temperature_c': temp, 'source': 'nvidia-smi'})
    return rows


def _amd_sysfs_gpus():
    rows = []
    roots = ['/host/sys/class/drm', '/sys/class/drm']
    seen = set()
    for root in roots:
        if not os.path.isdir(root):
            continue
        for card in glob.glob(os.path.join(root, 'card[0-9]*')):
            device = os.path.join(card, 'device')
            total_path = os.path.join(device, 'mem_info_vram_total')
            if not os.path.isfile(total_path):
                continue
            try:
                total = int(Path(total_path).read_text().strip())
                used_path = os.path.join(device, 'mem_info_vram_used')
                used = int(Path(used_path).read_text().strip()) if os.path.isfile(used_path) else 0
            except Exception:
                continue
            if total <= 0 or total in seen:
                continue
            seen.add(total)
            name = 'AMD GPU'
            product_path = os.path.join(device, 'product_name')
            if os.path.isfile(product_path):
                try: name = Path(product_path).read_text().strip() or name
                except Exception: pass
            rows.append({'name': name, 'total_bytes': total, 'used_bytes': used, 'utilization': None, 'temperature_c': None, 'source': 'sysfs'})
        if rows:
            break
    return rows


def _hardware_snapshot():
    mem = _read_meminfo()
    seed = _seed_hardware()
    # Prefer live host-side telemetry from the localhost helper. The WebUI
    # container intentionally has no privileged GPU access, while the helper can
    # query stable driver interfaces (nvidia-smi / amdgpu sysfs) on the host.
    helper_gpu = None
    if HELPER_TOKEN:
        try:
            helper_gpu = _helper_relay('GET', 'gpu', timeout=6)
        except Exception:
            helper_gpu = None
    gpus = []
    if isinstance(helper_gpu, dict) and isinstance(helper_gpu.get('devices'), list):
        gpus = [x for x in helper_gpu.get('devices') if isinstance(x, dict)]
    if not gpus:
        gpus = _nvidia_smi_gpus() or _amd_sysfs_gpus()
    if not gpus and isinstance(seed.get('gpus'), list):
        for item in seed['gpus']:
            if not isinstance(item, dict):
                continue
            try: total = int(item.get('total_bytes') or 0)
            except (TypeError, ValueError): total = 0
            if total > 0:
                gpus.append({
                    'name': str(item.get('name') or 'GPU'), 'total_bytes': total,
                    'used_bytes': None, 'utilization': None, 'temperature_c': None,
                    'source': 'host preflight',
                })
    try:
        ps = _safe_ollama_json('GET', '/api/ps', timeout=4).get('models') or []
    except Exception:
        ps = []
    ollama_vram = sum(int(m.get('size_vram') or 0) for m in ps if isinstance(m, dict))
    total_gpu = sum(int(g.get('total_bytes') or 0) for g in gpus)
    live_gpu_used = sum(int(g.get('used_bytes') or 0) for g in gpus if g.get('used_bytes') is not None)
    gpu_used = live_gpu_used if any(g.get('used_bytes') is not None for g in gpus) else ollama_vram
    total_ram = int(mem.get('MemTotal') or seed.get('system_ram_bytes') or 0)
    available_ram = int(mem.get('MemAvailable') or mem.get('MemFree') or 0)
    return {
        'system': {
            'total_bytes': total_ram,
            'available_bytes': available_ram,
            'used_bytes': max(0, total_ram - available_ram) if total_ram else 0,
            'cpu_count': int(seed.get('cpu_count') or _host_cpu_count()),
            'source': 'host /proc' if os.path.exists('/host/proc/meminfo') else 'container /proc',
        },
        'gpu': {
            'devices': gpus,
            'total_bytes': total_gpu,
            'used_bytes': gpu_used,
            'free_bytes': max(0, total_gpu - gpu_used) if total_gpu else 0,
            'ollama_used_bytes': ollama_vram,
            'detected': bool(total_gpu),
            'source': (gpus[0].get('source') if gpus else 'Ollama /api/ps only'),
            'utilization': (helper_gpu.get('utilization') if isinstance(helper_gpu, dict) else None),
            'nvtop_installed': (helper_gpu.get('nvtop_installed') if isinstance(helper_gpu, dict) else None),
        },
        'loaded_models': ps,
        'captured_at': _utc_now(),
        'service': {
            'ollama_env': seed.get('ollama_env') if isinstance(seed.get('ollama_env'), dict) else {},
            'kv_cache_type': ((seed.get('ollama_env') or {}).get('OLLAMA_KV_CACHE_TYPE') if isinstance(seed.get('ollama_env'), dict) else None),
            'flash_attention': ((seed.get('ollama_env') or {}).get('OLLAMA_FLASH_ATTENTION') if isinstance(seed.get('ollama_env'), dict) else None),
        },
        'diagnostics': {
            'hardware_seed_loaded': bool(seed),
            'hardware_seed_path': os.environ.get('HARDWARE_SEED_PATH', '/data/hardware.json'),
            'host_proc_mounted': os.path.exists('/host/proc/meminfo'),
            'host_sys_mounted': os.path.exists('/host/sys'),
            'gpu_seed_count': len(seed.get('gpus') or []) if isinstance(seed.get('gpus'), list) else 0,
            'helper_gpu_live': bool(isinstance(helper_gpu, dict) and helper_gpu.get('detected')),
            'nvtop_installed': (helper_gpu.get('nvtop_installed') if isinstance(helper_gpu, dict) else None),
        },
    }


@app.get('/api/hardware')
def hardware_api():
    remote_host = _active_managed_host()
    if remote_host:
        snapshot = _host_summary(remote_host).get('snapshot') or {}
        gpu = snapshot.get('gpu') if isinstance(snapshot.get('gpu'), dict) else {}
        system = snapshot.get('system') if isinstance(snapshot.get('system'), dict) else {}
        devices = [item for item in (gpu.get('devices') or []) if isinstance(item, dict)]
        total = int(gpu.get('total_bytes') or sum(int(item.get('total_bytes') or 0) for item in devices))
        used = int(gpu.get('used_bytes') or sum(int(item.get('used_bytes') or 0) for item in devices if item.get('used_bytes') is not None))
        ram_total = int(system.get('total_bytes') or 0); ram_available = int(system.get('available_bytes') or 0)
        return jsonify({'system': {'total_bytes': ram_total, 'available_bytes': ram_available, 'used_bytes': max(0, ram_total-ram_available), 'cpu_count': int(system.get('cpu_count') or 0), 'source': f"managed host · {remote_host['name']}"},
                        'gpu': {'devices': devices, 'total_bytes': total, 'used_bytes': used, 'free_bytes': max(0,total-used), 'ollama_used_bytes': 0, 'detected': bool(total), 'source': f"managed host · {remote_host['name']}", 'utilization': gpu.get('utilization'), 'nvtop_installed': gpu.get('nvtop_installed')},
                        'loaded_models': [], 'captured_at': _utc_now(), 'service': {'ollama_env': ((snapshot.get('helper') or {}).get('effective_env') or {}), 'kv_cache_type': ((snapshot.get('helper') or {}).get('effective_env') or {}).get('OLLAMA_KV_CACHE_TYPE')},
                        'diagnostics': {'managed_host': _host_summary(remote_host), 'helper_gpu_live': bool(gpu.get('detected'))}})
    return jsonify(_hardware_snapshot())


@app.route('/api/manager/runtime', methods=['GET', 'POST'])
def manager_runtime():
    try:
        if request.method == 'GET':
            data = _safe_ollama_json('GET', '/api/ps', timeout=10)
            return jsonify({'models': data.get('models') or []})
        data = request.get_json(silent=True) or {}
        model = str(data.get('model') or '').strip()
        action = str(data.get('action') or '').strip().lower()
        if not model or action not in ('load', 'unload'):
            return jsonify({'error': 'model and action=load|unload are required'}), 400
        payload = {'model': model, 'prompt': '', 'stream': False}
        if action == 'load':
            payload['keep_alive'] = -1
        else:
            payload['keep_alive'] = 0
        result = _safe_ollama_json('POST', '/api/generate', json=payload, timeout=3600)
        return jsonify({'success': True, 'action': action, 'model': model, 'result': result})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/api/manager/parameters')
def manager_parameters():
    return jsonify({'parameters': PARAMETER_DEFINITIONS})


@app.route('/api/manager/model-info', methods=['POST'])
def manager_model_info():
    try:
        model = (request.json or {}).get('model')
        if not model:
            return jsonify({'error': 'Missing model name'}), 400
        data = _safe_ollama_json('POST', '/api/show', json={'model': model, 'verbose': True}, timeout=60)
        parsed = _parse_parameters(data.get('parameters', ''))
        data['parsed_parameters'] = parsed
        data['mtp_capability'] = _mtp_capability(data)
        data['manager_preferences'] = _effective_mtp_pref(model, parsed)
        data['thinking_detected'] = _thinking_profile(data, model)
        data['thinking_override'] = _thinking_override(model) or {'enabled': False}
        data['thinking_profile'] = _effective_thinking_profile(data, model)
        data['model_limits'] = _model_limits(data)
        try:
            row = _installed_model_row(model) or {}
            data['memory_estimate'] = _model_memory_estimate(data, row.get('size') or 0, parsed.get('num_ctx'), parsed.get('num_gpu', -1), model_name=model)
        except Exception:
            data['memory_estimate'] = None
        return jsonify(data)
    except requests.HTTPError as e:
        detail = e.response.text if e.response is not None else str(e)
        return jsonify({'error': detail}), getattr(e.response, 'status_code', 500)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/manager/modify-model', methods=['POST'])
def manager_modify_model():
    data = request.json or {}
    source = str(data.get('source') or '').strip()
    target = str(data.get('target') or '').strip()
    parameters = dict(data.get('parameters') or {})
    mtp = data.get('mtp') or {}
    reasoning_override = data.get('reasoning_override', None)
    if not source or not target:
        return jsonify({'error': 'Source and target model names are required'}), 400
    allowed = {p['name'] for p in PARAMETER_DEFINITIONS}
    unknown = sorted(set(parameters) - allowed)
    if unknown:
        return jsonify({'error': f'Unsupported parameters: {", ".join(unknown)}'}), 400

    definitions = {p['name']: p for p in PARAMETER_DEFINITIONS}
    for key, value in list(parameters.items()):
        definition = definitions.get(key) or {}
        if definition.get('type') not in ('int', 'float') or key in ('num_gpu', 'num_ctx', 'draft_num_predict'):
            continue
        try:
            numeric = int(value) if definition.get('type') == 'int' else float(value)
        except (TypeError, ValueError):
            return jsonify({'error': f'{key} must be a valid {definition.get("type") or "number"}'}), 400
        min_value, max_value = definition.get('min'), definition.get('max')
        if min_value is not None and numeric < min_value:
            return jsonify({'error': f'{key} cannot be below {min_value}'}), 400
        if max_value is not None and numeric > max_value:
            return jsonify({'error': f'{key} cannot be above {max_value}'}), 400
        parameters[key] = numeric

    mtp_enabled = bool(mtp.get('enabled', False))
    try:
        mtp_depth = int(mtp.get('draft_n_max', 2))
    except (TypeError, ValueError):
        return jsonify({'error': 'MTP draft depth must be an integer'}), 400
    if mtp_depth < 1 or mtp_depth > 8:
        return jsonify({'error': 'MTP draft depth must be between 1 and 8'}), 400

    if reasoning_override is not None and not isinstance(reasoning_override, dict):
        return jsonify({'error': 'reasoning_override must be an object'}), 400
    if isinstance(reasoning_override, dict) and reasoning_override.get('enabled'):
        manual_levels = []
        for level in (reasoning_override.get('levels') or []):
            level = str(level).strip().lower()
            if level in ('off','on','low','medium','high','xhigh','max') and level not in manual_levels:
                manual_levels.append(level)
        if not manual_levels:
            return jsonify({'error': 'Manual thinking override requires at least one enabled level.'}), 400
        manual_method = str(reasoning_override.get('method') or 'native').strip().lower()
        if manual_method not in ('native','xhigh-default','system-directive'):
            return jsonify({'error': 'Unsupported thinking override control method.'}), 400
        manual_default = str(reasoning_override.get('default') or '').strip().lower()
        if manual_default not in manual_levels:
            return jsonify({'error': 'Default thinking level must be one of the enabled levels.'}), 400

    capability = {'status': 'unknown', 'summary': 'Capability was not checked.'}
    show = None
    limits = {}
    needs_show = mtp_enabled or 'num_gpu' in parameters or 'num_ctx' in parameters
    if needs_show:
        try:
            show = _safe_ollama_json('POST', '/api/show', json={'model': source, 'verbose': True}, timeout=60)
            limits = _model_limits(show)
        except Exception as e:
            if mtp_enabled:
                capability = {'status': 'unknown', 'summary': f'Could not verify MTP capability: {e}', 'evidence': []}

    if 'num_gpu' in parameters:
        try:
            requested_gpu = int(parameters['num_gpu'])
        except (TypeError, ValueError):
            return jsonify({'error': 'num_gpu must be an integer'}), 400
        gpu_max = ((limits.get('num_gpu') or {}).get('max'))
        if gpu_max is not None and requested_gpu > int(gpu_max):
            return jsonify({'error': f'num_gpu is above this model maximum. Maximum is {int(gpu_max)} layers for full offload.'}), 400
        if requested_gpu < -1:
            return jsonify({'error': 'num_gpu must be -1 (auto), 0 (CPU), or a positive layer count.'}), 400

    if 'num_ctx' in parameters:
        try:
            requested_ctx = int(parameters['num_ctx'])
        except (TypeError, ValueError):
            return jsonify({'error': 'num_ctx must be an integer'}), 400
        if requested_ctx < 512:
            return jsonify({'error': 'num_ctx cannot be below 512 tokens.'}), 400
        ctx_max = ((limits.get('num_ctx') or {}).get('max'))
        if ctx_max is not None and requested_ctx > int(ctx_max):
            return jsonify({'error': f'num_ctx is above this model training context. Maximum is {int(ctx_max):,} tokens.'}), 400

    if mtp_enabled:
        if show is not None:
            capability = _mtp_capability(show)
        if capability.get('status') == 'unsupported':
            return jsonify({
                'error': 'MTP was blocked: this model has verbose tensor metadata but no embedded MTP / NextN layers were detected.',
                'mtp_capability': capability,
            }), 400

    # Ollama 0.32.14+ auto-detects embedded MTP in the GGUF. When that model
    # is loaded and DraftNumPredict > 0, Ollama launches llama-server with
    # --spec-type draft-mtp and --spec-draft-n-max N (plus backend sampling).
    # A zero value makes Ollama omit those draft arguments. Do not force the
    # equivalent LLAMA_ARG_* environment variables globally on the host service.
    parameters['draft_num_predict'] = mtp_depth if mtp_enabled else 0

    def generate():
        temp_source = None
        saved_ok = False
        try:
            create_from = source
            if mtp_enabled and capability.get('status') == 'unknown':
                yield json.dumps({'status': 'MTP capability could not be proven; saving with a cautious warning.', 'warning': capability.get('summary')}) + '\n'
            if target == source:
                temp_source = f'ollama-control-temp-{uuid.uuid4().hex[:10]}'
                yield json.dumps({'status': 'Creating a temporary source alias…'}) + '\n'
                copy_resp = requests.post(f'{OLLAMA_API}/api/copy', json={'source': source, 'destination': temp_source}, timeout=60)
                if not copy_resp.ok:
                    yield json.dumps({'error': copy_resp.text or 'Could not prepare model for replacement'}) + '\n'
                    return
                create_from = temp_source

            yield json.dumps({'status': f'Saving {target} with {len(parameters)} enabled parameter overrides…'}) + '\n'
            resp = requests.post(f'{OLLAMA_API}/api/create', json={'model': target, 'from': create_from, 'parameters': parameters, 'stream': True}, stream=True, timeout=3600)
            if not resp.ok:
                yield json.dumps({'error': resp.text or f'Ollama returned HTTP {resp.status_code}'}) + '\n'
                return
            for line in resp.iter_lines():
                if line:
                    try:
                        obj = json.loads(line)
                        if obj.get('error'):
                            yield line.decode('utf-8') + '\n'
                            return
                    except Exception:
                        pass
                    yield line.decode('utf-8') + '\n'
            saved_ok = True
        except Exception as e:
            yield json.dumps({'error': str(e)}) + '\n'
        finally:
            if saved_ok:
                try:
                    _save_model_pref(target, mtp_enabled, mtp_depth)
                except Exception as e:
                    app.logger.warning('Could not store MTP manager preference: %s', e)
                if reasoning_override is not None:
                    try:
                        _save_thinking_override(target, reasoning_override)
                    except Exception as e:
                        app.logger.warning('Could not store manual thinking override: %s', e)
            if temp_source:
                try:
                    requests.delete(f'{OLLAMA_API}/api/delete', json={'model': temp_source}, timeout=30)
                except Exception:
                    pass
    return Response(generate(), mimetype='application/x-ndjson')


@app.route('/api/manager/dashboard')
def manager_dashboard():
    since = datetime.now(timezone.utc) - timedelta(hours=24)
    since_iso = since.isoformat()
    with _db() as conn:
        summary = conn.execute("""SELECT COUNT(*) requests, COALESCE(SUM(prompt_tokens),0) prompt_tokens, COALESCE(SUM(output_tokens),0) output_tokens, AVG(latency_ms) avg_latency_ms, COALESCE(SUM(eval_duration_ns),0) eval_duration_ns FROM request_log WHERE created_at >= ?""", (since_iso,)).fetchone()
        by_model = [dict(r) for r in conn.execute("""SELECT COALESCE(model,'unknown') model, COUNT(*) requests, COALESCE(SUM(prompt_tokens),0) prompt_tokens, COALESCE(SUM(output_tokens),0) output_tokens FROM request_log WHERE created_at >= ? GROUP BY COALESCE(model,'unknown') ORDER BY requests DESC LIMIT 8""", (since_iso,))]
        recent = [dict(r) for r in conn.execute("""SELECT id, created_at, endpoint, model, status_code, latency_ms, prompt_tokens, output_tokens, eval_duration_ns, client_name FROM request_log ORDER BY id DESC LIMIT 12""")]
        rows = [dict(r) for r in conn.execute('SELECT created_at FROM request_log WHERE created_at >= ?', (since_iso,))]
    eval_s = (summary['eval_duration_ns'] or 0) / 1_000_000_000
    tps = (summary['output_tokens'] or 0) / eval_s if eval_s > 0 else None
    buckets=[]
    now=datetime.now(timezone.utc).replace(minute=0,second=0,microsecond=0)
    counts={}
    for r in rows:
        try:
            dt=datetime.fromisoformat(r['created_at']).astimezone(timezone.utc).replace(minute=0,second=0,microsecond=0)
            counts[dt]=counts.get(dt,0)+1
        except Exception:
            pass
    for i in range(23,-1,-1):
        dt=now-timedelta(hours=i)
        buckets.append({'label':dt.strftime('%H:%M'),'requests':counts.get(dt,0)})
    try:
        running = _safe_ollama_json('GET', '/api/ps', timeout=5).get('models', [])
    except Exception:
        running = []
    return jsonify({'summary':{'requests':summary['requests'] or 0,'prompt_tokens':summary['prompt_tokens'] or 0,'output_tokens':summary['output_tokens'] or 0,'avg_latency_ms':summary['avg_latency_ms'],'tokens_per_second':tps},'timeline':buckets,'by_model':by_model,'recent':recent,'running':running})


@app.delete('/api/manager/telemetry')
def manager_clear_telemetry():
    user = _current_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'error': 'Administrator access is required'}), 403
    with _db() as conn:
        deleted = int(conn.execute('SELECT COUNT(*) FROM request_log').fetchone()[0])
        conn.execute('DELETE FROM request_log')
    with _live_lock:
        local_recent = len(_recent_generations)
        _recent_generations.clear()
    proxy_recent_cleared = False
    try:
        response = requests.post(
            f'{PROXY_INTERNAL_URL}/__ollama_control/live/clear',
            headers={'X-Aperyn-Internal-Token': _telemetry_clear_identity()},
            timeout=2,
        )
        response.raise_for_status()
        proxy_recent_cleared = bool((response.json() or {}).get('cleared'))
    except Exception as exc:
        app.logger.warning('Persistent telemetry cleared, but proxy recent summaries could not be cleared: %s', exc)
    return jsonify({'success': True, 'deleted_requests': deleted, 'cleared_local_recent': local_recent,
                    'proxy_recent_cleared': proxy_recent_cleared,
                    'active_generations_preserved': True})


@app.route('/api/manager/live')
def manager_live():
    now_perf = time.perf_counter()
    with _live_lock:
        active = [_live_snapshot(dict(entry), now_perf) for entry in _active_generations.values()]
        recent = [dict(entry) for entry in list(_recent_generations)[:6]]

    proxy_online = False
    proxy_error = None
    try:
        proxy_response = requests.get(f'{PROXY_INTERNAL_URL}/__ollama_control/live', timeout=0.8)
        proxy_response.raise_for_status()
        proxy_data = proxy_response.json()
        active.extend(proxy_data.get('active') or [])
        recent.extend(proxy_data.get('recent') or [])
        proxy_online = True
    except Exception as exc:
        proxy_error = str(exc)

    active.sort(key=lambda x: x.get('started_at') or '')
    recent.sort(key=lambda x: x.get('finished_at') or '', reverse=True)
    return jsonify({
        'active': active,
        'recent': recent[:8],
        'proxy_online': proxy_online,
        'proxy_error': proxy_error,
        'note': 'Live token count/TPS are estimated from streamed chunks until Ollama returns final eval_count and eval_duration; native Ollama completions use exact completed TPS.'
    })


@app.route('/ollama/<path:subpath>', methods=['GET','POST','PUT','PATCH','DELETE','OPTIONS'])
def observable_proxy(subpath):
    # Legacy compatibility proxy. The primary drop-in endpoint is the dedicated
    # proxy listener on host port 11435, where clients can use normal /api/* paths.
    target = f'{OLLAMA_API}/{subpath}'
    payload = request.get_json(silent=True) or {}
    model = payload.get('model')
    track = subpath in ('api/generate','api/chat')
    mtp_pref = _model_pref(model) if model else None

    # Apply the manager's per-model MTP preference only when the caller did not
    # explicitly provide draft_num_predict. This keeps API callers in control.
    if track and model and mtp_pref:
        options = payload.setdefault('options', {})
        if isinstance(options, dict) and 'draft_num_predict' not in options:
            options['draft_num_predict'] = mtp_pref['mtp_draft_n_max'] if mtp_pref['mtp_enabled'] else 0

    body = json.dumps(payload).encode('utf-8') if payload else request.get_data()
    headers = {k:v for k,v in request.headers if k.lower() not in {'host','content-length','connection'}}
    if payload:
        headers['Content-Type'] = 'application/json'

    start = time.perf_counter()
    client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)
    live_id = _start_live_generation(model, '/'+subpath, mtp_pref) if track else None
    try:
        upstream = requests.request(request.method, target, params=request.args, data=body, headers=headers, stream=True, timeout=3600)
    except Exception as e:
        if track:
            _log_request('/'+subpath, model, 502, (time.perf_counter()-start)*1000, {}, client_ip)
            _finish_live_generation(live_id, 502, {}, e)
        return jsonify({'error': str(e)}), 502

    content_type = upstream.headers.get('Content-Type', 'application/octet-stream')

    def stream_response():
        final = {}
        stream_error = None
        try:
            for line in upstream.iter_lines():
                if line:
                    if track:
                        try:
                            obj = json.loads(line)
                            _update_live_generation(live_id, obj)
                            if obj.get('done'):
                                final.update(obj)
                        except Exception:
                            pass
                    yield line + b'\n'
        except Exception as e:
            stream_error = e
            raise
        finally:
            if track:
                latency_ms = (time.perf_counter()-start)*1000
                _log_request('/'+subpath, model, upstream.status_code, latency_ms, final, client_ip)
                _finish_live_generation(live_id, upstream.status_code, final, stream_error)
            upstream.close()
    return Response(stream_response(), status=upstream.status_code, content_type=content_type)


# --- v1.5 conversation workspace + model discovery --------------------------

HF_TOKEN = os.environ.get('HF_TOKEN', '').strip()
_download_lock = threading.Lock()
_download_jobs = {}


def _utc_now():
    return datetime.now(timezone.utc).isoformat()


def _chat_title(text):
    clean = re.sub(r'\s+', ' ', (text or '').strip())
    return (clean[:52] + ('…' if len(clean) > 52 else '')) or 'New chat'


@app.get('/api/chats')
def list_chats():
    with _db() as conn:
        rows = [dict(r) for r in conn.execute(
            'SELECT id, title, model, created_at, updated_at, parent_id, branched_from_message_id FROM conversations ORDER BY updated_at DESC LIMIT 250'
        )]
    return jsonify({'chats': rows})


@app.post('/api/chats')
def create_chat():
    data = request.get_json(silent=True) or {}
    ident = uuid.uuid4().hex
    now = _utc_now()
    title = (data.get('title') or 'New chat').strip()[:120] or 'New chat'
    model = (data.get('model') or '').strip() or None
    with _db() as conn:
        conn.execute('INSERT INTO conversations(id,title,model,created_at,updated_at) VALUES(?,?,?,?,?)',
                     (ident, title, model, now, now))
    return jsonify({'id': ident, 'title': title, 'model': model, 'created_at': now, 'updated_at': now})


@app.get('/api/chats/<chat_id>')
def get_chat(chat_id):
    with _db() as conn:
        row = conn.execute('SELECT id,title,model,created_at,updated_at,parent_id,branched_from_message_id FROM conversations WHERE id=?', (chat_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Chat not found'}), 404
        raw = conn.execute(
            'SELECT id,role,content,thinking,attachments_json,tool_calls_json,output_tokens,eval_duration_ns,created_at FROM conversation_messages WHERE conversation_id=? ORDER BY id', (chat_id,)
        ).fetchall()
    messages = []
    for r in raw:
        item = dict(r)
        try:
            attachments = json.loads(item.pop('attachments_json') or '[]')
        except Exception:
            attachments = []
        # The browser only needs display metadata when reopening a chat; image
        # bytes and extracted document text remain server-side in SQLite.
        item['attachments'] = []
        for idx, a in enumerate(attachments if isinstance(attachments, list) else []):
            if not isinstance(a, dict):
                continue
            meta = {**{k: a.get(k) for k in ('name','mime','kind','size')}, 'context_chars': len(str(a.get('text') or ''))}
            if a.get('kind') == 'image':
                meta['preview_url'] = f'/api/chats/{chat_id}/messages/{item["id"]}/attachments/{idx}'
            item['attachments'].append(meta)
        try:
            item['tool_calls'] = json.loads(item.pop('tool_calls_json') or '[]')
        except Exception:
            item['tool_calls'] = []
        eval_duration_ns = int(item.pop('eval_duration_ns') or 0)
        output_tokens = int(item.get('output_tokens') or 0)
        item['tokens_per_second'] = output_tokens / (eval_duration_ns / 1_000_000_000) if eval_duration_ns > 0 else None
        messages.append(item)
    out = dict(row)
    out['messages'] = messages
    return jsonify(out)


@app.get('/api/chats/<chat_id>/messages/<int:message_id>/attachments/<int:attachment_index>')
def get_chat_image_attachment(chat_id, message_id, attachment_index):
    """Serve a saved image attachment inline for the Chat preview/lightbox."""
    with _db() as conn:
        row = conn.execute(
            'SELECT attachments_json FROM conversation_messages WHERE conversation_id=? AND id=?',
            (chat_id, message_id),
        ).fetchone()
    if not row:
        return jsonify({'error': 'Attachment not found'}), 404
    try:
        items = json.loads(row['attachments_json'] or '[]')
        item = items[attachment_index]
    except Exception:
        return jsonify({'error': 'Attachment not found'}), 404
    if not isinstance(item, dict) or item.get('kind') != 'image':
        return jsonify({'error': 'Attachment is not an image'}), 404
    mime = str(item.get('mime') or '').lower()
    if mime not in _IMAGE_MIMES:
        return jsonify({'error': 'Unsupported image type'}), 415
    try:
        raw = base64.b64decode(item.get('data') or '', validate=True)
    except Exception:
        return jsonify({'error': 'Image data is unavailable'}), 404
    response = Response(raw, mimetype=mime)
    response.headers['Content-Disposition'] = f'inline; filename="{secure_filename(str(item.get("name") or "image"))}"'
    response.headers['Cache-Control'] = 'private, max-age=3600'
    return response


@app.patch('/api/chats/<chat_id>')
def update_chat(chat_id):
    data = request.get_json(silent=True) or {}
    title = data.get('title')
    model = data.get('model')
    with _db() as conn:
        if not conn.execute('SELECT 1 FROM conversations WHERE id=?', (chat_id,)).fetchone():
            return jsonify({'error': 'Chat not found'}), 404
        if title is not None:
            conn.execute('UPDATE conversations SET title=?,updated_at=? WHERE id=?', ((str(title).strip()[:120] or 'New chat'), _utc_now(), chat_id))
        if model is not None:
            conn.execute('UPDATE conversations SET model=?,updated_at=? WHERE id=?', ((str(model).strip() or None), _utc_now(), chat_id))
    return jsonify({'success': True})


@app.delete('/api/chats/<chat_id>')
def delete_chat(chat_id):
    with _db() as conn:
        conn.execute('DELETE FROM conversation_messages WHERE conversation_id=?', (chat_id,))
        cur = conn.execute('DELETE FROM conversations WHERE id=?', (chat_id,))
    return jsonify({'success': bool(cur.rowcount)})


_TEXT_EXTS = {'.txt','.md','.markdown','.json','.jsonl','.csv','.tsv','.py','.js','.ts','.tsx','.jsx','.html','.css','.scss','.yaml','.yml','.toml','.ini','.cfg','.conf','.xml','.sh','.bash','.zsh','.ps1','.sql','.go','.rs','.java','.c','.cc','.cpp','.h','.hpp','.cs','.php','.rb','.swift','.kt','.log'}
_IMAGE_MIMES = {'image/png','image/jpeg','image/webp','image/gif'}
_MAX_ATTACHMENT_BYTES = 12 * 1024 * 1024
_MAX_ATTACHMENT_TOTAL = 30 * 1024 * 1024
_MAX_EXTRACTED_CHARS = 180_000


def _extract_attachment_text(name, mime, raw):
    ext = os.path.splitext(name or '')[1].lower()
    if mime == 'application/pdf' or ext == '.pdf':
        reader = PdfReader(io.BytesIO(raw))
        chunks = []
        for page in reader.pages[:120]:
            try: chunks.append(page.extract_text() or '')
            except Exception: pass
            if sum(map(len, chunks)) >= _MAX_EXTRACTED_CHARS:
                break
        return '\n\n'.join(chunks)[:_MAX_EXTRACTED_CHARS]
    if mime == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' or ext == '.docx':
        doc = Document(io.BytesIO(raw))
        return '\n'.join(p.text for p in doc.paragraphs)[:_MAX_EXTRACTED_CHARS]
    if ext in ('.xlsx','.xls'):
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines = []
        for ws in workbook.worksheets[:8]:
            lines.append(f'--- Sheet: {ws.title} ---')
            for row in ws.iter_rows(values_only=True):
                lines.append('\t'.join('' if v is None else str(v) for v in row))
                if sum(map(len, lines)) >= _MAX_EXTRACTED_CHARS:
                    break
            if sum(map(len, lines)) >= _MAX_EXTRACTED_CHARS:
                break
        return '\n'.join(lines)[:_MAX_EXTRACTED_CHARS]
    if mime.startswith('text/') or ext in _TEXT_EXTS or mime in ('application/json','application/xml','application/yaml'):
        return raw.decode('utf-8', errors='replace')[:_MAX_EXTRACTED_CHARS]
    raise ValueError(f'{name}: unsupported file type. Use an image, PDF, DOCX, XLSX, or text/code file.')


def _prepare_attachments(raw_attachments):
    if not isinstance(raw_attachments, list):
        return [], [], ''
    stored, images, contexts = [], [], []
    total = 0
    for item in raw_attachments[:6]:
        if not isinstance(item, dict):
            continue
        name = secure_filename(str(item.get('name') or 'attachment'))[:180] or 'attachment'
        mime = str(item.get('mime') or 'application/octet-stream').lower()
        encoded = item.get('data') or ''
        if not isinstance(encoded, str):
            continue
        # Browser sends raw base64 (not a data URL).
        try:
            raw = base64.b64decode(encoded, validate=True)
        except Exception:
            raise ValueError(f'{name}: invalid attachment data')
        if len(raw) > _MAX_ATTACHMENT_BYTES:
            raise ValueError(f'{name}: file is larger than 12 MB')
        total += len(raw)
        if total > _MAX_ATTACHMENT_TOTAL:
            raise ValueError('Attachments exceed the 30 MB per-message limit')
        if mime in _IMAGE_MIMES:
            images.append(encoded)
            stored.append({'name':name,'mime':mime,'kind':'image','size':len(raw),'data':encoded})
            continue
        text = _extract_attachment_text(name, mime, raw)
        stored.append({'name':name,'mime':mime,'kind':'document','size':len(raw),'text':text})
        contexts.append(f'\n\n<attached_file name="{name}">\n{text}\n</attached_file>')
    return stored, images, ''.join(contexts)


def _stored_to_ollama_message(role, content, attachments_json='[]', tool_calls_json='[]'):
    try: attachments = json.loads(attachments_json or '[]')
    except Exception: attachments = []
    images, contexts = [], []
    for a in attachments if isinstance(attachments, list) else []:
        if not isinstance(a, dict): continue
        if a.get('kind') == 'image' and a.get('data'):
            images.append(a['data'])
        elif a.get('kind') == 'document':
            text = a.get('text') or ''
            if not text and a.get('data'):
                try:
                    raw = base64.b64decode(a.get('data'), validate=True)
                    text = _extract_attachment_text(str(a.get('name') or 'file'), str(a.get('mime') or 'application/octet-stream'), raw)
                except Exception:
                    text = ''
            if text:
                contexts.append(f'\n\n<attached_file name="{a.get("name","file")}">\n{text}\n</attached_file>')
    msg = {'role': role, 'content': (content or '') + ''.join(contexts)}
    if images: msg['images'] = images
    if role == 'assistant':
        try: tool_calls = json.loads(tool_calls_json or '[]')
        except Exception: tool_calls = []
        if tool_calls: msg['tool_calls'] = tool_calls
    return msg


def _vision_profile(show_data, model_name=''):
    """Conservative vision capability report for Chat image attachments.

    Ollama's capabilities list is authoritative for whether the installed model
    can currently accept images.  Some custom GGUFs contain multimodal weights
    but are installed without their projector/mmproj; those must not be treated
    as usable vision models merely because their name looks multimodal.
    """
    show_data = show_data or {}
    caps = [str(x).strip().lower() for x in (show_data.get('capabilities') or [])]
    info = show_data.get('model_info') or {}
    details = show_data.get('details') or {}
    identity = ' '.join([
        str(model_name or ''),
        str(details.get('family') or ''),
        ' '.join(str(x) for x in (details.get('families') or [])),
        str(info.get('general.architecture') or ''),
        ' '.join(str(k) for k in info.keys()),
    ]).lower()
    supported = 'vision' in caps
    projector_markers = ('projector', 'mmproj', 'vision_encoder', 'vision.encoder', 'clip.', 'clip_')
    projector_evidence = any(any(m in str(k).lower() for m in projector_markers) for k in info.keys())
    looks_multimodal = any(x in identity for x in (
        'vision', 'qwen3vl', 'qwen2vl', 'qwen2.5vl', 'llava', 'minicpm-v',
        'moondream', 'bakllava', 'pixtral', 'gemma3', 'mllama', 'multimodal',
    )) or projector_evidence
    if supported:
        summary = 'Ollama reports vision support for this installed model.'
        hint = 'Images can be sent directly through Ollama Chat.'
    elif looks_multimodal:
        summary = 'This looks like a multimodal model, but Ollama does not report a usable vision capability.'
        hint = 'The installed GGUF may be missing its matching projector/mmproj. Use a vision-capable Ollama model or reinstall/import the model with its required projector.'
    else:
        summary = 'Ollama does not report vision support for this installed model.'
        hint = 'Select a vision-capable model before sending images. Split multimodal GGUFs also require their matching projector/mmproj.'
    return {
        'supported': supported,
        'capability_reported': supported,
        'projector_evidence': projector_evidence,
        'looks_multimodal': looks_multimodal,
        'summary': summary,
        'hint': hint,
    }


@app.get('/api/chat/model-capabilities')
def chat_model_capabilities_v1_10():
    model = str(request.args.get('model') or '').strip()
    if not model:
        return jsonify({'error': 'model is required'}), 400
    try:
        show = _safe_ollama_json('POST', '/api/show', json={'model': model, 'verbose': True}, timeout=30)
        thinking = _effective_thinking_profile(show, model)
        vision = _vision_profile(show, model)
        limits = _model_limits(show)
        parsed = _parse_parameters(show.get('parameters', ''))
        return jsonify({
            'thinking': thinking,
            'vision': vision,
            'context_max': (limits.get('num_ctx') or {}).get('max'),
            'configured_context': parsed.get('num_ctx'),
            'capabilities': show.get('capabilities') or [],
            'limits': limits,
        })
    except requests.HTTPError as exc:
        detail = exc.response.text if exc.response is not None else str(exc)
        return jsonify({'error': detail}), getattr(exc.response, 'status_code', 500)
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.post('/api/chat/stream')
def chat_stream_v1_9():
    data = request.get_json(silent=True) or {}
    model = (data.get('model') or '').strip()
    provider = 'ollama'
    provider_model = model
    if ':' in model and model.split(':', 1)[0] in ('ollama', 'openai', 'anthropic', 'google'):
        provider, provider_model = model.split(':', 1)
    if provider == 'ollama' and model.startswith('ollama:'):
        provider_model = model[7:]
    prompt = (data.get('message') or '').strip()
    chat_id = (data.get('conversation_id') or '').strip() or None
    temporary = bool(data.get('temporary', not bool(chat_id)))
    think = data.get('think', None)
    if isinstance(think, str):
        think_l = think.strip().lower()
        if think_l in ('true','on'): think = True
        elif think_l in ('false','off'): think = False
        elif think_l in ('low','medium','high','xhigh','max'): think = think_l
        elif think_l in ('','auto'): think = None
        else: return jsonify({'error':'think must be auto, on, off, low, medium, high, xhigh, or max'}), 400
    elif think is not None and not isinstance(think, bool):
        think = None
    try:
        stored_attachments, current_images, attachment_context = _prepare_attachments(data.get('attachments') or [])
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    if not model:
        return jsonify({'error': 'Select a model first'}), 400
    thinking_profile = None
    thinking_system_instruction = None
    if provider != 'ollama':
        think = None
    if think is not None:
        try:
            thinking_profile = _thinking_profile_for_model(provider_model)
        except Exception as exc:
            return jsonify({'error': f'Could not verify thinking controls for {model}: {exc}'}), 502
        if not _think_value_allowed(thinking_profile, think):
            allowed = ', '.join(str(x.get('value')) for x in (thinking_profile.get('options') or [])) or 'none'
            return jsonify({'error': f'Thinking value is not supported by this model. Available controls: {allowed}'}), 400
        thinking_system_instruction = _thinking_system_instruction(thinking_profile, think)
        think = _thinking_wire_value(thinking_profile, think)
    if not prompt and not stored_attachments:
        return jsonify({'error': 'Message is empty'}), 400
    if not prompt:
        prompt = 'Please review the attached file(s).'
    if provider != 'ollama' and current_images:
        return jsonify({'error': 'External provider image input is not enabled yet. Remove the image or choose an Ollama vision model.'}), 400
    if provider != 'ollama':
        try:
            connection = provider_store.connection(provider)
            if provider_model not in connection['models']:
                return jsonify({'error': 'That external model is not enabled in Settings'}), 400
        except ProviderError as exc:
            return jsonify({'error': str(exc)}), 400

    messages = []
    created_chat = False
    now = _utc_now()
    attachments_json = json.dumps(stored_attachments, ensure_ascii=False)
    if chat_id and not temporary:
        with _db() as conn:
            chat = conn.execute('SELECT id,title,model FROM conversations WHERE id=?', (chat_id,)).fetchone()
            if not chat:
                return jsonify({'error': 'Chat not found'}), 404
            rows = conn.execute(
                'SELECT role,content,attachments_json,tool_calls_json FROM conversation_messages WHERE conversation_id=? ORDER BY id', (chat_id,)
            ).fetchall()
            messages = [_stored_to_ollama_message(r['role'], r['content'], r['attachments_json'], r['tool_calls_json']) for r in rows]
            conn.execute('INSERT INTO conversation_messages(conversation_id,role,content,thinking,attachments_json,tool_calls_json,created_at) VALUES(?,?,?,?,?,?,?)',
                         (chat_id, 'user', prompt, '', attachments_json, '[]', now))
            title = chat['title']
            if title == 'New chat' and not messages:
                title = _chat_title(prompt)
            conn.execute('UPDATE conversations SET title=?,model=?,updated_at=? WHERE id=?', (title, model, now, chat_id))
    elif not temporary:
        chat_id = uuid.uuid4().hex
        created_chat = True
        title = _chat_title(prompt)
        with _db() as conn:
            conn.execute('INSERT INTO conversations(id,title,model,created_at,updated_at) VALUES(?,?,?,?,?)', (chat_id,title,model,now,now))
            conn.execute('INSERT INTO conversation_messages(conversation_id,role,content,thinking,attachments_json,tool_calls_json,created_at) VALUES(?,?,?,?,?,?,?)',
                         (chat_id,'user',prompt,'',attachments_json,'[]',now))
    else:
        raw_messages = data.get('messages') or []
        if isinstance(raw_messages, list):
            for m in raw_messages[-80:]:
                if not isinstance(m, dict): continue
                role = str(m.get('role') or 'user')
                content = str(m.get('content') or '')
                try:
                    a_json = json.dumps(m.get('attachments') or [], ensure_ascii=False)
                except Exception:
                    a_json = '[]'
                try:
                    tc_json = json.dumps(m.get('tool_calls') or [], ensure_ascii=False)
                except Exception:
                    tc_json = '[]'
                messages.append(_stored_to_ollama_message(role, content, a_json, tc_json))

    if thinking_system_instruction:
        # Template-level controls such as Muse Glimmer's reasoning_strength are
        # represented as a system directive because Ollama's native /api/chat does
        # not currently expose arbitrary chat_template_kwargs.
        if messages and messages[0].get('role') == 'system':
            base = str(messages[0].get('content') or '')
            if thinking_system_instruction.lower() not in base.lower():
                messages[0]['content'] = (base.rstrip() + '\n' + thinking_system_instruction).strip()
        else:
            messages.insert(0, {'role': 'system', 'content': thinking_system_instruction})

    current_msg = {'role': 'user', 'content': prompt + attachment_context}
    if current_images:
        current_msg['images'] = current_images
    messages.append(current_msg)

    payload = {'model': provider_model, 'messages': messages, 'stream': True}
    if think is not None:
        payload['think'] = think
    external_live_id = _start_live_generation(model, '/api/chat', None) if provider != 'ollama' else None
    external_started = time.perf_counter()
    external_client_ip = request.headers.get('X-Forwarded-For', request.remote_addr)

    def generate():
        assistant_text, assistant_thinking, assistant_tool_calls = [], [], []
        final_output_tokens, final_eval_duration_ns = 0, 0
        external_final, external_error, external_status = {}, None, 500
        try:
            if chat_id:
                yield json.dumps({'control': {'conversation_id': chat_id, 'created': created_chat, 'title': _chat_title(prompt) if created_chat else None}}) + '\n'
            if provider != 'ollama':
                for obj in provider_store.stream_chat(provider, provider_model, messages):
                    _update_live_generation(external_live_id, obj)
                    if obj.get('done'):
                        external_final.update(obj)
                        final_output_tokens = int(obj.get('eval_count') or 0)
                        final_eval_duration_ns = int(obj.get('eval_duration') or 0)
                    msg = obj.get('message') or {}
                    chunk = msg.get('content') or ''
                    thinking_chunk = msg.get('thinking') or ''
                    if chunk: assistant_text.append(chunk)
                    if thinking_chunk: assistant_thinking.append(thinking_chunk)
                    yield json.dumps(obj) + '\n'
                external_status = 200
                return
            upstream = requests.post(f'{OLLAMA_API}/api/chat', json=payload, stream=True, timeout=3600, headers={'X-Ollama-Control-Client':'WebUI Chat','User-Agent':'Aperyn-WebUI/1.27.11'})
            if not upstream.ok:
                detail = upstream.text or f'Ollama returned HTTP {upstream.status_code}'
                low_detail = detail.lower()
                if current_images and ('image input is not supported' in low_detail or 'mmproj' in low_detail or 'projector' in low_detail):
                    detail = ('This installed model cannot currently accept images through Ollama. '
                              'Its vision projector/mmproj is missing or Ollama does not expose vision for this model. '
                              'Select a vision-capable Ollama model, or reinstall the multimodal model with its matching projector.')
                yield json.dumps({'error': detail}) + '\n'
                return
            for line in upstream.iter_lines():
                if not line:
                    continue
                try: obj = json.loads(line)
                except Exception: continue
                if obj.get('done'):
                    final_output_tokens = int(obj.get('eval_count') or 0)
                    final_eval_duration_ns = int(obj.get('eval_duration') or 0)
                msg = obj.get('message') or {}
                if isinstance(msg, dict):
                    chunk = msg.get('content') or ''
                    thinking_chunk = msg.get('thinking') or ''
                    if chunk: assistant_text.append(chunk)
                    if thinking_chunk: assistant_thinking.append(thinking_chunk)
                    calls = msg.get('tool_calls') or []
                    if isinstance(calls, list) and calls:
                        assistant_tool_calls.extend(calls)
                yield json.dumps(obj) + '\n'
        except Exception as exc:
            external_error = exc
            yield json.dumps({'error': str(exc)}) + '\n'
        finally:
            if external_live_id:
                latency_ms = (time.perf_counter() - external_started) * 1000
                try:
                    _log_request('/api/chat', model, external_status, latency_ms, external_final, external_client_ip,
                                 request_meta={'provider': provider, 'client_name': 'WebUI Chat'})
                finally:
                    _finish_live_generation(external_live_id, external_status, external_final, external_error)
            if chat_id and (assistant_text or assistant_thinking or assistant_tool_calls):
                text = ''.join(assistant_text)
                thinking_text = ''.join(assistant_thinking)
                with _db() as conn:
                    # De-duplicate identical streamed tool-call objects before persistence.
                    seen_calls=set(); stored_calls=[]
                    for call in assistant_tool_calls:
                        try: key=json.dumps(call,sort_keys=True,ensure_ascii=False)
                        except Exception: continue
                        if key in seen_calls: continue
                        seen_calls.add(key); stored_calls.append(call)
                    conn.execute('INSERT INTO conversation_messages(conversation_id,role,content,thinking,attachments_json,tool_calls_json,output_tokens,eval_duration_ns,created_at) VALUES(?,?,?,?,?,?,?,?,?)',
                                 (chat_id, 'assistant', text, thinking_text, '[]', json.dumps(stored_calls,ensure_ascii=False), final_output_tokens, final_eval_duration_ns, _utc_now()))
                    conn.execute('UPDATE conversations SET model=?,updated_at=? WHERE id=?', (model, _utc_now(), chat_id))
    return Response(generate(), mimetype='application/x-ndjson')


def _hf_headers():
    h = {'User-Agent': 'Ollama-Control/10'}
    if HF_TOKEN:
        h['Authorization'] = f'Bearer {HF_TOKEN}'
    return h


def _ollama_search(query, limit=16):
    response = requests.get('https://ollama.com/search', params={'q': query}, timeout=15, headers={'User-Agent':'Ollama-Control/10'})
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    found = []
    seen = set()
    for a in soup.find_all('a', href=True):
        href = a.get('href','')
        m = re.fullmatch(r'/library/([^/?#]+)', href)
        if not m:
            continue
        name = m.group(1)
        if name in seen:
            continue
        seen.add(name)
        parent = a.parent
        text = re.sub(r'\s+', ' ', (parent.get_text(' ', strip=True) if parent else a.get_text(' ', strip=True)))
        desc = text
        if desc.lower().startswith(name.lower()):
            desc = desc[len(name):].strip(' ·-')
        found.append({'source':'ollama','id':name,'name':name,'description':desc[:260],'url':f'https://ollama.com/library/{name}'})
        if len(found) >= limit:
            break
    return found


def _hf_search(query, limit=16):
    clean = (query or '').strip()
    url_match = re.match(r'https?://huggingface\.co/([^/]+/[^/?#]+)', clean, re.I)
    exact_repo = url_match.group(1).strip('/') if url_match else (clean if re.fullmatch(r'[^/\s]+/[^/\s]+', clean) else None)
    if exact_repo:
        response = requests.get(f'https://huggingface.co/api/models/{exact_repo}', timeout=8, headers=_hf_headers())
        if response.ok:
            item = response.json() or {}
            ident = item.get('id') or item.get('modelId') or exact_repo
            tags = item.get('tags') or []
            return [{'source':'huggingface','id':ident,'name':ident,'description':item.get('pipeline_tag') or 'GGUF repository','downloads':item.get('downloads') or 0,'likes':item.get('likes') or 0,'tags':tags[:8],'url':f'https://huggingface.co/{ident}'}]
    params = {'search': clean, 'filter': 'gguf', 'sort': 'downloads', 'direction': '-1', 'limit': limit, 'full': 'true'}
    response = requests.get('https://huggingface.co/api/models', params=params, timeout=8, headers=_hf_headers())
    response.raise_for_status()
    rows = []
    for item in response.json() or []:
        ident = item.get('id') or item.get('modelId')
        if not ident:
            continue
        tags = item.get('tags') or []
        rows.append({'source':'huggingface','id':ident,'name':ident,'description':item.get('pipeline_tag') or 'GGUF repository','downloads':item.get('downloads') or 0,'likes':item.get('likes') or 0,'tags':tags[:8],'url':f'https://huggingface.co/{ident}'})
    return rows



_POPULAR_OLLAMA_FALLBACK = [
    ('llama3.1','Meta Llama 3.1 family with 8B, 70B and 405B sizes.'),
    ('deepseek-r1','Open reasoning model family with multiple local sizes.'),
    ('llama3.2','Compact Meta Llama models for local use.'),
    ('gemma3','Google Gemma family with vision-capable variants.'),
    ('qwen2.5','Qwen multilingual general-purpose model family.'),
    ('qwen3','Qwen reasoning and tool-capable model family.'),
    ('mistral','Mistral general-purpose local model.'),
    ('gemma4','Gemma 4 multimodal and reasoning model family.'),
    ('qwen3.5','Qwen 3.5 multimodal, tools and thinking models.'),
    ('gpt-oss','Open-weight reasoning and agentic models.'),
    ('qwen3-coder','Qwen coding-focused long-context models.'),
    ('llava','LLaVA vision-language model family.'),
]
_POPULAR_HF_FALLBACK = [
    'unsloth/Qwen3.8-27B-GGUF',
    'unsloth/Muse-Glimmer-30B-GGUF',
    'bartowski/Qwen3.8-27B-GGUF',
    'unsloth/NVIDIA-Nemotron-3.5-Lightning-30B-A3B-GGUF',
    'unsloth/Qwen3.6-35B-A3B-GGUF',
    'unsloth/Qwen3.6-27B-MTP-GGUF',
    'LiquidAI/LFM2.5-2.6B-GGUF',
    'ornith-ai/Ornith-1.0-35B-GGUF',
]


def _fallback_ollama(limit=12):
    return [{'source':'ollama','id':n,'name':n,'description':d,'url':f'https://ollama.com/library/{n}','fallback':True} for n,d in _POPULAR_OLLAMA_FALLBACK[:limit]]


def _fallback_hf(limit=12):
    return [{'source':'huggingface','id':n,'name':n,'description':'Popular GGUF repository (cached fallback)','downloads':0,'likes':0,'tags':['gguf'],'url':f'https://huggingface.co/{n}','fallback':True} for n in _POPULAR_HF_FALLBACK[:limit]]

def _ollama_popular(limit=12):
    response = requests.get('https://ollama.com/library', params={'sort':'popular'}, timeout=8, headers={'User-Agent':'Ollama-Control/10'})
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    found, seen = [], set()
    for a in soup.find_all('a', href=True):
        href = a.get('href', '')
        m = re.fullmatch(r'/library/([^/?#]+)', href)
        if not m:
            continue
        name = m.group(1)
        if name in seen:
            continue
        seen.add(name)
        parent = a.parent
        text = re.sub(r'\s+', ' ', (parent.get_text(' ', strip=True) if parent else a.get_text(' ', strip=True)))
        desc = text
        if desc.lower().startswith(name.lower()):
            desc = desc[len(name):].strip(' ·-')
        found.append({'source':'ollama','id':name,'name':name,'description':desc[:260],'url':f'https://ollama.com/library/{name}'})
        if len(found) >= limit:
            break
    return found or _fallback_ollama(limit)


def _hf_popular(limit=12, trending=False):
    # Hugging Face exposes a live trending score as well as all-time downloads.
    params = {'filter':'gguf','sort':'trendingScore' if trending else 'downloads','direction':'-1','limit':limit,'full':'true'}
    response = requests.get('https://huggingface.co/api/models', params=params, timeout=8, headers=_hf_headers())
    response.raise_for_status()
    rows=[]
    for item in response.json() or []:
        ident=item.get('id') or item.get('modelId')
        if not ident:
            continue
        rows.append({'source':'huggingface','id':ident,'name':ident,'description':item.get('pipeline_tag') or 'GGUF repository','downloads':item.get('downloads') or 0,'likes':item.get('likes') or 0,'tags':(item.get('tags') or [])[:8],'url':f'https://huggingface.co/{ident}'})
    return rows or _fallback_hf(limit)


@app.get('/api/catalog/popular')
def catalog_popular():
    source = (request.args.get('source') or 'all').strip().lower()
    mode = (request.args.get('mode') or 'trending').strip().lower()
    try: limit = max(8, min(60, int(request.args.get('limit') or 24)))
    except (TypeError, ValueError): limit = 24
    trending = mode == 'trending'
    tasks = []
    if source in ('all','ollama'):
        # Ollama currently exposes popularity rather than a separate trending feed.
        tasks.append(('ollama','Popular on Ollama',lambda: _ollama_popular(limit),lambda: _fallback_ollama(limit)))
    if source in ('all','huggingface','hf'):
        title = 'Trending GGUF on Hugging Face' if trending else 'Popular GGUF on Hugging Face'
        tasks.append(('huggingface',title,lambda: _hf_popular(limit, trending=trending),lambda: _fallback_hf(limit)))
    sections, errors = [], []
    with ThreadPoolExecutor(max_workers=max(1,len(tasks))) as pool:
        futures = {pool.submit(fn): (src,title,fallback) for src,title,fn,fallback in tasks}
        for future in as_completed(futures):
            src,title,fallback = futures[future]
            try:
                rows = future.result()
                if not rows:
                    rows = fallback()
                sections.append({'source':src,'title':title,'results':rows,'fallback':bool(rows and rows[0].get('fallback'))})
            except Exception as exc:
                rows = fallback()
                sections.append({'source':src,'title':title,'results':rows,'fallback':True})
                errors.append(f'{title}: live catalogue unavailable; showing cached popular models ({exc})')
    order = {'ollama':0,'huggingface':1}
    sections.sort(key=lambda x: order.get(x['source'],9))
    response = jsonify({'sections':sections,'errors':errors,'mode':'trending' if trending else 'popular','limit':limit})
    response.headers['Cache-Control'] = 'private, max-age=180'
    return response


@app.get('/api/catalog/search')
def catalog_search():
    query = (request.args.get('q') or '').strip()
    source = (request.args.get('source') or 'all').strip().lower()
    if len(query) < 2:
        return jsonify({'results': [], 'message': 'Type at least 2 characters to search Ollama and Hugging Face.'})
    results, errors = [], []
    if source in ('all','ollama'):
        try: results.extend(_ollama_search(query))
        except Exception as exc: errors.append(f'Ollama catalogue: {exc}')
    if source in ('all','huggingface','hf'):
        try: results.extend(_hf_search(query))
        except Exception as exc: errors.append(f'Hugging Face: {exc}')
    return jsonify({'results': results, 'errors': errors})


def _parse_size_text(text):
    m = re.search(r'(?i)(\d+(?:\.\d+)?)\s*(KB|MB|GB|TB)', text or '')
    if not m: return 0
    scale = {'KB':1024,'MB':1024**2,'GB':1024**3,'TB':1024**4}[m.group(2).upper()]
    return int(float(m.group(1))*scale)


def _quant_from_name(name):
    upper = (name or '').upper()
    patterns = [r'(IQ\d(?:_[A-Z0-9]+)+)', r'(Q\d(?:_[A-Z0-9]+)+)', r'\b(FP16|F16|BF16|F32)\b']
    for pat in patterns:
        m = re.search(pat, upper)
        if m: return m.group(1)
    return 'default'


@app.get('/api/catalog/ollama/<path:model>/variants')
def catalog_ollama_variants(model):
    response = requests.get(f'https://ollama.com/library/{model}/tags', timeout=20, headers={'User-Agent':'Ollama-Control/10'})
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    text = soup.get_text('\n', strip=True)
    variants, seen = [], set()
    for match in re.finditer(rf'\b{re.escape(model)}:[A-Za-z0-9._-]+\b', text):
        tag = match.group(0)
        if tag in seen: continue
        seen.add(tag)
        around = text[match.start():match.start()+220]
        size = _parse_size_text(around)
        ctx = None
        cm = re.search(r'(?i)(\d+(?:\.\d+)?)\s*([KM])\s*context', around)
        if cm: ctx = int(float(cm.group(1))*(1000 if cm.group(2).upper()=='K' else 1_000_000))
        quant = _quant_from_name(tag)
        if quant == 'default':
            quant = _quant_from_name(around)
        variants.append({'source':'ollama','model':tag,'label':tag.split(':',1)[1],'quant':quant,'size':size,'context':ctx})
        if len(variants) >= 120: break
    if not variants:
        variants = [{'source':'ollama','model':f'{model}:latest','label':'latest','quant':'default','size':0,'context':None}]
    variants.sort(key=lambda x:(x['size'] or 10**18,x['label']))
    return jsonify({'model': model, 'variants': variants})


def _hf_file_digest(sibling):
    lfs = sibling.get('lfs') or {}
    value = lfs.get('sha256') or lfs.get('oid') or sibling.get('blobId') or sibling.get('oid')
    if value and str(value).startswith('sha256:'): return str(value).split(':',1)[1]
    if value and re.fullmatch(r'[a-fA-F0-9]{64}', str(value)): return str(value).lower()
    return None


@app.get('/api/catalog/huggingface/<path:repo>/variants')
def catalog_hf_variants(repo):
    response = requests.get(f'https://huggingface.co/api/models/{repo}', params={'blobs':'true'}, timeout=25, headers=_hf_headers())
    if response.status_code == 401:
        return jsonify({'error':'This Hugging Face repository requires authentication. Set HF_TOKEN in .env and recreate the WebUI container.'}), 401
    response.raise_for_status()
    info = response.json() or {}
    groups = {}
    for sibling in info.get('siblings') or []:
        filename = sibling.get('rfilename') or sibling.get('path') or ''
        low = filename.lower()
        if not low.endswith('.gguf') or any(x in low for x in ('mmproj','projector','imatrix')):
            continue
        quant = _quant_from_name(filename)
        # Split GGUF shards with the same quant belong to one install choice.
        key = quant
        size = sibling.get('size') or (sibling.get('lfs') or {}).get('size') or 0
        rec = {'name':filename,'size':int(size or 0),'sha256':_hf_file_digest(sibling)}
        groups.setdefault(key, {'source':'huggingface','repo':repo,'quant':quant,'label':quant,'size':0,'files':[]})
        groups[key]['files'].append(rec)
        groups[key]['size'] += rec['size']
    variants = list(groups.values())
    for v in variants:
        v['files'].sort(key=lambda x:x['name'])
    variants.sort(key=lambda x:(x['size'] or 10**18,x['quant']))
    
    for v in variants:
        v['revision'] = info.get('sha') or ''
    return jsonify({'repo':repo,'variants':variants,'downloads':info.get('downloads') or 0,'likes':info.get('likes') or 0,'private':bool(info.get('private')),'gated':info.get('gated'),'revision':info.get('sha') or ''})


def _set_download_job(job_id, **changes):
    with _download_lock:
        job = _download_jobs.get(job_id)
        if job:
            job.update(changes)
            job['updated_at'] = _utc_now()


def _run_hf_install(job_id, repo, files, model_name):
    try:
        total = sum(int(f.get('size') or 0) for f in files)
        completed_base = 0
        ollama_files = {}
        for idx, f in enumerate(files, start=1):
            filename = f['name']
            size = int(f.get('size') or 0)
            digest = f.get('sha256')
            _set_download_job(job_id, phase='download', status=f'Downloading {filename}', file=filename, file_index=idx, file_count=len(files), completed=completed_base, total=total)
            url = f'https://huggingface.co/{repo}/resolve/main/{quote(filename, safe="/")}'
            if digest:
                blob_url = f'{OLLAMA_API}/api/blobs/sha256:{digest}'
                exists = requests.head(blob_url, timeout=20)
                if exists.status_code == 200:
                    ollama_files[filename] = f'sha256:{digest}'
                    completed_base += size
                    _set_download_job(job_id, completed=completed_base)
                    continue
                hf = requests.get(url, stream=True, timeout=120, headers=_hf_headers(), allow_redirects=True)
                hf.raise_for_status()
                if not size:
                    size = int(hf.headers.get('Content-Length') or 0)
                    total = max(total, completed_base + size)
                sent = 0
                def chunks():
                    nonlocal sent
                    for chunk in hf.iter_content(chunk_size=4*1024*1024):
                        if not chunk: continue
                        sent += len(chunk)
                        _set_download_job(job_id, completed=completed_base+sent, total=total, status=f'Streaming {filename} into Ollama')
                        yield chunk
                headers = {'Content-Type':'application/octet-stream'}
                if size: headers['Content-Length'] = str(size)
                pushed = requests.post(blob_url, data=chunks(), headers=headers, timeout=3600)
                hf.close()
                if pushed.status_code not in (200,201):
                    raise RuntimeError(pushed.text or f'Ollama blob upload failed ({pushed.status_code})')
                ollama_files[filename] = f'sha256:{digest}'
                completed_base += size or sent
            else:
                # Fallback for non-LFS GGUFs where the Hub did not expose the SHA256.
                fd, temp_path = tempfile.mkstemp(prefix='hf-', suffix='.gguf', dir='/data')
                os.close(fd)
                try:
                    h = hashlib.sha256(); got = 0
                    hf = requests.get(url, stream=True, timeout=120, headers=_hf_headers(), allow_redirects=True); hf.raise_for_status()
                    with open(temp_path,'wb') as out:
                        for chunk in hf.iter_content(chunk_size=4*1024*1024):
                            if not chunk: continue
                            out.write(chunk); h.update(chunk); got += len(chunk)
                            _set_download_job(job_id, completed=completed_base+got, total=total or completed_base+got, status=f'Downloading {filename}')
                    digest = h.hexdigest(); size = got
                    blob_url = f'{OLLAMA_API}/api/blobs/sha256:{digest}'
                    if requests.head(blob_url, timeout=20).status_code != 200:
                        _set_download_job(job_id, phase='upload', status=f'Uploading {filename} to Ollama')
                        with open(temp_path,'rb') as source:
                            pushed = requests.post(blob_url, data=source, headers={'Content-Length':str(size),'Content-Type':'application/octet-stream'}, timeout=3600)
                        if pushed.status_code not in (200,201): raise RuntimeError(pushed.text or 'Ollama blob upload failed')
                    ollama_files[filename] = f'sha256:{digest}'
                    completed_base += size
                finally:
                    try: os.remove(temp_path)
                    except OSError: pass
        _set_download_job(job_id, phase='create', status=f'Creating Ollama model {model_name}', completed=total, total=total)
        created = requests.post(f'{OLLAMA_API}/api/create', json={'model':model_name,'files':ollama_files,'stream':False}, timeout=3600)
        if not created.ok:
            raise RuntimeError(created.text or f'Ollama create failed ({created.status_code})')
        _set_download_job(job_id, phase='done', status='Installed', done=True, success=True, completed=total, total=total, model=model_name)
    except Exception as exc:
        _set_download_job(job_id, phase='error', status='Failed', done=True, success=False, error=str(exc))


@app.post('/api/catalog/huggingface/install')
def catalog_hf_install():
    data = request.get_json(silent=True) or {}
    repo = (data.get('repo') or '').strip()
    files = data.get('files') or []
    model_name = (data.get('model_name') or '').strip().lower()
    if not repo or not files or not model_name:
        return jsonify({'error':'repo, files and model_name are required'}), 400
    if not re.fullmatch(r'[a-z0-9][a-z0-9._/-]*(?::[a-z0-9._-]+)?', model_name):
        return jsonify({'error':'Use a lowercase Ollama model name containing letters, numbers, ., _, -, / and an optional :tag.'}), 400
    total=sum(int(f.get('size') or 0) for f in files)
    job_id=_create_download_job('huggingface','huggingface',model_name,{'repo':repo,'files':files,'revision':data.get('revision') or '','quant':data.get('quant')},total)
    _ensure_download_worker()
    return jsonify({'job_id':job_id})


@app.get('/api/catalog/jobs/<job_id>')
def catalog_job(job_id):
    db_job = _download_job_row(job_id) if '_download_job_row' in globals() else None
    if db_job:
        return jsonify(db_job)
    with _download_lock:
        job = dict(_download_jobs.get(job_id) or {})
    if not job:
        return jsonify({'error':'Download job not found'}), 404
    return jsonify(job)


# --- v1.10 model operations, storage, downloads and inspection ---------------

def _normalize_digest(value):
    value = str(value or '').strip().lower()
    if value.startswith('sha256:'):
        value = value.split(':', 1)[1]
    return value


def _model_source(model):
    with _db() as conn:
        row = conn.execute('SELECT model,source,remote_id,revision,extra_json,updated_at FROM model_sources WHERE model=?', (model,)).fetchone()
    if not row:
        return None
    out = dict(row)
    try: out['extra'] = json.loads(out.pop('extra_json') or '{}')
    except Exception: out['extra'] = {}
    return out


def _record_model_source(model, source, remote_id='', revision='', extra=None):
    now = _utc_now()
    with _db() as conn:
        conn.execute("""INSERT INTO model_sources(model,source,remote_id,revision,extra_json,updated_at) VALUES(?,?,?,?,?,?)
                        ON CONFLICT(model) DO UPDATE SET source=excluded.source,remote_id=excluded.remote_id,
                        revision=excluded.revision,extra_json=excluded.extra_json,updated_at=excluded.updated_at""",
                     (model, source, remote_id or '', revision or '', json.dumps(extra or {}, ensure_ascii=False), now))


def _installed_models():
    data = _safe_ollama_json('GET', '/api/tags', timeout=12)
    return data.get('models') or []


def _installed_model_row(name):
    for row in _installed_models():
        if (row.get('name') or row.get('model')) == name:
            return row
    return None


def _info_number(info, suffixes):
    if not isinstance(info, dict):
        return None, None
    suffixes = tuple(str(s).lower() for s in suffixes)
    preferred = []
    rest = []
    arch = str(info.get('general.architecture') or '').lower()
    for key, value in info.items():
        key_l = str(key).lower()
        item = (key, value)
        if arch and key_l.startswith(arch + '.'):
            preferred.append(item)
        else:
            rest.append(item)
    for key, value in preferred + rest:
        key_l = str(key).lower()
        if any(key_l == s or key_l.endswith('.' + s) for s in suffixes):
            try:
                return float(value), str(key)
            except (TypeError, ValueError):
                continue
    return None, None


def _normalize_kv_cache_type(value, allow_auto=False):
    value = str(value or '').strip().lower().replace('-', '_')
    aliases = {'fp16':'f16','float16':'f16','q8':'q8_0','q4':'q4_0'}
    value = aliases.get(value, value)
    allowed = {'f16','q8_0','q4_0'} | ({'auto'} if allow_auto else set())
    return value if value in allowed else ('auto' if allow_auto else '')


def _model_memory_estimate(show_data, model_size, num_ctx=None, num_gpu=None, hardware=None, model_name=None, kv_cache_type=None):
    """Best-effort model + KV cache estimator.

    The KV estimate uses the standard attention K/V shape when the GGUF exposes
    embedding length, attention head counts and layer count. When metadata is
    incomplete, a conservative context reserve is used and labelled estimated.
    """
    show_data = show_data or {}
    info = show_data.get('model_info') or {}
    seed = _seed_hardware()
    service_env = seed.get('ollama_env') if isinstance(seed.get('ollama_env'), dict) else {}
    service_kv = _normalize_kv_cache_type(service_env.get('OLLAMA_KV_CACHE_TYPE'))
    global_kv = _normalize_kv_cache_type(_get_setting('estimator_kv_cache_type','auto'), allow_auto=True)
    model_kv = 'auto'
    if model_name:
        model_kv = _normalize_kv_cache_type(_get_setting(f'estimator_kv_cache_type:{model_name}','auto'), allow_auto=True)
    requested_kv = _normalize_kv_cache_type(kv_cache_type)
    if requested_kv:
        effective_kv, kv_source = requested_kv, 'request override'
    elif model_kv != 'auto':
        effective_kv, kv_source = model_kv, 'per-model estimator override'
    elif global_kv != 'auto':
        effective_kv, kv_source = global_kv, 'Settings estimator override'
    elif service_kv:
        effective_kv, kv_source = service_kv, 'detected Ollama service environment'
    else:
        effective_kv, kv_source = 'f16', 'Ollama default assumption'
    # Quantised llama.cpp KV blocks carry small per-block headers. Account for
    # those rather than treating q4/q8 as ideal half/one-byte scalars.
    kv_bpe = {'f16':2.0,'q8_0':34.0 / 32.0,'q4_0':18.0 / 32.0}[effective_kv]

    limits = _model_limits(show_data)
    ctx_max = (limits.get('num_ctx') or {}).get('max') or 4096
    try: ctx = int(num_ctx or 0) or int(ctx_max or 4096)
    except Exception: ctx = int(ctx_max or 4096)
    if ctx_max:
        ctx = max(512, min(ctx, int(ctx_max)))
    else:
        ctx = max(512, ctx)

    block_count = limits.get('block_count')
    emb, emb_key = _info_number(info, ('embedding_length',))
    heads, heads_key = _info_number(info, ('attention.head_count', 'head_count'))
    kv_heads, kv_key = _info_number(info, ('attention.head_count_kv', 'head_count_kv'))
    key_length, _ = _info_number(info, ('attention.key_length', 'key_length'))
    value_length, _ = _info_number(info, ('attention.value_length', 'value_length'))
    full_attention_interval, _ = _info_number(info, ('full_attention_interval',))
    if not block_count:
        bc, _ = _info_number(info, ('block_count',))
        block_count = int(bc) if bc else None
    if not kv_heads and heads:
        kv_heads = heads

    # Qwen hybrid/Gated DeltaNet models expose a full-attention interval. Only
    # those full-attention transformer blocks retain ordinary per-token K/V;
    # treating every DeltaNet layer as attention can overstate high-context
    # VRAM by several gigabytes. The linear recurrent state is bounded rather
    # than proportional to conversation length, so it belongs in overhead.
    kv_layers = int(block_count or 0)
    kv_attention_mode = 'all attention layers'
    if full_attention_interval and full_attention_interval > 1 and block_count:
        kv_layers = max(1, int(block_count) // int(full_attention_interval))
        kv_attention_mode = f'hybrid full-attention layers (1 in {int(full_attention_interval)})'

    kv_exactish = False
    kv_bytes = 0
    if kv_layers and emb and heads and kv_heads and heads > 0:
        head_dim = emb / heads
        # K + V, fp16-ish 2 bytes/element. Ollama may use other cache types;
        # this is deliberately surfaced as an estimate.
        key_dim = key_length or head_dim
        value_dim = value_length or head_dim
        kv_bytes = int(ctx * kv_layers * kv_heads * (key_dim + value_dim) * kv_bpe)
        kv_exactish = True
    else:
        # Fallback scales a modest reserve with context and model weight size.
        size_gb = max(0.25, float(model_size or 0) / (1024**3))
        kv_bytes = int((0.35 + size_gb * 0.018) * (ctx / 4096.0) * (1024**3) * (kv_bpe / 2.0))

    weights = int(model_size or 0)
    runtime_overhead = int(max(384 * 1024**2, weights * 0.055))
    gpu_max = (limits.get('num_gpu') or {}).get('max')
    try: gpu_layers = int(num_gpu) if num_gpu is not None else -1
    except Exception: gpu_layers = -1
    if gpu_max and gpu_layers >= 0:
        frac = min(1.0, max(0.0, gpu_layers / float(gpu_max)))
    else:
        frac = 1.0
    gpu_weights = int(weights * frac)
    cpu_weights = max(0, weights - gpu_weights)
    # KV usually follows the inference backend/GPU for accelerated models.
    gpu_kv = kv_bytes if frac > 0 else 0
    cpu_kv = kv_bytes if frac <= 0 else 0
    gpu_bytes = gpu_weights + gpu_kv + runtime_overhead
    base_ram_bytes = cpu_weights + cpu_kv + int(runtime_overhead * 0.55)

    hardware = hardware or _hardware_snapshot()
    gpu_total = int((hardware.get('gpu') or {}).get('total_bytes') or 0)
    ram_total = int((hardware.get('system') or {}).get('total_bytes') or 0)
    try: gpu_override_gb = max(0.0, float(_get_setting('gpu_vram_override_gb','0') or 0))
    except Exception: gpu_override_gb = 0.0
    try: ram_override_gb = max(0.0, float(_get_setting('system_ram_override_gb','0') or 0))
    except Exception: ram_override_gb = 0.0
    if gpu_override_gb > 0: gpu_total = int(gpu_override_gb * (1024**3))
    if ram_override_gb > 0: ram_total = int(ram_override_gb * (1024**3))
    gpu_budget = int(gpu_total * .94) if gpu_total else 0
    spill_bytes = max(0, gpu_bytes - gpu_budget) if gpu_budget else 0
    ram_bytes = base_ram_bytes + spill_bytes
    recommended_gpu_layers = None
    if gpu_total and gpu_max and weights > 0:
        # Reserve KV + runtime first, then estimate how many weight layers fit.
        weight_budget = max(0, gpu_budget - gpu_kv - runtime_overhead)
        recommended_gpu_layers = max(0, min(int(gpu_max), int(math.floor((weight_budget / weights) * gpu_max))))
    if gpu_total:
        if gpu_bytes <= gpu_budget:
            fit = 'full' if frac >= .999 else 'partial'
            fit_label = 'Full GPU offload' if frac >= .999 else 'Partial GPU offload'
        elif ram_total and ram_bytes <= ram_total * .82:
            fit, fit_label = 'partial', 'Partial offload required'
        else:
            fit, fit_label = 'no', 'Likely OOM / not recommended'
    else:
        fit, fit_label = 'unknown', 'VRAM capacity unknown'
    return {
        'context_tokens': ctx,
        'context_max': int(ctx_max) if ctx_max else None,
        'model_weights_bytes': weights,
        'kv_cache_bytes': kv_bytes,
        'kv_metadata_based': kv_exactish,
        'kv_layer_count': kv_layers or None,
        'kv_attention_mode': kv_attention_mode,
        'kv_cache_type': effective_kv,
        'kv_bytes_per_element': kv_bpe,
        'kv_cache_source': kv_source,
        'runtime_overhead_bytes': runtime_overhead,
        'estimated_gpu_bytes': gpu_bytes,
        'estimated_ram_bytes': ram_bytes,
        'estimated_spill_bytes': spill_bytes,
        'recommended_gpu_layers': recommended_gpu_layers,
        'gpu_layers': gpu_layers,
        'gpu_layers_max': gpu_max,
        'offload_fraction': frac,
        'fit': fit,
        'fit_label': fit_label,
        'hardware_gpu_total_bytes': gpu_total,
        'hardware_ram_total_bytes': ram_total,
        'hardware_gpu_source': 'Settings override' if gpu_override_gb > 0 else (hardware.get('gpu') or {}).get('source'),
        'hardware_ram_source': 'Settings override' if ram_override_gb > 0 else (hardware.get('system') or {}).get('source'),
        'metadata_keys': {'embedding': emb_key, 'heads': heads_key, 'kv_heads': kv_key},
        'note': f'Memory values are estimates. KV cache is calculated as {effective_kv} ({kv_source}); flash attention, parallelism and backend fallbacks can change real usage.',
    }


def _model_inspector_data(model):
    show = _safe_ollama_json('POST', '/api/show', json={'model': model, 'verbose': True}, timeout=45)
    row = _installed_model_row(model) or {}
    runtime = []
    try: runtime = _safe_ollama_json('GET', '/api/ps', timeout=8).get('models') or []
    except Exception: pass
    running = next((r for r in runtime if (r.get('name') or r.get('model')) == model), None)
    parsed = _parse_parameters(show.get('parameters', ''))
    limits = _model_limits(show)
    context = parsed.get('num_ctx') or (running or {}).get('context_length') or (limits.get('num_ctx') or {}).get('max') or 4096
    gpu_layers = parsed.get('num_gpu', -1)
    memory = _model_memory_estimate(show, row.get('size') or 0, context, gpu_layers, model_name=model)
    with _db() as conn:
        usage = conn.execute("""SELECT COUNT(*) requests, COALESCE(SUM(prompt_tokens),0) prompt_tokens,
                            COALESCE(SUM(output_tokens),0) output_tokens, AVG(latency_ms) avg_latency_ms,
                            COALESCE(SUM(eval_duration_ns),0) eval_duration_ns, MAX(created_at) last_used
                            FROM request_log WHERE model=?""", (model,)).fetchone()
    usage = dict(usage) if usage else {}
    eval_s = (usage.get('eval_duration_ns') or 0) / 1e9
    usage['tokens_per_second'] = ((usage.get('output_tokens') or 0) / eval_s) if eval_s > 0 else None
    usage.pop('eval_duration_ns', None)
    return {
        'model': model, 'installed': row, 'show': {
            'details': show.get('details') or {}, 'model_info': show.get('model_info') or {},
            'capabilities': show.get('capabilities') or [], 'parameters': parsed,
            'template_present': bool(show.get('template')),
        },
        'limits': limits, 'memory': memory, 'running': running,
        'mtp': {'capability': _mtp_capability(show), 'preference': _effective_mtp_pref(model, parsed)},
        'thinking': _effective_thinking_profile(show, model), 'source': _model_source(model), 'usage': usage,
    }


def _client_name_from_request_meta(meta):
    explicit = str((meta or {}).get('explicit_client') or '').strip()
    if explicit:
        return explicit[:120]
    ua = str((meta or {}).get('user_agent') or '').lower()
    if 'opencode' in ua: return 'OpenCode'
    if 'aider' in ua: return 'Aider'
    if 'continue' in ua: return 'Continue'
    if 'ollama-control-webui' in ua or 'webui chat' in ua: return 'WebUI Chat'
    if 'python-requests' in ua: return 'Python client'
    if 'curl/' in ua: return 'curl'
    return str((meta or {}).get('user_agent') or 'Unknown client')[:80]


def _request_meta_from_payload(payload=None):
    payload = payload if isinstance(payload, dict) else {}
    options = payload.get('options') if isinstance(payload.get('options'), dict) else {}
    safe_options = {str(k): v for k, v in options.items() if isinstance(v, (str,int,float,bool,type(None)))}
    meta = {
        'user_agent': request.headers.get('User-Agent', ''),
        'explicit_client': request.headers.get('X-Ollama-Control-Client', ''),
        'stream': payload.get('stream'), 'keep_alive': payload.get('keep_alive'),
        'think': payload.get('think'), 'format': ('json-schema' if isinstance(payload.get('format'), dict) else payload.get('format')),
        'messages_count': len(payload.get('messages') or []) if isinstance(payload.get('messages'), list) else None,
        'tools_count': len(payload.get('tools') or []) if isinstance(payload.get('tools'), list) else 0,
        'options': safe_options,
    }
    meta['client_name'] = _client_name_from_request_meta(meta)
    return meta


def _download_job_row(job_id):
    with _db() as conn:
        row = conn.execute('SELECT * FROM download_jobs WHERE id=?', (job_id,)).fetchone()
    if not row: return None
    out = dict(row)
    try: out['payload'] = json.loads(out.pop('payload_json') or '{}')
    except Exception: out['payload'] = {}
    out['done'] = bool(out['done']); out['success'] = None if out['success'] is None else bool(out['success']); out['cancel_requested'] = bool(out['cancel_requested'])
    return out


def _create_download_job(kind, source, target_model, payload, total=0):
    ident = uuid.uuid4().hex[:16]
    now = _utc_now()
    with _db() as conn:
        conn.execute("""INSERT INTO download_jobs(id,kind,source,target_model,payload_json,phase,status,completed,total,done,success,error,cancel_requested,created_at,updated_at)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                     (ident, kind, source, target_model, json.dumps(payload, ensure_ascii=False), 'queued', 'Queued', 0, int(total or 0), 0, None, '', 0, now, now))
    return ident


def _update_download_db(job_id, **changes):
    allowed = {'phase','status','completed','total','done','success','error','cancel_requested','updated_at'}
    values = {k:v for k,v in changes.items() if k in allowed}
    values['updated_at'] = _utc_now()
    if not values: return
    fields = ','.join(f'{k}=?' for k in values)
    with _db() as conn:
        conn.execute(f'UPDATE download_jobs SET {fields} WHERE id=?', tuple(values.values()) + (job_id,))


def _download_cancelled(job_id):
    row = _download_job_row(job_id)
    return bool(row and row.get('cancel_requested'))


def _run_ollama_queue_job(job):
    model = (job.get('payload') or {}).get('model') or job['target_model']
    _update_download_db(job['id'], phase='pull', status=f'Pulling {model}')
    resp = requests.post(f'{OLLAMA_API}/api/pull', json={'model': model, 'stream': True}, stream=True, timeout=3600)
    if not resp.ok: raise RuntimeError(resp.text or f'Ollama pull failed ({resp.status_code})')
    completed = total = 0
    for line in resp.iter_lines():
        if _download_cancelled(job['id']):
            resp.close(); raise InterruptedError('Cancelled by user')
        if not line: continue
        try: obj = json.loads(line)
        except Exception: continue
        if obj.get('error'): raise RuntimeError(obj['error'])
        completed = int(obj.get('completed') or completed or 0); total = int(obj.get('total') or total or 0)
        _update_download_db(job['id'], status=obj.get('status') or 'Pulling…', completed=completed, total=total)
    _record_model_source(job['target_model'], 'ollama', model, '', {'managed_by':'download_queue'})


def _run_hf_queue_job(job):
    payload = job.get('payload') or {}
    repo, files, model_name = payload.get('repo'), payload.get('files') or [], job['target_model']
    revision = payload.get('revision') or ''
    total = sum(int(f.get('size') or 0) for f in files)
    completed_base = 0; ollama_files = {}
    for idx, f in enumerate(files, start=1):
        if _download_cancelled(job['id']): raise InterruptedError('Cancelled by user')
        filename = f['name']; size = int(f.get('size') or 0); digest = f.get('sha256')
        _update_download_db(job['id'], phase='download', status=f'Downloading {filename} ({idx}/{len(files)})', completed=completed_base, total=total)
        url = f'https://huggingface.co/{repo}/resolve/{quote(revision or "main", safe="")}/{quote(filename, safe="/")}'
        if digest:
            blob_url = f'{OLLAMA_API}/api/blobs/sha256:{digest}'
            if requests.head(blob_url, timeout=20).status_code == 200:
                ollama_files[filename] = f'sha256:{digest}'; completed_base += size; continue
            hf = requests.get(url, stream=True, timeout=120, headers=_hf_headers(), allow_redirects=True); hf.raise_for_status()
            if not size:
                size = int(hf.headers.get('Content-Length') or 0); total = max(total, completed_base + size)
            sent = 0
            def chunks():
                nonlocal sent
                for chunk in hf.iter_content(chunk_size=4*1024*1024):
                    if _download_cancelled(job['id']): raise InterruptedError('Cancelled by user')
                    if not chunk: continue
                    sent += len(chunk); _update_download_db(job['id'], completed=completed_base+sent, total=total, status=f'Streaming {filename} into Ollama'); yield chunk
            headers={'Content-Type':'application/octet-stream'}
            if size: headers['Content-Length']=str(size)
            pushed=requests.post(blob_url,data=chunks(),headers=headers,timeout=3600); hf.close()
            if pushed.status_code not in (200,201): raise RuntimeError(pushed.text or f'Ollama blob upload failed ({pushed.status_code})')
            ollama_files[filename]=f'sha256:{digest}'; completed_base += size or sent
        else:
            fd,temp_path=tempfile.mkstemp(prefix='hf-',suffix='.gguf',dir='/data'); os.close(fd)
            try:
                h=hashlib.sha256(); got=0; hf=requests.get(url,stream=True,timeout=120,headers=_hf_headers(),allow_redirects=True); hf.raise_for_status()
                with open(temp_path,'wb') as out:
                    for chunk in hf.iter_content(chunk_size=4*1024*1024):
                        if _download_cancelled(job['id']): raise InterruptedError('Cancelled by user')
                        if not chunk: continue
                        out.write(chunk); h.update(chunk); got += len(chunk); _update_download_db(job['id'],completed=completed_base+got,total=total or completed_base+got,status=f'Downloading {filename}')
                digest=h.hexdigest(); size=got; blob_url=f'{OLLAMA_API}/api/blobs/sha256:{digest}'
                if requests.head(blob_url,timeout=20).status_code != 200:
                    _update_download_db(job['id'],phase='upload',status=f'Uploading {filename} to Ollama')
                    with open(temp_path,'rb') as source:
                        pushed=requests.post(blob_url,data=source,headers={'Content-Length':str(size),'Content-Type':'application/octet-stream'},timeout=3600)
                    if pushed.status_code not in (200,201): raise RuntimeError(pushed.text or 'Ollama blob upload failed')
                ollama_files[filename]=f'sha256:{digest}'; completed_base += size
            finally:
                try: os.remove(temp_path)
                except OSError: pass
    _update_download_db(job['id'],phase='create',status=f'Creating Ollama model {model_name}',completed=total,total=total)
    created=requests.post(f'{OLLAMA_API}/api/create',json={'model':model_name,'files':ollama_files,'stream':False},timeout=3600)
    if not created.ok: raise RuntimeError(created.text or f'Ollama create failed ({created.status_code})')
    _record_model_source(model_name,'huggingface',repo,revision,{'files':files,'quant':payload.get('quant')})


_DOWNLOAD_WORKER_LOCK = threading.Lock()
_DOWNLOAD_WORKER_STARTED = False

def _download_worker_loop():
    while True:
        try:
            with _db() as conn:
                row = conn.execute("SELECT * FROM download_jobs WHERE done=0 AND phase='queued' ORDER BY created_at LIMIT 1").fetchone()
            if not row:
                time.sleep(.8); continue
            job = dict(row)
            try: job['payload'] = json.loads(job.pop('payload_json') or '{}')
            except Exception: job['payload'] = {}
            if job.get('cancel_requested'):
                _update_download_db(job['id'],phase='cancelled',status='Cancelled',done=1,success=0,error='Cancelled by user'); continue
            _update_download_db(job['id'],phase='starting',status='Starting download')
            try:
                if job['kind']=='ollama': _run_ollama_queue_job(job)
                elif job['kind']=='huggingface': _run_hf_queue_job(job)
                else: raise RuntimeError(f'Unknown download type: {job["kind"]}')
                _update_download_db(job['id'],phase='done',status='Installed',done=1,success=1)
            except InterruptedError as exc:
                _update_download_db(job['id'],phase='cancelled',status='Cancelled',done=1,success=0,error=str(exc))
            except Exception as exc:
                _update_download_db(job['id'],phase='error',status='Failed',done=1,success=0,error=str(exc))
        except Exception:
            time.sleep(1.5)


def _ensure_download_worker():
    global _DOWNLOAD_WORKER_STARTED
    if os.environ.get('OLLAMA_CONTROL_ROLE','web') != 'web': return
    with _DOWNLOAD_WORKER_LOCK:
        if _DOWNLOAD_WORKER_STARTED: return
        # A container restart can interrupt an in-flight transfer. Put unfinished
        # jobs back in the persistent queue so they resume automatically.
        with _db() as conn:
            conn.execute("UPDATE download_jobs SET phase='queued', status='Resuming after restart', cancel_requested=0 WHERE done=0 AND phase!='queued'")
        _DOWNLOAD_WORKER_STARTED=True
        threading.Thread(target=_download_worker_loop,name='download-worker',daemon=True).start()


@app.get('/model/<path:model>')
def model_inspector_page(model):
    return render_template('model.html', model=model)


@app.get('/storage')
def storage_page():
    return render_template('storage.html')


@app.get('/downloads')
def downloads_page():
    return render_template('downloads.html')


@app.post('/api/manager/estimate')
def manager_estimate():
    data=request.get_json(silent=True) or {}; model=str(data.get('model') or '').strip()
    if not model: return jsonify({'error':'model is required'}),400
    show=_safe_ollama_json('POST','/api/show',json={'model':model,'verbose':True},timeout=45)
    row=_installed_model_row(model) or {}
    estimate=_model_memory_estimate(show,row.get('size') or 0,data.get('num_ctx'),data.get('num_gpu'),model_name=model,kv_cache_type=data.get('kv_cache_type'))
    return jsonify(estimate)


@app.route('/api/models/<path:model>/estimator-settings', methods=['GET','POST'])
def model_estimator_settings(model):
    model = str(model or '').strip()
    if not model:
        return jsonify({'error':'model is required'}), 400
    key = f'estimator_kv_cache_type:{model}'
    if request.method == 'POST':
        data = request.get_json(silent=True) or {}
        value = _normalize_kv_cache_type(data.get('estimator_kv_cache_type', data.get('kv_cache_type', 'auto')) or 'auto', allow_auto=True)
        if value not in ('auto','f16','q8_0','q4_0'):
            return jsonify({'error':'KV cache estimator override must be auto, f16, q8_0, or q4_0'}), 400
        _set_setting(key, value)
    override = _normalize_kv_cache_type(_get_setting(key,'auto'), allow_auto=True)
    seed = _seed_hardware()
    env = seed.get('ollama_env') if isinstance(seed.get('ollama_env'), dict) else {}
    service = _normalize_kv_cache_type(env.get('OLLAMA_KV_CACHE_TYPE'))
    global_value = _normalize_kv_cache_type(_get_setting('estimator_kv_cache_type','auto'), allow_auto=True)
    effective = override if override != 'auto' else (global_value if global_value != 'auto' else (service or 'f16'))
    return jsonify({'model':model,'kv_cache_type':override,'estimator_kv_cache_type':override,'global_kv_cache_type':global_value,'service_kv_cache_type':service or None,'detected_service_kv_cache_type':service or None,'effective_kv_cache_type':effective,'effective_estimator_kv_cache_type':effective})


@app.get('/api/models/<path:model>/inspector')
def model_inspector_api(model):
    try: return jsonify(_model_inspector_data(model))
    except Exception as exc: return jsonify({'error':str(exc)}),500


@app.get('/api/storage')
def storage_api():
    try: models=_installed_models()
    except Exception as exc: return jsonify({'error':str(exc)}),500
    with _db() as conn:
        last={r['model']:r['last_used'] for r in conn.execute('SELECT model,MAX(created_at) last_used FROM request_log WHERE model IS NOT NULL GROUP BY model')}
    items=[]; total=0
    for m in models:
        name=m.get('name') or m.get('model'); size=int(m.get('size') or 0); total+=size
        items.append({**m,'name':name,'size':size,'last_used':last.get(name),'source':_model_source(name)})
    items.sort(key=lambda x:x.get('size') or 0,reverse=True)
    stat=os.statvfs('/data')
    return jsonify({'models':items,'model_bytes':total,'count':len(items),'data_volume':{'free_bytes':stat.f_bavail*stat.f_frsize,'total_bytes':stat.f_blocks*stat.f_frsize}})


def _ollama_remote_digest(remote_id):
    remote_id=str(remote_id or '').strip()
    # Custom registry/namespace names are not guessed. Managed pulls record their
    # exact remote id; ordinary library names use Ollama defaults.
    name=remote_id
    tag='latest'
    if ':' in name.rsplit('/',1)[-1]:
        name,tag=name.rsplit(':',1)
    if name.startswith('registry.ollama.ai/'):
        name=name[len('registry.ollama.ai/'):]
    if '/' not in name: path=f'library/{name}'
    else: path=name
    url=f'https://registry.ollama.ai/v2/{path}/manifests/{quote(tag,safe="")}'
    headers={'Accept':'application/vnd.docker.distribution.manifest.v2+json, application/vnd.oci.image.manifest.v1+json','User-Agent':'Ollama-Control/10'}
    r=requests.head(url,headers=headers,timeout=8,allow_redirects=True)
    if not r.ok:
        r=requests.get(url,headers=headers,timeout=8,stream=True); r.raise_for_status()
    digest=r.headers.get('ollama-content-digest') or r.headers.get('docker-content-digest') or ''
    if not digest:
        body=r.content if hasattr(r,'content') else b''
        if body: digest=hashlib.sha256(body).hexdigest()
    return _normalize_digest(digest)


def _check_model_update(model):
    row=_installed_model_row(model) or {}; local=_normalize_digest(row.get('digest'))
    source=_model_source(model)
    if source and source.get('source')=='huggingface':
        repo=source.get('remote_id'); rev=source.get('revision') or ''
        r=requests.get(f'https://huggingface.co/api/models/{repo}',timeout=8,headers=_hf_headers()); r.raise_for_status(); current=str((r.json() or {}).get('sha') or '')
        return {'model':model,'source':'huggingface','remote_id':repo,'installed_revision':rev,'remote_revision':current,'update_available':bool(rev and current and rev!=current),'can_update':True if repo else False,'note':'Hugging Face revision comparison'}
    remote=(source or {}).get('remote_id') or model
    # Local tuned aliases cannot be reliably mapped upstream unless source was recorded.
    if not source and (model.startswith('ollama-control-') or '/' in model and not model.startswith('library/')):
        return {'model':model,'source':'local','update_available':False,'can_update':False,'note':'No upstream source recorded for this local/tuned model.'}
    try: remote_digest=_ollama_remote_digest(remote)
    except Exception as exc: return {'model':model,'source':'ollama','remote_id':remote,'update_available':False,'can_update':False,'error':str(exc)}
    return {'model':model,'source':'ollama','remote_id':remote,'installed_digest':local,'remote_digest':remote_digest,'update_available':bool(local and remote_digest and local!=remote_digest),'can_update':True}


@app.get('/api/storage/updates')
def storage_updates():
    models=[m.get('name') or m.get('model') for m in _installed_models()]
    results=[]
    with ThreadPoolExecutor(max_workers=min(6,max(1,len(models)))) as pool:
        futs={pool.submit(_check_model_update,m):m for m in models}
        for fut in as_completed(futs):
            try: results.append(fut.result())
            except Exception as exc: results.append({'model':futs[fut],'update_available':False,'can_update':False,'error':str(exc)})
    return jsonify({'updates':sorted(results,key=lambda x:(not x.get('update_available'),x.get('model','')))})


@app.post('/api/storage/update')
def storage_update_model():
    data=request.get_json(silent=True) or {}; model=str(data.get('model') or '').strip()
    if not model: return jsonify({'error':'model is required'}),400
    source=_model_source(model)
    if source and source.get('source')=='huggingface':
        repo=source.get('remote_id'); extra=source.get('extra') or {}; quant=extra.get('quant')
        if not repo: return jsonify({'error':'Hugging Face source repository is not recorded for this model.'}),400
        response=requests.get(f'https://huggingface.co/api/models/{repo}',params={'blobs':'true'},timeout=20,headers=_hf_headers()); response.raise_for_status(); info=response.json() or {}
        groups={}
        for sibling in info.get('siblings') or []:
            filename=sibling.get('rfilename') or sibling.get('path') or ''
            if not filename.lower().endswith('.gguf') or any(x in filename.lower() for x in ('mmproj','projector','imatrix')): continue
            q=_quant_from_name(filename); size=sibling.get('size') or (sibling.get('lfs') or {}).get('size') or 0
            groups.setdefault(q,[]).append({'name':filename,'size':int(size or 0),'sha256':_hf_file_digest(sibling)})
        files=groups.get(quant) if quant else None
        if not files:
            return jsonify({'error':f'Could not find the previously installed quant ({quant or "unknown"}) in the current Hugging Face revision. Open Model Library to choose a quant manually.'}),409
        total=sum(int(f.get('size') or 0) for f in files); job_id=_create_download_job('huggingface','huggingface',model,{'repo':repo,'files':files,'revision':info.get('sha') or '','quant':quant},total)
    else:
        remote=(source or {}).get('remote_id') or model; job_id=_create_download_job('ollama','ollama',model,{'model':remote})
    _ensure_download_worker(); return jsonify({'job_id':job_id})


@app.post('/api/storage/delete')
def storage_delete_models():
    data=request.get_json(silent=True) or {}; models=data.get('models') or []
    if not isinstance(models,list) or not models: return jsonify({'error':'models list is required'}),400
    deleted=[]; errors=[]
    for model in models[:100]:
        try:
            r=requests.delete(f'{OLLAMA_API}/api/delete',json={'model':model},timeout=45)
            if not r.ok: raise RuntimeError(r.text or f'HTTP {r.status_code}')
            _delete_model_pref(model)
            _set_setting(_thinking_override_key(model), '')
            with _db() as conn: conn.execute('DELETE FROM model_sources WHERE model=?',(model,))
            deleted.append(model)
        except Exception as exc: errors.append({'model':model,'error':str(exc)})
    return jsonify({'deleted':deleted,'errors':errors})


@app.get('/api/manager/request/<int:req_id>')
def manager_request_detail(req_id):
    with _db() as conn: row=conn.execute('SELECT * FROM request_log WHERE id=?',(req_id,)).fetchone()
    if not row: return jsonify({'error':'Request not found'}),404
    out=dict(row)
    try: out['request_meta']=json.loads(out.pop('request_meta_json') or '{}')
    except Exception: out['request_meta']={}
    eval_s=(out.get('eval_duration_ns') or 0)/1e9
    out['tokens_per_second']=(out.get('output_tokens') or 0)/eval_s if eval_s>0 else None
    return jsonify(out)


@app.get('/api/downloads')
def downloads_api():
    _ensure_download_worker()
    with _db() as conn: rows=conn.execute('SELECT * FROM download_jobs ORDER BY created_at DESC LIMIT 250').fetchall()
    jobs=[]
    for row in rows:
        item=dict(row)
        try: item['payload']=json.loads(item.pop('payload_json') or '{}')
        except Exception: item['payload']={}
        item['done']=bool(item['done']); item['success']=None if item['success'] is None else bool(item['success']); item['cancel_requested']=bool(item['cancel_requested']); jobs.append(item)
    return jsonify({'jobs':jobs})


@app.delete('/api/downloads')
def downloads_clear_completed():
    """Clear queue history without interrupting queued or active transfers."""
    user = _current_user()
    if not user or user.get('role') != 'admin':
        return jsonify({'error': 'Administrator access is required'}), 403
    with _db() as conn:
        deleted = conn.execute('DELETE FROM download_jobs WHERE done=1').rowcount
    return jsonify({'ok': True, 'deleted': int(deleted or 0)})


@app.post('/api/downloads/ollama')
def downloads_ollama_enqueue():
    data=request.get_json(silent=True) or {}; model=str(data.get('model') or '').strip()
    if not model: return jsonify({'error':'model is required'}),400
    ident=_create_download_job('ollama','ollama',model,{'model':model})
    _ensure_download_worker(); return jsonify({'job_id':ident})


@app.post('/api/downloads/huggingface')
def downloads_hf_enqueue():
    data=request.get_json(silent=True) or {}; repo=str(data.get('repo') or '').strip(); files=data.get('files') or []; model=str(data.get('model_name') or '').strip().lower()
    if not repo or not files or not model: return jsonify({'error':'repo, files and model_name are required'}),400
    if not re.fullmatch(r'[a-z0-9][a-z0-9._/-]*(?::[a-z0-9._-]+)?',model): return jsonify({'error':'Invalid Ollama model name'}),400
    total=sum(int(f.get('size') or 0) for f in files)
    ident=_create_download_job('huggingface','huggingface',model,{'repo':repo,'files':files,'revision':data.get('revision') or '','quant':data.get('quant')},total)
    _ensure_download_worker(); return jsonify({'job_id':ident})


@app.post('/api/downloads/<job_id>/cancel')
def download_cancel(job_id):
    if not _download_job_row(job_id): return jsonify({'error':'Download job not found'}),404
    _update_download_db(job_id,cancel_requested=1,status='Cancelling…'); return jsonify({'success':True})


@app.post('/api/downloads/<job_id>/retry')
def download_retry(job_id):
    old=_download_job_row(job_id)
    if not old: return jsonify({'error':'Download job not found'}),404
    ident=_create_download_job(old['kind'],old['source'],old['target_model'],old.get('payload') or {},old.get('total') or 0)
    _ensure_download_worker(); return jsonify({'job_id':ident})


@app.post('/api/chats/<chat_id>/branch')
def branch_chat(chat_id):
    data=request.get_json(silent=True) or {}
    try: message_id=int(data.get('message_id'))
    except Exception: return jsonify({'error':'message_id is required'}),400
    with _db() as conn:
        chat=conn.execute('SELECT * FROM conversations WHERE id=?',(chat_id,)).fetchone()
        target=conn.execute('SELECT * FROM conversation_messages WHERE conversation_id=? AND id=?',(chat_id,message_id)).fetchone()
        if not chat or not target: return jsonify({'error':'Chat/message not found'}),404
        if target['role']!='user': return jsonify({'error':'Branches must start from a user message'}),400
        new_id=uuid.uuid4().hex; now=_utc_now(); title=(str(data.get('title') or chat['title']).strip()[:100] or 'Branched chat')
        if not title.endswith(' · branch'): title=(title[:108]+' · branch')
        conn.execute('INSERT INTO conversations(id,title,model,created_at,updated_at,parent_id,branched_from_message_id) VALUES(?,?,?,?,?,?,?)',(new_id,title,chat['model'],now,now,chat_id,message_id))
        rows=conn.execute('SELECT role,content,thinking,attachments_json,tool_calls_json,output_tokens,eval_duration_ns,created_at FROM conversation_messages WHERE conversation_id=? AND id<? ORDER BY id',(chat_id,message_id)).fetchall()
        for r in rows:
            conn.execute('INSERT INTO conversation_messages(conversation_id,role,content,thinking,attachments_json,tool_calls_json,output_tokens,eval_duration_ns,created_at) VALUES(?,?,?,?,?,?,?,?,?)',(new_id,r['role'],r['content'],r['thinking'],r['attachments_json'],r['tool_calls_json'],r['output_tokens'],r['eval_duration_ns'],r['created_at']))
    return jsonify({'id':new_id,'message':str(data.get('content') if data.get('content') is not None else target['content']),'parent_id':chat_id,'branched_from_message_id':message_id})


_ensure_download_worker()

@app.get('/sw.js')
def service_worker():
    response = send_file(Path(app.static_folder) / 'sw.js', mimetype='application/javascript')
    response.headers['Service-Worker-Allowed'] = '/'
    response.headers['Cache-Control'] = 'no-cache'
    return response


@app.route('/health')
def health():
    return jsonify({'status': 'healthy'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=MANAGER_LISTEN_PORT, debug=os.environ.get('FLASK_DEBUG', 'false').lower() == 'true')
